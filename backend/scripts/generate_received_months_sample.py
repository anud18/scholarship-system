"""
Generate the Excel template for 已領月份數 import.

Produces docs/samples/received-months-template.xlsx with a header row and
a few example data rows. Run from the repo root:

    docker exec scholarship_backend_dev python backend/scripts/generate_received_months_sample.py

The output is checked into the repo so admins can download it directly;
regenerate if the column format ever changes.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).resolve().parents[2] / "docs" / "samples" / "received-months-template.xlsx"

HEADER = ("學號", "已領月份數")
SAMPLE_ROWS = (
    ("310551005", 12),
    ("312551183", 6),
    ("412551016", 0),
    ("412551012", 18),
    ("412551010", 24),
)


def build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "已領月份數"

    header_fill = PatternFill(start_color="FFE2E8F0", end_color="FFE2E8F0", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for col, value in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx, (sid, months) in enumerate(SAMPLE_ROWS, start=2):
        ws.cell(row=row_idx, column=1, value=sid).alignment = center
        ws.cell(row=row_idx, column=2, value=months).alignment = center

    ws.column_dimensions[get_column_letter(1)].width = 16
    ws.column_dimensions[get_column_letter(2)].width = 14
    ws.freeze_panes = "A2"

    return wb


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH.relative_to(OUTPUT_PATH.parents[2])}")


if __name__ == "__main__":
    main()
