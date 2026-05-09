"""Unit tests for CollegeRankingExportService cell rendering and column ordering.

These tests use plain dicts/dataclasses as input — no DB fixtures — so they
focus purely on the rendering logic.
"""

import io
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

import pytest
from openpyxl import load_workbook

from app.services.college_ranking_export_service import (
    CollegeRankingExportService,
    DynamicFieldSpec,
    ExportRow,
)


@dataclass
class FakeApplication:
    sub_type_preferences: Optional[List[str]] = None
    sub_scholarship_type: str = ""
    student_data: Dict[str, Any] = field(default_factory=dict)
    submitted_form_data: Dict[str, Any] = field(default_factory=dict)


def _make_row(
    rank_position: int = 1,
    app: Optional[FakeApplication] = None,
    bank_account: Optional[str] = None,
    advisor_names: Optional[str] = None,
) -> ExportRow:
    return ExportRow(
        rank_position=rank_position,
        application=app or FakeApplication(),
        bank_account=bank_account,
        advisor_names=advisor_names,
    )


def _full_student_data() -> Dict[str, Any]:
    return {
        "std_stdcode": "411511000",
        "std_cname": "王小明",
        "std_ename": "WANG, XIAO-MING",
        "std_pid": "A123456789",
        "std_nation": "中華民國",
        "std_sex": 1,
        "std_enrollyear": 110,
        "std_enrollterm": 1,
        "std_enrolltype": 1,  # 招生考試一般生 (not direct-track PhD)
        "trm_year": 114,
        "trm_term": 1,
        "trm_termcount": 5,
        "trm_academyname": "工學院",
        "trm_depname": "土木工程學系",
        "com_email": "test@nycu.edu.tw",
    }


def _build_workbook(
    rows: List[ExportRow],
    dynamic_fields: Optional[List[DynamicFieldSpec]] = None,
    sub_type_labels: Optional[Dict[str, str]] = None,
    title: str = "114學年度博士生獎學金學生資料彙整表",
    sheet_name: str = "114學年",
) -> List[List[Any]]:
    svc = CollegeRankingExportService()
    blob = svc.build_workbook(
        rows=rows,
        dynamic_fields=dynamic_fields or [],
        sub_type_labels=sub_type_labels or {},
        title=title,
        sheet_name=sheet_name,
    )
    wb = load_workbook(io.BytesIO(blob))
    ws = wb[sheet_name]
    # openpyxl 3.x suppresses empty-string cell values during XML
    # serialization (openpyxl.cell._writer early-returns on value == "",
    # emitting <c r="..."/> with no value subelement). On reload such
    # cells read back as None. The service writes "" for empty static and
    # dynamic fields; normalise None -> "" here so the per-test assertions
    # can compare against literal "" as written below.
    return [[("" if value is None else value) for value in row] for row in ws.iter_rows(values_only=True)]


class TestStaticColumns:
    def test_fully_populated_row(self):
        app = FakeApplication(
            sub_type_preferences=["nstc"],
            student_data=_full_student_data(),
        )
        rows = _build_workbook(
            [
                _make_row(
                    rank_position=2,
                    app=app,
                    bank_account="277506027171",
                    advisor_names="林教授、陳教授",
                )
            ],
            sub_type_labels={"nstc": "國科會"},
        )
        # rows[0] = title, rows[1] = headers, rows[2] = first data row
        data = rows[2]
        assert data[0] == 1  # NO. (1-based row index)
        assert data[1] == 2  # 學院初審會議之學院排序
        assert data[2] == "國科會"  # 申請獎學金類別 (sub_type_labels lookup happens externally)
        assert data[3] == "工學院"
        assert data[4] == "土木工程學系"
        assert data[5] == 3  # 年級 = ceil(5/2)
        assert data[6] == "否"  # 是否為逕博學生 (std_enrolltype=1)
        assert data[7] == "王小明"
        assert data[8] == "WANG, XIAO-MING"
        assert data[9] == "中華民國"
        assert data[10] == "男"
        assert data[11] == "110.9.1"
        assert data[12] == "411511000"
        assert data[13] == "A123456789"
        assert data[14] == "277506027171"  # 學生匯款帳號 (from user_profile.account_number)
        assert data[15] == "test@nycu.edu.tw"
        assert data[16] == "林教授、陳教授"  # 指導教授姓名 (joined when multiple)

    def test_header_row(self):
        rows = _build_workbook([])
        headers = rows[1]
        assert headers[:17] == [
            "NO.",
            "學院初審會議之學院排序",
            "申請獎學金類別",
            "學院",
            "系所",
            "年級",
            "是否為逕博學生",
            "學生中文姓名",
            "學生英文姓名",
            "國籍",
            "性別",
            "註冊入學日期",
            "學號",
            "學生身分證字號",
            "學生匯款帳號",
            "學生E-mail",
            "指導教授姓名",
        ]


class TestStaticFieldEdgeCases:
    @pytest.mark.parametrize(
        "sex,expected",
        [(1, "男"), (2, "女"), (0, ""), (None, ""), ("invalid", "")],
    )
    def test_gender_mapping(self, sex, expected):
        app = FakeApplication(student_data={"std_sex": sex})
        rows = _build_workbook([_make_row(app=app)])
        assert rows[2][10] == expected

    @pytest.mark.parametrize(
        "bank_account,expected",
        [
            ("277506027171", "277506027171"),
            ("12341234123412", "12341234123412"),
            (None, ""),
            ("", ""),
        ],
    )
    def test_bank_account_column(self, bank_account, expected):
        rows = _build_workbook([_make_row(bank_account=bank_account)])
        assert rows[2][14] == expected

    @pytest.mark.parametrize(
        "advisor_names,expected",
        [
            ("黃藍", "黃藍"),
            ("林教授、陳教授", "林教授、陳教授"),
            (None, ""),
            ("", ""),
        ],
    )
    def test_advisor_names_column(self, advisor_names, expected):
        rows = _build_workbook([_make_row(advisor_names=advisor_names)])
        assert rows[2][16] == expected

    @pytest.mark.parametrize(
        "enrolltype,expected",
        [
            (1, "否"),  # 招生考試一般生
            (4, "否"),  # 推甄一般生
            (8, "是"),  # 大學逕博
            (9, "是"),  # 碩士逕博
            (10, "是"),  # 跨校學士逕博
            (11, "是"),  # 跨校碩士逕博
            (12, "否"),  # 雙聯學位
            (29, "否"),  # TIGP
            (None, ""),
            ("invalid", ""),
        ],
    )
    def test_direct_phd_mapping(self, enrolltype, expected):
        app = FakeApplication(student_data={"std_enrolltype": enrolltype})
        rows = _build_workbook([_make_row(app=app)])
        assert rows[2][6] == expected

    @pytest.mark.parametrize(
        "termcount,expected",
        [(None, ""), (0, ""), (1, 1), (2, 1), (3, 2), (5, 3), (8, 4)],
    )
    def test_grade_formula(self, termcount, expected):
        app = FakeApplication(student_data={"trm_termcount": termcount})
        rows = _build_workbook([_make_row(app=app)])
        assert rows[2][5] == expected

    @pytest.mark.parametrize(
        "year,term,expected",
        [
            (110, 1, "110.9.1"),
            (110, 2, "110.2.1"),
            (114, 1, "114.9.1"),
            (None, 1, ""),
            (110, None, "110.9.1"),  # Default to first-term style
        ],
    )
    def test_enrollment_date(self, year, term, expected):
        app = FakeApplication(student_data={"std_enrollyear": year, "std_enrollterm": term})
        rows = _build_workbook([_make_row(app=app)])
        assert rows[2][11] == expected

    def test_missing_student_data_renders_empty_strings(self):
        app = FakeApplication(student_data={})
        rows = _build_workbook([_make_row(app=app)])
        data = rows[2]
        # Static text columns should be "" not None. Indices map to:
        # 3=學院, 4=系所, 6=年級, 7=是否為逕博, 8=中文姓名, 9=英文姓名,
        # 12=註冊入學日期, 13=學號, 14=身分證, 15=匯款帳號 (None via default),
        # 16=E-mail (col 17 = 指導教授, also defaults None -> "").
        for idx in [3, 4, 6, 7, 8, 9, 12, 13, 14, 15, 16]:
            assert data[idx] == "", f"col {idx} expected empty string, got {data[idx]!r}"


class TestSubTypeRendering:
    def test_zero_preferences_uses_fallback(self):
        app = FakeApplication(sub_type_preferences=[], sub_scholarship_type="moe_1w")
        rows = _build_workbook([_make_row(app=app)], sub_type_labels={"moe_1w": "教育部"})
        assert rows[2][2] == "教育部"

    def test_zero_preferences_unknown_code_uses_raw_code(self):
        app = FakeApplication(sub_type_preferences=None, sub_scholarship_type="custom_x")
        rows = _build_workbook([_make_row(app=app)])
        assert rows[2][2] == "custom_x"

    def test_one_preference(self):
        app = FakeApplication(sub_type_preferences=["nstc"])
        rows = _build_workbook([_make_row(app=app)], sub_type_labels={"nstc": "國科會"})
        assert rows[2][2] == "國科會"

    def test_two_preferences(self):
        app = FakeApplication(sub_type_preferences=["nstc", "moe_1w"])
        rows = _build_workbook(
            [_make_row(app=app)],
            sub_type_labels={"nstc": "國科會", "moe_1w": "教育部"},
        )
        assert rows[2][2] == "國科會(第一志願)暨教育部(第二志願)"

    def test_unknown_sub_type_in_preferences_uses_code(self):
        app = FakeApplication(sub_type_preferences=["nstc", "unknown_code"])
        rows = _build_workbook([_make_row(app=app)], sub_type_labels={"nstc": "國科會"})
        assert rows[2][2] == "國科會(第一志願)暨unknown_code(第二志願)"


class TestDynamicColumns:
    def test_dynamic_field_appended_after_static(self):
        app = FakeApplication(
            student_data=_full_student_data(),
            submitted_form_data={"fields": {"master_school": {"value": "台大土木系"}}},
        )
        fields = [
            DynamicFieldSpec(
                field_name="master_school",
                field_label="碩士畢業學校",
                export_column_label=None,
                display_order=10,
            )
        ]
        rows = _build_workbook([_make_row(app=app)], dynamic_fields=fields)
        assert rows[1][17] == "碩士畢業學校"
        assert rows[2][17] == "台大土木系"

    def test_export_column_label_overrides_field_label(self):
        app = FakeApplication(submitted_form_data={"fields": {"phone": {"value": "0912345678"}}})
        fields = [
            DynamicFieldSpec(
                field_name="phone",
                field_label="原始 label",
                export_column_label="學生手機",
                display_order=5,
            )
        ]
        rows = _build_workbook([_make_row(app=app)], dynamic_fields=fields)
        assert rows[1][17] == "學生手機"
        assert rows[2][17] == "0912345678"

    def test_dynamic_fields_sorted_by_display_order(self):
        app = FakeApplication(
            submitted_form_data={
                "fields": {
                    "a_first": {"value": "A"},
                    "b_second": {"value": "B"},
                }
            }
        )
        fields = [
            DynamicFieldSpec("b_second", "Second", None, display_order=20),
            DynamicFieldSpec("a_first", "First", None, display_order=10),
        ]
        rows = _build_workbook([_make_row(app=app)], dynamic_fields=fields)
        assert rows[1][17:19] == ["First", "Second"]
        assert rows[2][17:19] == ["A", "B"]

    def test_missing_dynamic_value_renders_empty(self):
        app = FakeApplication(submitted_form_data={"fields": {}})
        fields = [DynamicFieldSpec("phone", "手機", None, display_order=1)]
        rows = _build_workbook([_make_row(app=app)], dynamic_fields=fields)
        assert rows[2][17] == ""

    def test_none_value_renders_empty(self):
        app = FakeApplication(submitted_form_data={"fields": {"phone": {"value": None}}})
        fields = [DynamicFieldSpec("phone", "手機", None, display_order=1)]
        rows = _build_workbook([_make_row(app=app)], dynamic_fields=fields)
        assert rows[2][17] == ""

    def test_non_string_value_coerced_to_string(self):
        app = FakeApplication(submitted_form_data={"fields": {"count": {"value": 42}}})
        fields = [DynamicFieldSpec("count", "次數", None, display_order=1)]
        rows = _build_workbook([_make_row(app=app)], dynamic_fields=fields)
        assert rows[2][17] == "42"


class TestEmptyRanking:
    def test_empty_rows_produces_valid_workbook_with_headers(self):
        rows = _build_workbook([])
        assert rows[0][0].startswith("114學年度")  # title row
        assert rows[1][0] == "NO."  # header row
        assert len(rows) == 2  # title + header, no data


class TestTitleAndSheet:
    def test_workbook_title_row(self):
        rows = _build_workbook([], title="115學年度大學部獎學金學生資料彙整表")
        assert rows[0][0] == "115學年度大學部獎學金學生資料彙整表"

    def test_sheet_name_used(self):
        svc = CollegeRankingExportService()
        blob = svc.build_workbook(
            rows=[],
            dynamic_fields=[],
            sub_type_labels={},
            title="X",
            sheet_name="115學年",
        )
        wb = load_workbook(io.BytesIO(blob))
        assert "115學年" in wb.sheetnames
