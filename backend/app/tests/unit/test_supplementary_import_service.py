"""Unit tests for SupplementaryImportService pure helpers."""

import io
import pytest
from openpyxl import Workbook

from app.services.supplementary_import_service import (
    SupplementaryImportService,
    SupplementaryRow,
    parse_scholarship_type_cell,
)


def _make_excel(rows: list[list]) -> bytes:
    """Build a minimal 學生資料彙整表-format xlsx in memory."""
    wb = Workbook()
    ws = wb.active
    # Row 1: title (merged in real export, plain text here is fine for parsing)
    ws.cell(row=1, column=1, value="Title")
    # Row 2: static headers (18 columns)
    headers = [
        "NO.", "學院初審會議之學院排序", "申請獎學金類別", "學院", "系所",
        "年級", "是否為逕博學生", "學生中文姓名", "學生英文姓名", "國籍",
        "性別", "註冊入學日期", "學號", "學生身分證字號", "學生匯款帳號",
        "學生E-mail", "學生通訊地址", "指導教授姓名",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=h)
    # Data rows (row 3+)
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


LABEL_TO_CODE = {
    "國科會博士生研究獎學金": "nstc",
    "教育部博士生獎學金": "moe_1w",
}


class TestParseScholarshipTypeCell:
    def test_single_preference(self):
        prefs = parse_scholarship_type_cell("國科會博士生研究獎學金", LABEL_TO_CODE)
        assert prefs == ["nstc"]

    def test_dual_preference_first_nstc(self):
        cell = "國科會博士生研究獎學金(第一志願)暨教育部博士生獎學金(第二志願)"
        prefs = parse_scholarship_type_cell(cell, LABEL_TO_CODE)
        assert prefs == ["nstc", "moe_1w"]

    def test_dual_preference_first_moe(self):
        cell = "教育部博士生獎學金(第一志願)暨國科會博士生研究獎學金(第二志願)"
        prefs = parse_scholarship_type_cell(cell, LABEL_TO_CODE)
        assert prefs == ["moe_1w", "nstc"]

    def test_unknown_label_raises(self):
        with pytest.raises(ValueError, match="無法識別的獎學金類別"):
            parse_scholarship_type_cell("不存在的獎學金", LABEL_TO_CODE)

    def test_empty_cell_raises(self):
        with pytest.raises(ValueError, match="無法識別的獎學金類別"):
            parse_scholarship_type_cell("", LABEL_TO_CODE)


class TestParseExcel:
    def test_parses_student_id_rank_and_scholarship(self):
        row = [1, 1, "國科會博士生研究獎學金", "工學院", "電機系",
               2, "否", "王小明", "Wang", "台灣", "男", "113.9.1",
               "310460001", "A123456789", "12345678", "test@nycu.edu.tw",
               "新竹市", "指導教授A"]
        file_bytes = _make_excel([row])
        dynamic_field_names: list[str] = []
        rows, errors = SupplementaryImportService.parse_excel(
            file_bytes, LABEL_TO_CODE, dynamic_field_names
        )
        assert not errors
        assert len(rows) == 1
        r = rows[0]
        assert r.student_id == "310460001"
        assert r.excel_rank == 1
        assert r.sub_type_preferences == ["nstc"]
        assert r.bank_account == "12345678"
        assert r.advisor_name == "指導教授A"
        assert r.submitted_form_fields == {}

    def test_duplicate_student_ids_reported(self):
        row = [1, 1, "國科會博士生研究獎學金", "", "", 2, "", "王A", "", "", "", "",
               "310460001", "", "", "", "", ""]
        row2 = [2, 2, "國科會博士生研究獎學金", "", "", 2, "", "王B", "", "", "", "",
                "310460001", "", "", "", "", ""]
        file_bytes = _make_excel([row, row2])
        rows, errors = SupplementaryImportService.parse_excel(
            file_bytes, LABEL_TO_CODE, []
        )
        assert any("重複" in e for e in errors)

    def test_non_integer_rank_reported(self):
        row = [1, "abc", "國科會博士生研究獎學金", "", "", 2, "", "王A", "", "", "", "",
               "310460001", "", "", "", "", ""]
        file_bytes = _make_excel([row])
        rows, errors = SupplementaryImportService.parse_excel(
            file_bytes, LABEL_TO_CODE, []
        )
        assert any("排名" in e for e in errors)

    def test_skips_empty_rows(self):
        empty_row = [""] * 18
        real_row = [1, 1, "國科會博士生研究獎學金", "", "", 2, "", "王A", "", "", "", "",
                    "310460001", "", "", "", "", ""]
        file_bytes = _make_excel([empty_row, real_row])
        rows, errors = SupplementaryImportService.parse_excel(
            file_bytes, LABEL_TO_CODE, []
        )
        assert not errors
        assert len(rows) == 1
