"""Integration tests for the export-excel `template` flag (blank rank column)."""

import io

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import require_scholarship_manager
from app.main import app
from app.models.application import Application, ApplicationStatus
from app.models.audit_log import AuditAction, AuditLog
from app.models.college_review import CollegeRanking, CollegeRankingItem
from app.models.scholarship import (
    ScholarshipConfiguration,
    ScholarshipType,
    SubTypeSelectionMode,
)
from app.models.user import AdminScholarship, User, UserRole, UserType


@pytest_asyncio.fixture
async def admin_user(db: AsyncSession) -> User:
    user = User(
        nycu_id="admin900",
        name="Admin",
        email="admin900@nycu.edu.tw",
        user_type=UserType.employee,
        role=UserRole.admin,
    )
    db.add(user)
    await db.flush()
    return user


@pytest_asyncio.fixture
async def scholarship(db: AsyncSession) -> ScholarshipType:
    s = ScholarshipType(
        code="phd_tmpl_test",
        name="Test Template PhD",
        sub_type_selection_mode=SubTypeSelectionMode.single,
        status="active",
    )
    db.add(s)
    await db.flush()
    return s


@pytest_asyncio.fixture
async def configuration(db: AsyncSession, scholarship: ScholarshipType, admin_user: User) -> ScholarshipConfiguration:
    cfg = ScholarshipConfiguration(
        scholarship_type_id=scholarship.id,
        academic_year=114,
        semester=None,  # yearly
        config_name="Test PhD 114",
        config_code="test-tmpl-114",
        amount=40000,
        is_active=True,
    )
    db.add(cfg)
    # Grant admin permission so _check_scholarship_permission passes
    db.add(AdminScholarship(admin_id=admin_user.id, scholarship_id=scholarship.id))
    await db.flush()
    return cfg


@pytest_asyncio.fixture
async def ranking_with_item(
    db: AsyncSession,
    admin_user: User,
    scholarship: ScholarshipType,
    configuration: ScholarshipConfiguration,
) -> CollegeRanking:
    student = User(
        nycu_id="310460099",
        name="王小明",
        email="s99@nycu.edu.tw",
        user_type=UserType.student,
        role=UserRole.student,
    )
    db.add(student)
    await db.flush()

    r = CollegeRanking(
        scholarship_type_id=scholarship.id,
        sub_type_code="nstc",
        academic_year=114,
        ranking_name="Test",
        created_by=admin_user.id,
        is_finalized=False,
    )
    db.add(r)
    await db.flush()

    app_row = Application(
        app_id="APP-114-0-09999",
        user_id=student.id,
        scholarship_type_id=scholarship.id,
        academic_year=114,
        semester=None,
        status=ApplicationStatus.submitted,
        sub_type_selection_mode=SubTypeSelectionMode.single,
        student_data={"std_stdcode": "310460099", "std_cname": "王小明"},
        sub_type_preferences=["nstc"],
        scholarship_subtype_list=["nstc"],
        submitted_form_data={"fields": {}},
    )
    db.add(app_row)
    await db.flush()

    item = CollegeRankingItem(
        ranking_id=r.id,
        application_id=app_row.id,
        rank_position=1,
        status="ranked",
        college_rejected=False,
        is_allocated=False,
    )
    db.add(item)
    await db.flush()
    return r


@pytest.mark.asyncio
class TestExportTemplateFlag:
    async def test_template_true_blanks_rank_and_marks_filename(
        self,
        client: AsyncClient,
        db: AsyncSession,
        admin_user: User,
        ranking_with_item: CollegeRanking,
    ):
        app.dependency_overrides[require_scholarship_manager] = lambda: admin_user
        try:
            resp = await client.get(
                f"/api/v1/college-review/rankings/{ranking_with_item.id}/export-excel",
                params={"template": "true"},
            )
        finally:
            app.dependency_overrides.pop(require_scholarship_manager, None)

        assert resp.status_code == 200, resp.text
        # 範本 (U+7BC4 U+672C) url-encoded in Content-Disposition filename*
        assert "%E7%AF%84%E6%9C%AC" in resp.headers["content-disposition"]

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # row1 = title, row2 = headers, row3 = first data row
        assert ws.cell(row=2, column=2).value == "學院初審會議之學院排序"
        assert ws.cell(row=2, column=13).value == "學號"
        assert str(ws.cell(row=3, column=13).value) == "310460099"  # 學號 present
        assert (ws.cell(row=3, column=2).value or "") == ""  # rank BLANK

        # PII audit must be recorded even for the template: the 彙整表 still
        # contains plaintext std_pid, so the template export is logged like any
        # other, tagged is_template=True.
        result = await db.execute(
            select(AuditLog).where(
                AuditLog.action == AuditAction.pii_access.value,
                AuditLog.resource_type == "college_ranking",
                AuditLog.resource_id == str(ranking_with_item.id),
            )
        )
        audit = result.scalars().first()
        assert audit is not None, "template export must record a pii_access audit"
        assert audit.meta_data["is_template"] is True

    async def test_default_keeps_rank_filled(
        self,
        client: AsyncClient,
        db: AsyncSession,
        admin_user: User,
        ranking_with_item: CollegeRanking,
    ):
        app.dependency_overrides[require_scholarship_manager] = lambda: admin_user
        try:
            resp = await client.get(f"/api/v1/college-review/rankings/{ranking_with_item.id}/export-excel")
        finally:
            app.dependency_overrides.pop(require_scholarship_manager, None)

        assert resp.status_code == 200, resp.text
        assert "%E7%AF%84%E6%9C%AC" not in resp.headers["content-disposition"]
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.cell(row=3, column=2).value == 1  # rank present
