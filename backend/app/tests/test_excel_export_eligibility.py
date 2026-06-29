"""
Eligibility-verification enrichment of the roster (造冊) Excel.

Covers the feature added on 2026-06-30: the STD_UP_MIXLISTA payment sheet now
carries verification columns (學籍驗證 / 納入造冊 / 排除原因) plus ONE COLUMN PER
SCHOLARSHIP RULE, and paints the specific failing cells red (hard fail) or amber
(warning fail). See docs/superpowers/specs/2026-06-30-roster-excel-eligibility-
verification-design.md.

Design invariants pinned here:

- The per-rule pass/fail is NOT re-evaluated in the Excel layer. It is read from
  the frozen `item.rule_validation_result["details"]["rule_<id>"]` snapshot that
  RosterService writes at generation time. So these tests feed that snapshot
  shape directly (no DB, SimpleNamespace stubs like the sibling row-mapping test).
- Fill is a pure function of (column_name, rendered value, row) — a rule cell
  reading "未通過" is red, "警告" is amber, regardless of how it was produced.
- The review export shows ineligibility-excluded students (so they can be shown
  in red) but HIDES students manually removed after locking.
"""

from __future__ import annotations

import os
from types import SimpleNamespace
from typing import Any, Dict, List, Optional

import pytest
from openpyxl import load_workbook

from app.models.payment_roster import StudentVerificationStatus
from app.services.excel_export_service import ExcelExportService

RED = "FFC7CE"
AMBER = "FFEB9C"


@pytest.fixture
def service() -> ExcelExportService:
    return ExcelExportService()


def _rule_detail(rule_name: str, passed: bool, *, hard: bool = True, warning: bool = False) -> Dict[str, Any]:
    return {
        "passed": passed,
        "rule_name": rule_name,
        "rule_type": "gpa",
        "is_hard_rule": hard,
        "is_warning": warning,
        "message": f"{rule_name} 訊息",
    }


def _make_item(
    *,
    student_id_number: str = "A123456789",
    student_name: str = "王小明",
    bank_account: Optional[str] = "00012345678",
    is_included: bool = True,
    exclusion_reason: Optional[str] = None,
    verification_status: StudentVerificationStatus = StudentVerificationStatus.VERIFIED,
    rule_details: Optional[Dict[str, Dict[str, Any]]] = None,
) -> SimpleNamespace:
    rvr: Optional[Dict[str, Any]] = None
    if rule_details is not None:
        rvr = {"is_eligible": is_included, "failed_rules": [], "warning_rules": [], "details": rule_details}
    return SimpleNamespace(
        student_id_number=student_id_number,
        student_name=student_name,
        student_email="x@nycu.edu.tw",
        bank_account=bank_account,
        scholarship_name="博士班獎學金",
        scholarship_amount=50000,
        permanent_address=None,
        application_identity="114新申請",
        allocated_sub_type=None,
        allocation_year=None,
        is_included=is_included,
        exclusion_reason=exclusion_reason,
        verification_status=verification_status,
        rule_validation_result=rvr,
        excel_row_data=None,
        excel_remarks=None,
    )


def _make_roster(**kw: Any) -> SimpleNamespace:
    return SimpleNamespace(
        period_label=kw.get("period_label", "2025-H1"),
        roster_code=kw.get("roster_code", "ROSTER-114-2025-H1-PHD001"),
        academic_year=kw.get("academic_year", 114),
        items=kw.get("items", []),
    )


# ---------------------------------------------------------------------------
# _is_manual_removal — distinguishes intentional removals from ineligibility
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "reason,expected",
    [
        ("鎖定後移除[補充]：重複", True),
        ("比對分發移除：不在分發名單", True),
        ("缺少銀行帳戶資訊", False),
        ("學籍驗證未通過: graduated", False),
        ("不符合獎學金規則: GPA 不足", False),
        (None, False),
        ("", False),
    ],
)
def test_is_manual_removal(service: ExcelExportService, reason: Optional[str], expected: bool) -> None:
    item = _make_item(is_included=False, exclusion_reason=reason)
    assert service._is_manual_removal(item) is expected


# ---------------------------------------------------------------------------
# _get_roster_items — review scope
# ---------------------------------------------------------------------------


def test_get_roster_items_shows_ineligible_hides_manual_removal(service: ExcelExportService) -> None:
    included = _make_item(student_name="A_included")
    ineligible = _make_item(student_name="B_ineligible", is_included=False, exclusion_reason="缺少銀行帳戶資訊")
    manual = _make_item(student_name="C_manual", is_included=False, exclusion_reason="鎖定後移除：手動")
    roster = _make_roster(items=[included, ineligible, manual])

    names = [it.student_name for it in service._get_roster_items(roster, include_excluded=False)]

    assert "A_included" in names
    assert "B_ineligible" in names  # auto-excluded → shown (will be red)
    assert "C_manual" not in names  # manual removal → hidden


def test_get_roster_items_include_excluded_returns_everything(service: ExcelExportService) -> None:
    included = _make_item(student_name="A_included")
    manual = _make_item(student_name="C_manual", is_included=False, exclusion_reason="鎖定後移除：手動")
    roster = _make_roster(items=[included, manual])

    names = [it.student_name for it in service._get_roster_items(roster, include_excluded=True)]

    assert names == ["A_included", "C_manual"]  # sorted by name, manual removal included


# ---------------------------------------------------------------------------
# _collect_rule_columns — dynamic per-rule column set
# ---------------------------------------------------------------------------


def test_collect_rule_columns_orders_by_rule_id(service: ExcelExportService) -> None:
    item1 = _make_item(rule_details={"rule_5": _rule_detail("GPA門檻", True), "rule_2": _rule_detail("國籍", True)})
    item2 = _make_item(rule_details={"rule_2": _rule_detail("國籍", True), "rule_9": _rule_detail("年級", False)})

    cols = service._collect_rule_columns([item1, item2])

    assert [key for key, _ in cols] == ["rule_2", "rule_5", "rule_9"]
    assert [header for _, header in cols] == ["國籍", "GPA門檻", "年級"]


def test_collect_rule_columns_disambiguates_duplicate_names(service: ExcelExportService) -> None:
    item = _make_item(
        rule_details={"rule_2": _rule_detail("學業成績", True), "rule_7": _rule_detail("學業成績", False)}
    )

    cols = service._collect_rule_columns([item])

    headers = [header for _, header in cols]
    assert headers == ["學業成績 (2)", "學業成績 (7)"]


def test_collect_rule_columns_ignores_non_rule_detail_keys(service: ExcelExportService) -> None:
    no_rules = _make_item(rule_details={"no_rules_found": True})
    errored = _make_item(rule_details={"error": "boom"})
    none_item = _make_item(rule_details=None)

    assert service._collect_rule_columns([no_rules, errored, none_item]) == []


# ---------------------------------------------------------------------------
# _prepare_excel_data — verification + per-rule cell values
# ---------------------------------------------------------------------------


def test_prepare_adds_verification_columns_for_qualified(service: ExcelExportService) -> None:
    roster = _make_roster()
    item = _make_item(verification_status=StudentVerificationStatus.VERIFIED, is_included=True)

    row = service._prepare_excel_data(roster, [item])[0]

    assert row["學籍驗證"] == "已驗證"
    assert row["納入造冊"] == "是"
    assert row["排除原因"] == ""


def test_prepare_verification_columns_for_excluded(service: ExcelExportService) -> None:
    roster = _make_roster()
    item = _make_item(
        verification_status=StudentVerificationStatus.GRADUATED,
        is_included=False,
        exclusion_reason="學籍驗證未通過: graduated",
    )

    row = service._prepare_excel_data(roster, [item])[0]

    assert row["學籍驗證"] == "已畢業"
    assert row["納入造冊"] == "否"
    assert row["排除原因"] == "學籍驗證未通過: graduated"


def test_prepare_per_rule_cell_values(service: ExcelExportService) -> None:
    roster = _make_roster()
    item = _make_item(
        rule_details={
            "rule_2": _rule_detail("國籍", True),  # 通過
            "rule_5": _rule_detail("GPA門檻", False, hard=True),  # 未通過 (red)
            "rule_9": _rule_detail("年級", False, hard=False, warning=True),  # 警告 (amber)
        }
    )

    row = service._prepare_excel_data(roster, [item])[0]

    assert row["國籍"] == "通過"
    assert row["GPA門檻"] == "未通過"
    assert row["年級"] == "警告"


def test_prepare_per_rule_cell_dash_when_no_detail(service: ExcelExportService) -> None:
    """A rule column present on the sheet but with no detail for this item
    (rule not applicable to this student) renders '—', not blank/crash."""
    roster = _make_roster()
    has_rule = _make_item(student_name="A", rule_details={"rule_5": _rule_detail("GPA門檻", True)})
    no_rule = _make_item(student_name="B", rule_details={})

    rows = service._prepare_excel_data(roster, [has_rule, no_rule])

    # both rows expose the column; the one lacking detail shows —
    assert rows[0]["GPA門檻"] == "通過"
    assert rows[1]["GPA門檻"] == "—"


def test_prepare_sets_template_columns_dynamically(service: ExcelExportService) -> None:
    roster = _make_roster()
    item = _make_item(rule_details={"rule_5": _rule_detail("GPA門檻", True)})

    service._prepare_excel_data(roster, [item])

    cols = service.template_columns
    # base STD_UP_MIXLISTA columns survive, verification + rule columns appended
    assert "身分證字號" in cols and "分發獎學金" in cols
    assert cols.index("學籍驗證") < cols.index("GPA門檻") < cols.index("納入造冊") < cols.index("排除原因")


def test_prepare_is_idempotent_no_duplicate_columns(service: ExcelExportService) -> None:
    """Calling _prepare_excel_data twice must not append the verification
    columns twice (column list is rebuilt from the immutable base)."""
    roster = _make_roster()
    item = _make_item(rule_details={"rule_5": _rule_detail("GPA門檻", True)})

    service._prepare_excel_data(roster, [item])
    service._prepare_excel_data(roster, [item])

    cols = service.template_columns
    assert cols.count("學籍驗證") == 1
    assert cols.count("GPA門檻") == 1
    assert cols.count("納入造冊") == 1


# ---------------------------------------------------------------------------
# _resolve_cell_fill — pure red/amber/none decision
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "column,value,row,rule_headers,expected",
    [
        ("帳號", "", {}, set(), "red"),  # missing bank account
        ("帳號", "00012345678", {}, set(), None),
        ("學籍驗證", "已畢業", {}, set(), "red"),
        ("學籍驗證", "休學中", {}, set(), "red"),
        ("學籍驗證", "已驗證", {}, set(), None),
        ("學籍驗證", "—", {}, set(), None),  # unknown status never reds
        ("納入造冊", "否", {}, set(), "red"),
        ("納入造冊", "是", {}, set(), None),
        ("排除原因", "缺銀行", {"納入造冊": "否"}, set(), "red"),
        ("排除原因", "", {"納入造冊": "是"}, set(), None),
        ("GPA門檻", "未通過", {}, {"GPA門檻"}, "red"),
        ("GPA門檻", "警告", {}, {"GPA門檻"}, "amber"),
        ("GPA門檻", "通過", {}, {"GPA門檻"}, None),
        ("GPA門檻", "—", {}, {"GPA門檻"}, None),
        # a base column whose value happens to be "未通過" is NOT a rule col → no fill
        ("說明", "未通過", {}, {"GPA門檻"}, None),
    ],
)
def test_resolve_cell_fill(
    service: ExcelExportService,
    column: str,
    value: str,
    row: Dict[str, Any],
    rule_headers: set,
    expected: Optional[str],
) -> None:
    assert service._resolve_cell_fill(column, value, row, rule_headers) == expected


# ---------------------------------------------------------------------------
# End-to-end: real workbook carries the red/amber fills on the right cells
# ---------------------------------------------------------------------------


def _cell_rgb(ws, header_row: List[str], row_idx: int, column_name: str):
    col_idx = header_row.index(column_name) + 1
    return ws.cell(row=row_idx, column=col_idx).fill


def test_create_excel_file_paints_failing_cells(service: ExcelExportService, tmp_path) -> None:
    roster = _make_roster()
    item = _make_item(
        student_name="陳大文",
        bank_account=None,  # missing bank → 帳號 red
        is_included=False,
        exclusion_reason="不符合獎學金規則: GPA 不足",
        verification_status=StudentVerificationStatus.GRADUATED,  # 學籍驗證 red
        rule_details={
            "rule_2": _rule_detail("國籍", True),  # 通過 → no fill
            "rule_5": _rule_detail("GPA門檻", False, hard=True),  # 未通過 → red
            "rule_9": _rule_detail("年級", False, hard=False, warning=True),  # 警告 → amber
        },
    )

    excel_data = service._prepare_excel_data(roster, [item])
    out = os.path.join(str(tmp_path), "roster.xlsx")
    service._create_excel_file(
        excel_data,
        out,
        roster,
        template_path="/nonexistent/template.xlsx",
        include_header=True,
        include_statistics=False,
    )

    wb = load_workbook(out)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    data_row = 2

    def fill_of(col: str) -> str:
        return _cell_rgb(ws, headers, data_row, col).start_color.rgb or ""

    def has(col: str, rgb: str) -> bool:
        f = _cell_rgb(ws, headers, data_row, col)
        return f.fill_type == "solid" and (f.start_color.rgb or "").endswith(rgb)

    assert has("帳號", RED)
    assert has("學籍驗證", RED)
    assert has("GPA門檻", RED)
    assert has("年級", AMBER)
    assert has("納入造冊", RED)
    assert has("排除原因", RED)
    # the passing rule cell carries NO solid fill
    assert _cell_rgb(ws, headers, data_row, "國籍").fill_type != "solid"
