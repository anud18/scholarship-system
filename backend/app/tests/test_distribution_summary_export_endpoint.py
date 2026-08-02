"""Integration tests for GET /api/v1/manual-distribution/distribution-summary/export.

Auth is mocked via ``app.dependency_overrides[get_current_admin_user]`` — same
pattern as test_distribution_state_endpoint.py (both get_db functions must be
overridden, see the admin_client fixture note there).

Workbook assertions parse the xlsx with openpyxl, never scan response bytes —
xlsx is zip-compressed, so byte-scanning false-negatives (see
test_college_distribution_export_endpoint.py).
"""

import io

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_admin_user
from app.main import app
from app.models.application import Application
from app.models.college_review import CollegeRanking, CollegeRankingItem
from app.models.enums import ApplicationStatus, ReviewStage, SubTypeSelectionMode
from app.models.scholarship import ScholarshipConfiguration, ScholarshipSubTypeConfig, ScholarshipType
from app.models.user import User, UserRole, UserType
from app.services.manual_distribution_export_service import CHECK_CLEAR, CHECK_FLAGGED

ACADEMIC_YEAR = 114
EXPORT_PATH = "/api/v1/manual-distribution/distribution-summary/export"


@pytest_asyncio.fixture
async def admin_user(db: AsyncSession) -> User:
    user = User(
        nycu_id="admin_export",
        name="Export Admin",
        email="export_admin@university.edu",
        user_type=UserType.employee,
        role=UserRole.admin,
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return user


@pytest_asyncio.fixture
async def admin_client(client: AsyncClient, admin_user: User, db: AsyncSession):
    from app.core.deps import get_db as core_get_db

    async def override_admin():
        return admin_user

    async def override_db():
        yield db

    app.dependency_overrides[get_current_admin_user] = override_admin
    app.dependency_overrides[core_get_db] = override_db
    try:
        yield client
    finally:
        app.dependency_overrides.pop(get_current_admin_user, None)
        app.dependency_overrides.pop(core_get_db, None)


@pytest_asyncio.fixture
async def scholarship(db: AsyncSession) -> ScholarshipType:
    sch = ScholarshipType(
        code="export_sch",
        name="博士生獎學金",
        description="Fixture for distribution-summary export",
    )
    db.add(sch)
    await db.commit()
    await db.refresh(sch)

    config = ScholarshipConfiguration(
        scholarship_type_id=sch.id,
        academic_year=ACADEMIC_YEAR,
        semester=None,
        config_name="Export Config",
        config_code="export-config",
        amount=40000,
        currency="TWD",
        is_active=True,
        quotas={"nstc": {"I": 5}},
    )
    sub_type = ScholarshipSubTypeConfig(
        scholarship_type_id=sch.id,
        sub_type_code="nstc",
        name="國科會",
        is_active=True,
    )
    db.add_all([config, sub_type])
    await db.commit()
    return sch


_COLLEGE_NAMES = {"I": "工學院", "E": "電機學院"}


def _student_data(*, stdcode: str, name: str, identity: int = 1, college: str = "I") -> dict:
    return {
        "std_stdcode": stdcode,
        "std_cname": name,
        "std_nation": "中華民國",
        "std_sex": 2,
        "std_academyno": college,
        "trm_academyname": _COLLEGE_NAMES[college],
        "trm_depname": "土木工程學系",
        "std_enrollyear": 113,
        "std_enrollterm": 2,
        "std_highestschname": "台灣大學",
        "std_schoolid": 1,
        "std_enrolltype": 9,
        "std_identity": identity,
        "std_studingstatus": 1,
        "trm_studystatus": 1,
    }


async def _make_allocated_student(
    db: AsyncSession,
    *,
    scholarship: ScholarshipType,
    ranking: CollegeRanking,
    suffix: str,
    name: str,
    rank_position: int,
    allocation_config_id: int | None,
    identity: int = 1,
    allocated_sub_type: str = "nstc",
    is_allocated: bool = True,
    college: str = "I",
) -> CollegeRankingItem:
    user = User(
        nycu_id=f"stu_exp_{suffix}",
        name=name,
        email=f"stu_exp_{suffix}@university.edu",
        user_type=UserType.student,
        role=UserRole.student,
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)

    application = Application(
        app_id=f"APP-{ACADEMIC_YEAR}-0-9{suffix}",
        user_id=user.id,
        scholarship_type_id=scholarship.id,
        scholarship_subtype_list=[allocated_sub_type],
        sub_type_selection_mode=SubTypeSelectionMode.single,
        sub_scholarship_type=allocated_sub_type,
        academic_year=ACADEMIC_YEAR,
        semester=None,
        status=ApplicationStatus.approved,
        review_stage=ReviewStage.quota_distributed,
        agree_terms=True,
        student_data=_student_data(stdcode=f"3123456{suffix}", name=name, identity=identity, college=college),
        submitted_form_data={"fields": {"master_school_info": {"value": "台灣大學工學院土木工程學系"}}},
    )
    db.add(application)
    await db.commit()
    await db.refresh(application)

    item = CollegeRankingItem(
        ranking_id=ranking.id,
        application_id=application.id,
        rank_position=rank_position,
        is_allocated=is_allocated,
        allocated_sub_type=allocated_sub_type if is_allocated else None,
        allocation_config_id=allocation_config_id,
        status="allocated" if is_allocated else "ranked",
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return item


async def _make_ranking(db: AsyncSession, *, scholarship: ScholarshipType, executed: bool = True) -> CollegeRanking:
    ranking = CollegeRanking(
        scholarship_type_id=scholarship.id,
        sub_type_code="default",
        academic_year=ACADEMIC_YEAR,
        semester=None,
        is_finalized=True,
        distribution_executed=executed,
        ranking_status="finalized",
    )
    db.add(ranking)
    await db.commit()
    await db.refresh(ranking)
    return ranking


@pytest_asyncio.fixture
async def seeded_distribution(db: AsyncSession, scholarship: ScholarshipType) -> CollegeRanking:
    """Two allocated students (ranks 2 and 1, seeded out of order) + config quota."""
    from sqlalchemy import select

    config_id = (
        (
            await db.execute(
                select(ScholarshipConfiguration.id).where(
                    ScholarshipConfiguration.scholarship_type_id == scholarship.id
                )
            )
        )
        .scalars()
        .first()
    )
    ranking = await _make_ranking(db, scholarship=scholarship)
    await _make_allocated_student(
        db,
        scholarship=scholarship,
        ranking=ranking,
        suffix="02",
        name="李陸生",
        rank_position=2,
        allocation_config_id=config_id,
        identity=17,
    )
    await _make_allocated_student(
        db,
        scholarship=scholarship,
        ranking=ranking,
        suffix="01",
        name="王小美",
        rank_position=1,
        allocation_config_id=config_id,
    )
    return ranking


def _params(scholarship: ScholarshipType, **overrides):
    params = {
        "scholarship_type_id": scholarship.id,
        "academic_year": ACADEMIC_YEAR,
        "semester": "yearly",
    }
    params.update(overrides)
    return params


@pytest.mark.asyncio
async def test_xlsx_export_headers_and_rows(admin_client, scholarship, seeded_distribution):
    resp = await admin_client.get(EXPORT_PATH, params=_params(scholarship))
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert resp.content[:2] == b"PK"
    assert resp.headers["content-disposition"].startswith("attachment; filename*=UTF-8''")
    assert resp.headers["content-length"] == str(len(resp.content))

    wb = load_workbook(io.BytesIO(resp.content))
    # Sheet named after the ScholarshipSubTypeConfig label + consumed config's year.
    assert wb.sheetnames == [f"國科會_{ACADEMIC_YEAR}"]
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "序號"
    assert ws.cell(row=2, column=2).value == "基本資料"

    # Rows are ordered by rank_position — 王小美 (rank 1) first despite being
    # seeded second — and 序號 restarts from 1.
    first = [ws.cell(row=4, column=c).value for c in range(1, 14)]
    assert first == [
        1,
        "工學院",
        "土木工程學系",
        "王小美",
        "中華民國",
        "女",
        "台灣大學工學院土木工程學系",
        "113.02.01",
        "312345601",
        None,
        CHECK_CLEAR,
        CHECK_CLEAR,
        CHECK_CLEAR,
    ]
    # 陸生 (std_identity=17) row: cross-strait auto-check flags 有.
    second = [ws.cell(row=5, column=c).value for c in range(1, 14)]
    assert second[0] == 2
    assert second[3] == "李陸生"
    assert second[11] == CHECK_FLAGGED


@pytest.mark.asyncio
async def test_pdf_export_returns_pdf(admin_client, scholarship, seeded_distribution):
    resp = await admin_client.get(EXPORT_PATH, params=_params(scholarship, format="pdf"))
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF")
    assert resp.headers["content-length"] == str(len(resp.content))


@pytest.mark.asyncio
async def test_unknown_sub_type_label_falls_back_to_raw_code(admin_client, scholarship, db):
    ranking = await _make_ranking(db, scholarship=scholarship)
    await _make_allocated_student(
        db,
        scholarship=scholarship,
        ranking=ranking,
        suffix="77",
        name="自訂類別生",
        rank_position=1,
        allocation_config_id=None,
        allocated_sub_type="custom_new_type",
    )
    resp = await admin_client.get(EXPORT_PATH, params=_params(scholarship))
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == [f"custom_new_type_{ACADEMIC_YEAR}"]


@pytest.mark.asyncio
async def test_no_finalized_distribution_is_404(admin_client, scholarship):
    resp = await admin_client.get(EXPORT_PATH, params=_params(scholarship))
    assert resp.status_code == 404
    # The app-level HTTPException handler rewraps detail as {success, message}.
    assert "尚未完成分發" in resp.json()["message"]


@pytest.mark.asyncio
async def test_distribution_without_allocated_students_is_404(admin_client, scholarship, db):
    ranking = await _make_ranking(db, scholarship=scholarship)
    await _make_allocated_student(
        db,
        scholarship=scholarship,
        ranking=ranking,
        suffix="55",
        name="未分配生",
        rank_position=1,
        allocation_config_id=None,
        is_allocated=False,
    )
    resp = await admin_client.get(EXPORT_PATH, params=_params(scholarship))
    assert resp.status_code == 404
    assert "尚無已分配" in resp.json()["message"]


@pytest.mark.asyncio
async def test_unknown_format_is_422(admin_client, scholarship, seeded_distribution):
    resp = await admin_client.get(EXPORT_PATH, params=_params(scholarship, format="csv"))
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_rows_are_ordered_by_college_then_rank(admin_client, scholarship, db):
    """rank_position is scoped to ONE college's ranking, so ordering on it alone
    interleaves colleges and the roster reads as if ranks were global."""
    ranking_i = await _make_ranking(db, scholarship=scholarship)
    for suffix, name, rank, college in [
        ("31", "工一", 1, "I"),
        ("32", "工二", 2, "I"),
        ("41", "電一", 1, "E"),
        ("42", "電二", 2, "E"),
    ]:
        await _make_allocated_student(
            db,
            scholarship=scholarship,
            ranking=ranking_i,
            suffix=suffix,
            name=name,
            rank_position=rank,
            allocation_config_id=None,
            college=college,
        )

    resp = await admin_client.get(EXPORT_PATH, params=_params(scholarship))
    assert resp.status_code == 200
    ws = load_workbook(io.BytesIO(resp.content)).active
    names = [ws.cell(row=r, column=4).value for r in range(4, 8)]
    seqs = [ws.cell(row=r, column=1).value for r in range(4, 8)]
    # 學院 E sorts before I; within a college, by rank.
    assert names == ["電一", "電二", "工一", "工二"]
    assert seqs == [1, 2, 3, 4]


@pytest.mark.asyncio
async def test_multiple_groups_share_one_running_序號_in_sub_type_order(admin_client, scholarship, db):
    ranking = await _make_ranking(db, scholarship=scholarship)
    for suffix, name, rank, sub_type in [
        ("61", "科會一", 1, "nstc"),
        ("62", "科會二", 2, "nstc"),
        ("63", "教部一", 1, "moe_1w"),
    ]:
        await _make_allocated_student(
            db,
            scholarship=scholarship,
            ranking=ranking,
            suffix=suffix,
            name=name,
            rank_position=rank,
            allocation_config_id=None,
            allocated_sub_type=sub_type,
        )

    resp = await admin_client.get(EXPORT_PATH, params=_params(scholarship))
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    # Groups sort by sub_type code: moe_1w before nstc. moe_1w has no
    # ScholarshipSubTypeConfig row here, so it falls back to the raw code.
    assert wb.sheetnames == [f"moe_1w_{ACADEMIC_YEAR}", f"國科會_{ACADEMIC_YEAR}"]

    moe = wb[f"moe_1w_{ACADEMIC_YEAR}"]
    assert [moe.cell(row=4, column=1).value, moe.cell(row=4, column=4).value] == [1, "教部一"]
    assert moe.cell(row=5, column=1).value is None  # only one row in this group

    # 序號 is ONE running number across the whole document, not per sheet: the
    # 名冊 is filed as a single document, so the 承辦人 can cite 序號 N unambiguously.
    nstc = wb[f"國科會_{ACADEMIC_YEAR}"]
    assert [nstc.cell(row=r, column=1).value for r in (4, 5)] == [2, 3]
    assert [nstc.cell(row=r, column=4).value for r in (4, 5)] == ["科會一", "科會二"]


@pytest.mark.asyncio
async def test_export_writes_a_pii_access_audit_log(admin_client, scholarship, seeded_distribution, db, admin_user):
    """The file carries 國籍/性別/碩士畢業校系 + 學籍 flags for identified students,
    so the bulk export must be attributable."""
    from sqlalchemy import select

    from app.models.audit_log import AuditAction, AuditLog

    resp = await admin_client.get(EXPORT_PATH, params=_params(scholarship, format="pdf"))
    assert resp.status_code == 200

    logs = (await db.execute(select(AuditLog).where(AuditLog.action == AuditAction.pii_access.value))).scalars().all()
    assert len(logs) == 1
    entry = logs[0]
    assert entry.user_id == admin_user.id
    assert entry.resource_id == str(scholarship.id)
    assert entry.meta_data["export_format"] == "pdf"
    assert entry.meta_data["record_count"] == 2
    assert len(entry.meta_data["application_ids"]) == 2


@pytest.mark.asyncio
async def test_failed_export_writes_no_audit_log(admin_client, scholarship, db):
    """A 404 must not leave a pii_access record claiming data left the system."""
    from sqlalchemy import select

    from app.models.audit_log import AuditAction, AuditLog

    resp = await admin_client.get(EXPORT_PATH, params=_params(scholarship))
    assert resp.status_code == 404
    logs = (await db.execute(select(AuditLog).where(AuditLog.action == AuditAction.pii_access.value))).scalars().all()
    assert logs == []


@pytest.mark.asyncio
async def test_json_summary_endpoint_still_serves_the_same_students(admin_client, scholarship, seeded_distribution):
    """Regression: the JSON endpoint was refactored onto the shared loader."""
    resp = await admin_client.get("/api/v1/manual-distribution/distribution-summary", params=_params(scholarship))
    assert resp.status_code == 200
    body = resp.json()
    assert body["success"] is True
    assert body["data"]["total_allocated"] == 2
    assert len(body["data"]["groups"]) == 1
    group = body["data"]["groups"][0]
    assert set(group) == {"sub_type", "allocation_config_id", "allocation_year", "count", "students"}
    assert group["sub_type"] == "nstc"
    assert group["allocation_year"] == ACADEMIC_YEAR
    assert group["count"] == 2

    # Pin the per-student payload the panel consumes — this endpoint moved onto a
    # loader shared with the export, so a dropped key must fail here, not in the UI.
    students = sorted(group["students"], key=lambda s: s["rank_position"])
    assert set(students[0]) == {
        "ranking_item_id",
        "application_id",
        "student_name",
        "student_id",
        "college_code",
        "college_name",
        "department_name",
        "rank_position",
        "college_rejected",
        "is_supplementary",
        "is_renewal",
        "renewal_year",
    }
    first = students[0]
    assert first["student_name"] == "王小美"
    assert first["student_id"] == "312345601"
    assert first["college_code"] == "I"
    assert first["college_name"] == "工學院"
    assert first["department_name"] == "土木工程學系"
    assert first["rank_position"] == 1
    assert first["college_rejected"] is False
    assert first["is_supplementary"] is False
    assert first["is_renewal"] is False
    assert first["renewal_year"] is None
    assert [s["student_name"] for s in students] == ["王小美", "李陸生"]
