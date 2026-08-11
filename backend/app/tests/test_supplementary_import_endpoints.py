"""Integration tests for supplementary import endpoints."""

import io

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import require_admin, require_college
from app.main import app
from app.models.scholarship import (
    ScholarshipConfiguration,
    ScholarshipSubTypeConfig,
    ScholarshipType,
    SubTypeSelectionMode,
)
from app.models.user import AdminScholarship, User, UserRole, UserType


def _build_xlsx_bytes(student_id: str = "310460099") -> bytes:
    """A minimal 批次匯入-format workbook — the same shape the shared template
    generator emits, which is what 補充匯入 now accepts."""
    wb = Workbook()
    ws = wb.active
    headers = ["學號", "學生姓名", "郵局帳號", "指導教授本校人事編號", "國科會"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    for col_idx, val in enumerate([student_id, "王小明", "12345678", "P0001", 1], start=1):
        ws.cell(row=2, column=col_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest_asyncio.fixture
async def admin_user(db: AsyncSession) -> User:
    user = User(
        nycu_id="admin001",
        name="Admin",
        email="admin@nycu.edu.tw",
        user_type=UserType.employee,
        role=UserRole.admin,
    )
    db.add(user)
    await db.flush()
    return user


@pytest_asyncio.fixture
async def college_user(db: AsyncSession) -> User:
    user = User(
        nycu_id="col001",
        name="College",
        email="col@nycu.edu.tw",
        user_type=UserType.employee,
        role=UserRole.college,
        college_code="A",
    )
    db.add(user)
    await db.flush()
    return user


@pytest_asyncio.fixture
async def scholarship(db: AsyncSession) -> ScholarshipType:
    s = ScholarshipType(
        code="phd_supp_test",
        name="Test Supp PhD",
        sub_type_list=["nstc"],
        sub_type_selection_mode=SubTypeSelectionMode.single,
        status="active",
    )
    db.add(s)
    await db.flush()
    return s


@pytest_asyncio.fixture
async def configuration(db: AsyncSession, scholarship: ScholarshipType, admin_user: User) -> ScholarshipConfiguration:
    cfg = ScholarshipConfiguration(
        scholarship_type_id=scholarship.id,
        academic_year=114,
        semester=None,  # yearly
        config_name="Test PhD 114",
        config_code="test-phd-114",
        amount=40000,
        is_active=True,
        allow_supplementary_import=False,
    )
    db.add(cfg)
    await db.flush()

    # Grant admin + college permission to manage this scholarship type
    db.add(AdminScholarship(admin_id=admin_user.id, scholarship_id=scholarship.id))
    await db.flush()
    return cfg


@pytest_asyncio.fixture
async def college_grant(db: AsyncSession, college_user: User, scholarship: ScholarshipType) -> None:
    """_check_scholarship_permission requires an explicit grant for college users."""
    db.add(AdminScholarship(admin_id=college_user.id, scholarship_id=scholarship.id))
    await db.flush()


UPLOAD_URL = "/api/v1/college-review/supplementary-import/upload"
AVAILABILITY_URL = "/api/v1/college-review/supplementary-import/availability"
PERIOD_QUERY = {"scholarship_type": "phd_supp_test", "academic_year": 114, "semester": "yearly"}


def _xlsx_files():
    return {
        "file": (
            "test.xlsx",
            _build_xlsx_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }


@pytest.mark.asyncio
class TestAdminConfigToggle:
    async def test_admin_can_enable_supplementary_import(
        self,
        client: AsyncClient,
        db: AsyncSession,
        admin_user: User,
        configuration: ScholarshipConfiguration,
    ):
        app.dependency_overrides[require_admin] = lambda: admin_user
        try:
            resp = await client.patch(
                f"/api/v1/scholarship-configurations/configurations/{configuration.id}/supplementary-import",
                json={"allow": True},
            )
        finally:
            app.dependency_overrides.pop(require_admin, None)
        assert resp.status_code == 200, resp.text
        data = resp.json()
        assert data["data"]["allow_supplementary_import"] is True

    async def test_returns_404_for_unknown_configuration(self, client: AsyncClient, db: AsyncSession, admin_user: User):
        app.dependency_overrides[require_admin] = lambda: admin_user
        try:
            resp = await client.patch(
                "/api/v1/scholarship-configurations/configurations/999999/supplementary-import",
                json={"allow": True},
            )
        finally:
            app.dependency_overrides.pop(require_admin, None)
        assert resp.status_code == 404


@pytest.mark.asyncio
class TestSupplementaryImportAvailability:
    async def test_reports_closed_without_403(
        self,
        client: AsyncClient,
        db: AsyncSession,
        college_user: User,
        configuration: ScholarshipConfiguration,
        college_grant: None,
    ):
        """A closed period must answer allowed=false rather than 403 — the panel
        needs to explain why upload is disabled before the college picks a file."""
        app.dependency_overrides[require_college] = lambda: college_user
        try:
            resp = await client.get(AVAILABILITY_URL, params=PERIOD_QUERY)
        finally:
            app.dependency_overrides.pop(require_college, None)
        assert resp.status_code == 200, resp.text
        data = resp.json()["data"]
        assert data["allowed"] is False
        assert data["configuration_id"] == configuration.id

    async def test_reports_open_when_flag_on(
        self,
        client: AsyncClient,
        db: AsyncSession,
        college_user: User,
        configuration: ScholarshipConfiguration,
        college_grant: None,
    ):
        configuration.allow_supplementary_import = True
        await db.commit()

        app.dependency_overrides[require_college] = lambda: college_user
        try:
            resp = await client.get(AVAILABILITY_URL, params=PERIOD_QUERY)
        finally:
            app.dependency_overrides.pop(require_college, None)
        assert resp.status_code == 200, resp.text
        assert resp.json()["data"]["allowed"] is True


@pytest.mark.asyncio
class TestSupplementaryImportEndpoint:
    async def test_returns_403_when_flag_is_off(
        self,
        client: AsyncClient,
        db: AsyncSession,
        college_user: User,
        configuration: ScholarshipConfiguration,
        college_grant: None,
    ):
        app.dependency_overrides[require_college] = lambda: college_user
        try:
            resp = await client.post(UPLOAD_URL, params=PERIOD_QUERY, files=_xlsx_files())
        finally:
            app.dependency_overrides.pop(require_college, None)
        assert resp.status_code == 403
        body = resp.json()
        assert "補充匯入" in (body.get("message") or body.get("detail") or "")

    async def test_returns_404_for_unknown_scholarship_type(
        self, client: AsyncClient, db: AsyncSession, college_user: User
    ):
        app.dependency_overrides[require_college] = lambda: college_user
        try:
            resp = await client.post(
                UPLOAD_URL,
                params={**PERIOD_QUERY, "scholarship_type": "does_not_exist"},
                files=_xlsx_files(),
            )
        finally:
            app.dependency_overrides.pop(require_college, None)
        assert resp.status_code == 404

    async def test_returns_403_when_college_has_no_college_code(
        self,
        client: AsyncClient,
        db: AsyncSession,
        configuration: ScholarshipConfiguration,
    ):
        """Without a college binding there is no scope to import into — an
        unbound account must not be able to import arbitrary students."""
        unbound = User(
            nycu_id="col_unbound",
            name="Unbound",
            email="unbound@nycu.edu.tw",
            user_type=UserType.employee,
            role=UserRole.college,
            college_code=None,
        )
        db.add(unbound)
        await db.flush()

        app.dependency_overrides[require_college] = lambda: unbound
        try:
            resp = await client.post(UPLOAD_URL, params=PERIOD_QUERY, files=_xlsx_files())
        finally:
            app.dependency_overrides.pop(require_college, None)
        assert resp.status_code == 403
        assert "未綁定學院" in (resp.json().get("message") or resp.json().get("detail") or "")

    async def test_rejects_student_from_other_college(
        self,
        client: AsyncClient,
        db: AsyncSession,
        college_user: User,  # college_code = "A"
        configuration: ScholarshipConfiguration,
        scholarship: ScholarshipType,
        college_grant: None,
        monkeypatch,
    ):
        """Student whose SIS std_academyno doesn't match the caller's college
        must be rejected with 422 and the offending student ID listed."""
        # Enable the supplementary-import flag for the configuration
        configuration.allow_supplementary_import = True
        # Seed sub-type config so parse_excel can resolve the Excel label
        db.add(
            ScholarshipSubTypeConfig(
                scholarship_type_id=scholarship.id,
                sub_type_code="nstc",
                name="國科會博士生研究獎學金",
                is_active=True,
            )
        )
        await db.commit()

        # Bypass SIS: return a student whose std_academyno is "EE" (not "A")
        async def fake_fetch(self, student_ids, **kwargs):
            return (
                {sid: {"std_academyno": "EE", "std_cname": "他院學生"} for sid in student_ids},
                [],
                [],
            )

        from app.services.supplementary_import_service import SupplementaryImportService

        monkeypatch.setattr(SupplementaryImportService, "fetch_student_data_bulk", fake_fetch)

        app.dependency_overrides[require_college] = lambda: college_user
        try:
            resp = await client.post(UPLOAD_URL, params=PERIOD_QUERY, files=_xlsx_files())
        finally:
            app.dependency_overrides.pop(require_college, None)

        assert resp.status_code == 422, resp.text
        body = resp.json()
        # ApiResponse wrapper format
        message = body.get("message") or body.get("detail") or ""
        assert "不屬於本學院" in message
        assert "310460099" in message  # student_id from _build_xlsx_bytes
        assert "EE" in message  # surfaced mismatched college code

    async def test_creates_submitted_application_without_ranking_item(
        self,
        client: AsyncClient,
        db: AsyncSession,
        college_user: User,
        configuration: ScholarshipConfiguration,
        scholarship: ScholarshipType,
        college_grant: None,
        monkeypatch,
    ):
        """Happy path: the import mints an ordinary submitted application and
        leaves ranking untouched — 名次 is decided by the normal ranking flow."""
        from sqlalchemy import select

        from app.models.application import Application, ApplicationStatus
        from app.models.college_review import CollegeRankingItem
        from app.services.supplementary_import_service import SupplementaryImportService

        configuration.allow_supplementary_import = True
        db.add(
            ScholarshipSubTypeConfig(
                scholarship_type_id=scholarship.id,
                sub_type_code="nstc",
                name="國科會博士生研究獎學金",
                is_active=True,
            )
        )
        await db.commit()

        async def fake_fetch(self, student_ids, **kwargs):
            return (
                {sid: {"std_academyno": "A", "std_cname": "本院學生", "std_stdcode": sid} for sid in student_ids},
                [],
                [],
            )

        monkeypatch.setattr(SupplementaryImportService, "fetch_student_data_bulk", fake_fetch)

        app.dependency_overrides[require_college] = lambda: college_user
        try:
            resp = await client.post(UPLOAD_URL, params=PERIOD_QUERY, files=_xlsx_files())
        finally:
            app.dependency_overrides.pop(require_college, None)

        assert resp.status_code == 200, resp.text
        data = resp.json()["data"]
        assert data["imported_count"] == 1
        assert data["student_ids"] == ["310460099"]
        # No rank is reported anywhere — the old response advertised a rank range
        assert "new_rank_range" not in data
        assert "max_existing_rank" not in data

        app_row = (await db.execute(select(Application).where(Application.app_id.like("APP-114-0-%")))).scalar_one()
        assert app_row.status == ApplicationStatus.submitted
        assert app_row.scholarship_configuration_id == configuration.id
        assert app_row.import_source == "supplementary_import"

        items = (
            (await db.execute(select(CollegeRankingItem).where(CollegeRankingItem.application_id == app_row.id)))
            .scalars()
            .all()
        )
        assert items == []
