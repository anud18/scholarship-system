"""
Per-scholarship worksheet tabs in the payment-roster Excel export.

The admin 造冊分發 Excel keeps its full 印領清冊 main sheet (the
STD_UP_MIXLISTA list finance ingests), and additionally gets one worksheet
per scholarship, each holding only that scholarship's roster rows:

- Rows with an allocation snapshot group by「{allocation_year}年 {sub_type
  label}」— the exact same string as the 分發獎學金 column (e.g.
  「114年 國科會」,「113年 國科會」,「114年 教育部(5000)」).
- Rows without an allocated sub-type fall back to `scholarship_name`.
- Tabs are ordered year-descending first (114 → 113), then label; the
  造冊資訊 statistics sheet stays last.

Covers the label helper, sheet ordering, Excel-illegal sheet-title
sanitization (31-char cap, []:*?/\\ stripped, dedupe), and the end-to-end
`export_roster_to_excel` path including the skipped-invalid-row parallelism
(labels are built with the same `_has_required_export_fields` predicate as
`_prepare_excel_data`, so the strict zip in `_add_scholarship_sheets` must
not blow up when a row is dropped).

Stubs follow the SimpleNamespace pattern of `test_excel_export_service_rows.py`.
"""

from __future__ import annotations

from types import SimpleNamespace
from typing import Any, Optional

import openpyxl
import pytest

from app.services.excel_export_service import ExcelExportService

# ---------------------------------------------------------------------------
# Fixtures / builders
# ---------------------------------------------------------------------------


@pytest.fixture
def service() -> ExcelExportService:
    return ExcelExportService()


def _make_item(
    *,
    student_id_number: str = "A123456789",
    student_name: str = "王小明",
    student_email: str = "wang@nycu.edu.tw",
    bank_account: Optional[str] = "00012345678",
    scholarship_name: str = "博士班獎學金",
    scholarship_amount: Any = 50000,
    permanent_address: Optional[str] = None,
    application_identity: Optional[str] = "114新申請",
    allocated_sub_type: Optional[str] = None,
    allocation_year: Optional[int] = None,
    is_included: bool = True,
    verification_status: str = "verified",
    is_eligible: Optional[bool] = True,
    rule_validation_result: Optional[dict] = None,
    exclusion_reason: Optional[str] = None,
) -> SimpleNamespace:
    return SimpleNamespace(
        student_id_number=student_id_number,
        student_name=student_name,
        student_email=student_email,
        bank_account=bank_account,
        scholarship_name=scholarship_name,
        scholarship_amount=scholarship_amount,
        permanent_address=permanent_address,
        application_identity=application_identity,
        allocated_sub_type=allocated_sub_type,
        allocation_year=allocation_year,
        is_included=is_included,
        verification_status=verification_status,
        is_eligible=is_eligible,
        rule_validation_result=rule_validation_result,
        exclusion_reason=exclusion_reason,
        excel_row_data=None,
        excel_remarks=None,
    )


def _make_roster(items) -> SimpleNamespace:
    """Roster stub with everything `export_roster_to_excel` touches.

    id=None + skip_minio_upload=True keeps the export purely local; the
    excel_* attributes are write-back targets."""
    return SimpleNamespace(
        id=None,
        period_label="2025-H1",
        roster_code="ROSTER-114-2025-H1-PHD001",
        academic_year=114,
        items=items,
        generate_excel_filename=lambda: "roster_tabs_test.xlsx",
        minio_object_name=None,
        excel_filename=None,
        excel_file_path=None,
        excel_file_size=None,
        excel_file_hash=None,
        # _add_worksheet_info (造冊資訊 statistics sheet) reads these
        roster_cycle=SimpleNamespace(value="monthly"),
        trigger_type=SimpleNamespace(value="manual"),
        started_at=None,
        completed_at=None,
        total_applications=1,
        qualified_count=1,
        disqualified_count=0,
        total_amount=50000,
        student_verification_enabled=False,
        verification_api_failures=0,
    )


# ---------------------------------------------------------------------------
# _get_scholarship_sheet_label — grouping key per roster row
# ---------------------------------------------------------------------------


def test_sheet_label_uses_allocation_year_and_sub_type(service):
    item = _make_item(allocated_sub_type="nstc", allocation_year=114)
    assert service._get_scholarship_sheet_label(item) == "114年 國科會"


def test_sheet_label_sub_type_without_year(service):
    item = _make_item(allocated_sub_type="moe_1w", allocation_year=None)
    assert service._get_scholarship_sheet_label(item) == "教育部(5000)"


def test_sheet_label_falls_back_to_roster_year_when_item_year_null(service):
    """Monthly/legacy path leaves item.allocation_year NULL — the label must
    fall back to the roster's academic_year (same semantics as 查看名單's
    `_roster_item_dict_with_display_year`), so每個年度各自成頁."""
    unallocated = _make_item(allocated_sub_type="nstc", allocation_year=None)
    borrowed = _make_item(allocated_sub_type="nstc", allocation_year=113)

    assert service._get_scholarship_sheet_label(unallocated, 114) == "114年 國科會"
    # An explicit prior-year snapshot must NOT be overwritten by the fallback
    assert service._get_scholarship_sheet_label(borrowed, 114) == "113年 國科會"


def test_sheet_label_falls_back_to_scholarship_name(service):
    item = _make_item(allocated_sub_type=None, scholarship_name="逕博獎學金")
    assert service._get_scholarship_sheet_label(item) == "逕博獎學金"


def test_sheet_label_unknown_when_nothing_available(service):
    item = _make_item(allocated_sub_type=None, scholarship_name="")
    assert service._get_scholarship_sheet_label(item) == "未分類"


# ---------------------------------------------------------------------------
# _scholarship_sheet_order — year-descending, then label
# ---------------------------------------------------------------------------


def test_sheet_order_year_descending_then_no_year_groups(service):
    labels = ["教育部(5000)", "113年 國科會", "逕博獎學金", "114年 國科會", "114年 教育部(5000)"]

    ordered = sorted(labels, key=service._scholarship_sheet_order)

    assert ordered == [
        "114年 國科會",
        "114年 教育部(5000)",
        "113年 國科會",
        "教育部(5000)",
        "逕博獎學金",
    ]


# ---------------------------------------------------------------------------
# _safe_sheet_title — Excel-legal, 31-char cap, unique
# ---------------------------------------------------------------------------


def test_safe_sheet_title_strips_illegal_chars(service):
    used: set = set()
    assert service._safe_sheet_title("114年 國科會/教育部[A]", used) == "114年 國科會_教育部_A_"


def test_safe_sheet_title_truncates_to_31_chars(service):
    used: set = set()
    title = service._safe_sheet_title("x" * 40, used)
    assert title == "x" * 31


def test_safe_sheet_title_dedupes_with_counter_suffix(service):
    used: set = set()
    first = service._safe_sheet_title("國科會", used)
    second = service._safe_sheet_title("國科會", used)
    third = service._safe_sheet_title("國科會", used)

    assert first == "國科會"
    assert second == "國科會(2)"
    assert third == "國科會(3)"
    # Dedupe of an already-max-length title must stay within 31 chars
    long_dup = service._safe_sheet_title("x" * 40, {"x" * 31})
    assert len(long_dup) <= 31
    assert long_dup.endswith("(2)")


def test_safe_sheet_title_empty_label_becomes_unclassified(service):
    assert service._safe_sheet_title("", set()) == "未分類"


# ---------------------------------------------------------------------------
# End-to-end — export_roster_to_excel produces the per-scholarship tabs
# ---------------------------------------------------------------------------


def _export(service, tmp_path, items, **kwargs):
    service.export_base_path = str(tmp_path)
    roster = _make_roster(items)
    result = service.export_roster_to_excel(roster, skip_minio_upload=True, **kwargs)
    return openpyxl.load_workbook(result["file_path"])


def test_export_creates_one_tab_per_scholarship(service, tmp_path):
    items = [
        _make_item(student_id_number="A1", student_name="甲", allocated_sub_type="nstc", allocation_year=114),
        _make_item(student_id_number="A2", student_name="乙", allocated_sub_type="nstc", allocation_year=113),
        _make_item(student_id_number="A3", student_name="丙", allocated_sub_type="moe_1w", allocation_year=114),
        _make_item(student_id_number="A4", student_name="丁", allocated_sub_type="nstc", allocation_year=114),
        _make_item(student_id_number="A5", student_name="戊", allocated_sub_type=None, scholarship_name="逕博獎學金"),
    ]

    wb = _export(service, tmp_path, items, include_statistics=False)

    # Main full-list sheet first, then per-scholarship tabs year-descending
    assert wb.sheetnames == ["印領清冊", "114年 國科會", "114年 教育部(5000)", "113年 國科會", "逕博獎學金"]

    def names_on(sheet_title):
        ws = wb[sheet_title]
        return [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]

    # Main sheet keeps the full roster (sorted by student_name upstream)
    assert sorted(names_on("印領清冊")) == ["丁", "丙", "乙", "戊", "甲"]
    # Each tab holds exactly that scholarship's students
    assert sorted(names_on("114年 國科會")) == ["丁", "甲"]
    assert names_on("113年 國科會") == ["乙"]
    assert names_on("114年 教育部(5000)") == ["丙"]
    assert names_on("逕博獎學金") == ["戊"]


def test_export_tab_headers_match_main_sheet(service, tmp_path):
    items = [_make_item(student_id_number="A1", student_name="甲", allocated_sub_type="nstc", allocation_year=114)]

    wb = _export(service, tmp_path, items, include_statistics=False)

    main_header = [c.value for c in wb["印領清冊"][1]]
    tab_header = [c.value for c in wb["114年 國科會"][1]]
    assert tab_header == main_header
    assert "身分證字號" in tab_header and "分發獎學金" in tab_header


def test_export_tab_preserves_cell_fills(service, tmp_path):
    """The red missing-bank-account fill must carry into the scholarship tab,
    not just the main sheet — the operator reviews per-scholarship pages."""
    items = [
        _make_item(
            student_id_number="A1",
            student_name="甲",
            bank_account=None,
            allocated_sub_type="nstc",
            allocation_year=114,
        )
    ]

    wb = _export(service, tmp_path, items, include_statistics=False)

    ws = wb["114年 國科會"]
    header = [c.value for c in ws[1]]
    account_col = header.index("帳號") + 1
    assert "FFC7CE" in str(ws.cell(row=2, column=account_col).fill.start_color.rgb)


def test_export_statistics_sheet_stays_last(service, tmp_path):
    items = [_make_item(student_id_number="A1", student_name="甲", allocated_sub_type="nstc", allocation_year=114)]

    wb = _export(service, tmp_path, items, include_statistics=True)

    assert wb.sheetnames[-1] == "造冊資訊"
    assert "114年 國科會" in wb.sheetnames


def test_export_splits_nstc_per_year_even_without_allocation_snapshot(service, tmp_path):
    """Regression (user report): a roster mixing current-year rows
    (allocation_year=NULL — the normal generation path) with prior-year-quota
    rows (allocation_year=113) must yield SEPARATE 國科會 tabs per year.
    Before the fallback, NULL-year rows rendered a year-less「國科會」tab and
    every year collapsed into it."""
    items = [
        _make_item(student_id_number="A1", student_name="甲", allocated_sub_type="nstc", allocation_year=None),
        _make_item(student_id_number="A2", student_name="乙", allocated_sub_type="nstc", allocation_year=113),
        _make_item(student_id_number="A3", student_name="丙", allocated_sub_type="nstc", allocation_year=None),
    ]

    wb = _export(service, tmp_path, items, include_statistics=False)

    assert wb.sheetnames == ["印領清冊", "114年 國科會", "113年 國科會"]
    ws114, ws113 = wb["114年 國科會"], wb["113年 國科會"]
    assert sorted(ws114.cell(row=r, column=2).value for r in range(2, ws114.max_row + 1)) == ["丙", "甲"]
    assert [ws113.cell(row=r, column=2).value for r in range(2, ws113.max_row + 1)] == ["乙"]
    # The 分發獎學金 column shows the same resolved year as the tab
    header = [c.value for c in ws114[1]]
    alloc_col = header.index("分發獎學金") + 1
    assert ws114.cell(row=2, column=alloc_col).value == "114年 國科會"


def test_scholarship_named_like_statistics_sheet_cannot_take_its_title(service, tmp_path):
    """A scholarship whose name collides with the reserved 造冊資訊 title must
    NOT squat on it — openpyxl would then silently rename the real statistics
    sheet to 造冊資訊1, swapping the two sheets' identities. The reservation in
    `_add_scholarship_sheets` forces the scholarship tab to 造冊資訊(2) and the
    statistics sheet keeps its canonical name (and stays last)."""
    items = [
        _make_item(student_id_number="A1", student_name="甲", allocated_sub_type=None, scholarship_name="造冊資訊")
    ]

    wb = _export(service, tmp_path, items, include_statistics=True)

    assert wb.sheetnames == ["印領清冊", "造冊資訊(2)", "造冊資訊"]
    # The real statistics sheet is the label/value page, not a roster page
    assert wb["造冊資訊"].cell(row=1, column=1).value == "造冊代碼"
    assert wb["造冊資訊(2)"].cell(row=2, column=2).value == "甲"


def test_statistics_title_not_reserved_when_statistics_disabled(service, tmp_path):
    """include_statistics=False (client-settable via the export endpoint) means
    no statistics sheet exists — the reservation must not apply, or the
    scholarship tab would be named 造冊資訊(2) with no 造冊資訊 anywhere,
    reading as a lost sheet."""
    items = [
        _make_item(student_id_number="A1", student_name="甲", allocated_sub_type=None, scholarship_name="造冊資訊")
    ]

    wb = _export(service, tmp_path, items, include_statistics=False)

    assert wb.sheetnames == ["印領清冊", "造冊資訊"]
    assert wb["造冊資訊"].cell(row=2, column=2).value == "甲"


def test_export_skipped_invalid_rows_keep_labels_parallel(service, tmp_path):
    """A row dropped for missing 身分證字號 must also be dropped from the
    label list (shared `_has_required_export_fields` predicate) — otherwise
    the strict zip in `_add_scholarship_sheets` raises and the whole export
    fails. The invalid row's scholarship must not get a tab of its own."""
    items = [
        _make_item(student_id_number="", student_name="無ID", allocated_sub_type="moe_2w", allocation_year=114),
        _make_item(student_id_number="A1", student_name="甲", allocated_sub_type="nstc", allocation_year=114),
    ]

    wb = _export(service, tmp_path, items, include_statistics=False)

    assert wb.sheetnames == ["印領清冊", "114年 國科會"]
    ws = wb["114年 國科會"]
    assert ws.cell(row=2, column=2).value == "甲"
    assert ws.max_row == 2
