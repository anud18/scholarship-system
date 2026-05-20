"""
Tests for `WhitelistExcelService.generate_template` +
`export_whitelist`.

Wave 6a44 covered `parse_import_excel`. This wave fills the two
write-side helpers:

  - **generate_template(sub_types)**: writes a 4-column Excel
    file (學號 / 姓名 / 子獎學金類型 / 備註) with two example
    rows and a 使用說明 sheet listing the valid sub_types.

  - **export_whitelist(whitelist_data, scholarship_name)**:
    flattens a nested {sub_type: [students]} dict into a
    flat 4-column Excel listing.

10 cases. Pure helpers — service writes to in-memory BytesIO
and we read it back with openpyxl.
"""

import io

import pytest
from openpyxl import load_workbook

from app.services.whitelist_excel_service import WhitelistExcelService


@pytest.fixture
def service():
    return WhitelistExcelService()


# ─── generate_template ──────────────────────────────────────────────


def test_template_headers_are_canonical_four_columns(service):
    # Pin: ["學號", "姓名", "子獎學金類型", "備註"] in that order.
    # Drift in column order silently breaks import (column indices
    # are positional in parse_import_excel).
    buf = service.generate_template(["nstc", "moe_1w"])
    wb = load_workbook(buf)
    ws = wb["白名單匯入模板"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    assert headers == ["學號", "姓名", "子獎學金類型", "備註"]


def test_template_includes_two_example_rows(service):
    # Pin: 2 example rows show admins the expected format.
    buf = service.generate_template(["nstc", "moe_1w"])
    wb = load_workbook(buf)
    ws = wb["白名單匯入模板"]
    assert ws.cell(row=2, column=1).value == "0856001"
    assert ws.cell(row=3, column=1).value == "0856002"


def test_template_uses_first_sub_type_in_first_example(service):
    # Pin: first example row uses sub_types[0]. Pin so admins see
    # a real value.
    buf = service.generate_template(["nstc", "moe_1w"])
    wb = load_workbook(buf)
    ws = wb["白名單匯入模板"]
    assert ws.cell(row=2, column=3).value == "nstc"


def test_template_handles_single_sub_type(service):
    # Pin: when only one sub_type, second example row reuses it
    # (instead of crashing on index 1).
    buf = service.generate_template(["only_one"])
    wb = load_workbook(buf)
    ws = wb["白名單匯入模板"]
    assert ws.cell(row=2, column=3).value == "only_one"
    assert ws.cell(row=3, column=3).value == "only_one"


def test_template_handles_empty_sub_types_falls_back_to_general(service):
    # Pin: empty sub_types list → "general" placeholder. Pin so
    # the template still renders sensibly when admin hasn't defined
    # any sub-types yet.
    buf = service.generate_template([])
    wb = load_workbook(buf)
    ws = wb["白名單匯入模板"]
    assert ws.cell(row=2, column=3).value == "general"


def test_template_includes_instructions_sheet(service):
    # Pin: 使用說明 sheet exists with the documented rules.
    buf = service.generate_template(["nstc"])
    wb = load_workbook(buf)
    assert "使用說明" in wb.sheetnames

    ws = wb["使用說明"]
    text = "\n".join(ws.cell(row=r, column=1).value or "" for r in range(1, 20))
    assert "必填欄位：學號、子獎學金類型" in text
    assert "選填欄位：姓名、備註" in text


def test_template_lists_valid_sub_types_in_instructions(service):
    # Pin: each sub_type appears in the instruction sheet so admins
    # know which values are accepted.
    buf = service.generate_template(["nstc", "moe_1w", "moe_2w"])
    wb = load_workbook(buf)
    ws = wb["使用說明"]
    text = "\n".join(ws.cell(row=r, column=1).value or "" for r in range(1, 25))
    for st in ("nstc", "moe_1w", "moe_2w"):
        assert st in text


# ─── export_whitelist ───────────────────────────────────────────────


def test_export_flattens_nested_dict_to_rows(service):
    # Pin: dict of {sub_type: [students]} → flat rows. Each
    # student/sub_type combo gets one row.
    data = {
        "nstc": [
            {"nycu_id": "0856001", "name": "王小明", "note": "n1"},
            {"nycu_id": "0856002", "name": "李小華", "note": ""},
        ],
        "moe_1w": [
            {"nycu_id": "0856003", "name": "陳大文", "note": ""},
        ],
    }
    buf = service.export_whitelist(data, scholarship_name="獎學金A")
    wb = load_workbook(buf)
    ws = wb["白名單"]
    # 3 data rows after the header row → row 2..4
    rows = [
        (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
        for r in range(2, 5)
    ]
    assert ("0856001", "王小明", "nstc") in rows
    assert ("0856002", "李小華", "nstc") in rows
    assert ("0856003", "陳大文", "moe_1w") in rows


def test_export_empty_dict_yields_only_header(service):
    # Pin: empty whitelist → only the header row, no data rows.
    buf = service.export_whitelist({}, scholarship_name="獎學金A")
    wb = load_workbook(buf)
    ws = wb["白名單"]
    # Row 2 must be empty
    assert ws.cell(row=2, column=1).value is None


def test_export_missing_optional_fields_renders_empty_not_none_text(service):
    # Pin: student dict without "name" / "note" keys still produces
    # a valid row (no "None" rendered). openpyxl normalizes the
    # written empty string to None on reload — pin the spec: the
    # cell does NOT contain the literal text "None".
    data = {"nstc": [{"nycu_id": "0856001"}]}
    buf = service.export_whitelist(data, scholarship_name="獎學金A")
    wb = load_workbook(buf)
    ws = wb["白名單"]
    assert ws.cell(row=2, column=1).value == "0856001"
    # Empty/None is OK; "None" literal text is NOT.
    assert ws.cell(row=2, column=2).value in (None, "")
    assert ws.cell(row=2, column=4).value in (None, "")
