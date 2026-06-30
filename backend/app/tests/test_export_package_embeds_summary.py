"""Integration test: generate_export_zip embeds the 申請總表 workbooks and the
table rows match the student folders. Avoids DB / MinIO / reportlab font by
monkeypatching ensure_cjk_font, _query_applications, _get_scholarship_type,
_generate_summary_pdf, and load_export_aux_data.
"""

import io
import zipfile
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.models.application import Application
from app.services.export_package_service import ExportPackageService


def _mk_app(app_id, user_id, dep_no, dep_name, std_code, cname, academy="某學院"):
    a = Application(
        user_id=user_id,
        scholarship_type_id=1,
        academic_year=114,
        student_data={
            "trm_depno": dep_no,
            "trm_depname": dep_name,
            "trm_academyname": academy,
            "std_stdcode": std_code,
            "std_cname": cname,
        },
    )
    a.id = app_id
    a.files = []  # no uploaded files → no MinIO access
    return a


def _coro_returning(value):
    async def _inner(*args, **kwargs):
        return value

    return _inner


@pytest.mark.asyncio
async def test_export_zip_contains_summary_tables_matching_folders(monkeypatch):
    monkeypatch.setattr("app.services.export_package_service.ensure_cjk_font", lambda: None)

    async def _fake_aux(db, *, scholarship_type, applications):
        return ([], {}, {}, {})

    monkeypatch.setattr("app.services.export_summary_tables.load_export_aux_data", _fake_aux)

    apps = [
        _mk_app(1, 11, "1000", "A系", "001", "甲"),
        _mk_app(2, 12, "1000", "A系", "002", "乙"),
        _mk_app(3, 13, "2000", "B系", "003", "丙"),
    ]
    stype = SimpleNamespace(name="某獎學金", code="phd", sub_type_configs=[])

    svc = ExportPackageService(db=None, minio_service=None)
    monkeypatch.setattr(svc, "_get_scholarship_type", _coro_returning(stype))
    monkeypatch.setattr(svc, "_query_applications", _coro_returning(apps))
    monkeypatch.setattr(svc, "_generate_summary_pdf", lambda *a, **k: b"%PDF-1.4 fake")

    buf, fname = await svc.generate_export_zip(
        scholarship_type_id=1,
        academic_year=114,
        semester="first",
        college_code="A",
    )

    names = zipfile.ZipFile(buf).namelist()

    # College-level table at ZIP root
    assert "114學年度某獎學金學生資料彙整表_某學院.xlsx" in names
    # Per-department tables inside each dept folder
    assert "1000_A系/114學年度某獎學金學生資料彙整表_A系.xlsx" in names
    assert "2000_B系/114學年度某獎學金學生資料彙整表_B系.xlsx" in names

    # Consistency: A系 table 學號 set == A系 student folders
    zf = zipfile.ZipFile(buf)
    a_xlsx = zf.read("1000_A系/114學年度某獎學金學生資料彙整表_A系.xlsx")
    wb = load_workbook(io.BytesIO(a_xlsx))
    ws = wb.active
    table_codes = {ws.cell(row=r, column=13).value for r in range(3, ws.max_row + 1)}
    folder_codes = {
        n.split("/")[1].split("_")[0]
        for n in names
        if n.startswith("1000_A系/") and n.count("/") == 2  # dept/student/file
    }
    assert table_codes == folder_codes == {"001", "002"}


@pytest.mark.asyncio
async def test_export_zip_degrades_when_summary_build_fails_wholesale(monkeypatch):
    monkeypatch.setattr("app.services.export_package_service.ensure_cjk_font", lambda: None)

    async def _boom(*a, **k):
        raise RuntimeError("aux DB down")

    # Wholesale failure of the summary-table builder (e.g. an aux-data DB error
    # before the per-table try/except) must not lose the materials ZIP.
    monkeypatch.setattr("app.services.export_package_service.build_embedded_summary_tables", _boom)

    apps = [_mk_app(1, 11, "1000", "A系", "001", "甲")]
    stype = SimpleNamespace(name="某獎學金", code="phd", sub_type_configs=[])

    svc = ExportPackageService(db=None, minio_service=None)
    monkeypatch.setattr(svc, "_get_scholarship_type", _coro_returning(stype))
    monkeypatch.setattr(svc, "_query_applications", _coro_returning(apps))
    monkeypatch.setattr(svc, "_generate_summary_pdf", lambda *a, **k: b"%PDF-1.4 fake")

    buf, fname = await svc.generate_export_zip(
        scholarship_type_id=1,
        academic_year=114,
        semester="first",
        college_code="A",
    )

    names = zipfile.ZipFile(buf).namelist()
    # Materials ZIP still produced (the student's summary PDF is present)
    assert any(n.endswith("_學生資料彙整.pdf") for n in names)
    # Wholesale failure degraded to a root error placeholder, not an aborted export
    assert "_錯誤_申請總表生成失敗.txt" in names
    # No summary workbooks were emitted
    assert not any(n.endswith(".xlsx") for n in names)
