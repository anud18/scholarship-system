"""
Row-mapping tests for `ExcelExportService._prepare_excel_data`.

`_prepare_excel_data` (excel_export_service.py ~381) translates a list
of persisted `PaymentRosterItem` rows into a list of dicts keyed by the
30-column STD_UP_MIXLISTA template header names. Those dicts are then
written verbatim into each payment Excel row by `_create_excel_file`.

The mapping is the contract between the database model and what shows
up on the bank-payment voucher. A regression here:

- Routes payments to wrong accounts (bank_code 700 hardcoded for postal).
- Reports wrong amounts to the tax-exempt-payment column (immediate
  audit issue with the accounting system).
- Puts the student number into the 身分證字號 column instead of the
  national ID — PR #819 fixed the upstream `_create_roster_item` to
  store std_pid; this file pins that downstream the row dict's
  身分證字號 entry maps to `item.student_id_number` so the fix
  propagates end-to-end.

Tests use SimpleNamespace stubs (no DB) following the pattern in
`test_college_ranking_export_service.py`. `_prepare_excel_data` writes
back to `item.excel_row_data` and `item.excel_remarks` — that's fine
because the stubs accept arbitrary attribute assignment.

Also covers the header-only-file branch of `_validate_export_data`
(returns is_valid=True + a warning, NOT an error) so a roster with no
qualified items still ships a valid Excel header for finance to
acknowledge.
"""

from __future__ import annotations

from types import SimpleNamespace
from typing import Any, Optional

import pytest

from app.services.excel_export_service import ExcelExportService

# ---------------------------------------------------------------------------
# Fixtures / builders
# ---------------------------------------------------------------------------


@pytest.fixture
def service() -> ExcelExportService:
    """`ExcelExportService.__init__` reads template path settings and
    tries to load the .xlsx template file; both fall back gracefully if
    missing. `_prepare_excel_data` itself is pure (just maps dicts) — no
    DB, no filesystem."""
    return ExcelExportService()


def _make_roster(
    *,
    period_label: str = "2025-H1",
    roster_code: str = "ROSTER-114-2025-H1-PHD001",
    academic_year: int = 114,
) -> SimpleNamespace:
    """Minimal roster stub. `_prepare_excel_data` reads `period_label`
    via `hasattr` + `roster.period_label` for the remarks column."""
    return SimpleNamespace(
        period_label=period_label,
        roster_code=roster_code,
        academic_year=academic_year,
    )


def _make_item(
    *,
    student_id_number: str = "A123456789",
    student_name: str = "王小明",
    student_email: str = "wang@nycu.edu.tw",
    bank_account: Optional[str] = "00012345678",
    scholarship_name: str = "博士班獎學金",
    scholarship_amount: Any = 50000,
    permanent_address: Optional[str] = None,
    application_identity: Optional[str] = "114新申請",
    allocated_sub_type: Optional[str] = None,
    allocation_year: Optional[int] = None,
    is_included: bool = True,
    verification_status: str = "verified",
    is_eligible: Optional[bool] = True,
    rule_validation_result: Optional[dict] = None,
    exclusion_reason: Optional[str] = None,
) -> SimpleNamespace:
    """PaymentRosterItem stub with just the attrs _prepare_excel_data reads.

    Also exposes `excel_row_data` + `excel_remarks` as plain attrs so the
    method's write-back assignments don't AttributeError."""
    return SimpleNamespace(
        student_id_number=student_id_number,
        student_name=student_name,
        student_email=student_email,
        bank_account=bank_account,
        scholarship_name=scholarship_name,
        scholarship_amount=scholarship_amount,
        permanent_address=permanent_address,
        application_identity=application_identity,
        allocated_sub_type=allocated_sub_type,
        allocation_year=allocation_year,
        is_included=is_included,
        verification_status=verification_status,
        is_eligible=is_eligible,
        rule_validation_result=rule_validation_result,
        exclusion_reason=exclusion_reason,
        # Write-back targets — pre-init so SimpleNamespace accepts assignment
        excel_row_data=None,
        excel_remarks=None,
    )


# ---------------------------------------------------------------------------
# 身分證字號 column ← item.student_id_number (PR #819 propagation)
# ---------------------------------------------------------------------------


def test_row_data_national_id_comes_from_student_id_number(service: ExcelExportService) -> None:
    """Pin: the "身分證字號" key in row_data maps to `item.student_id_number`.

    Upstream (_create_roster_item, PR #819) stores std_pid into
    student_id_number. If THIS mapping later drifts to a different attr
    (e.g. someone introduces item.national_id and forgets to migrate
    here), payments would route by an inconsistent identifier."""
    roster = _make_roster()
    item = _make_item(student_id_number="A987654321", student_name="陳大文")

    data, _ = service._prepare_excel_data(roster, [item])

    assert len(data) == 1
    assert data[0]["身分證字號"] == "A987654321"
    assert data[0]["姓名"] == "陳大文"


def test_row_data_skips_items_missing_required_fields(service: ExcelExportService) -> None:
    """Items missing student_id_number OR student_name are silently
    dropped (logged + continued). They never reach the Excel because
    the bank rejects rows with blank 身分證字號. Pin so this defense
    doesn't get removed accidentally."""
    roster = _make_roster()
    items = [
        _make_item(student_id_number="", student_name="王小明"),  # missing ID
        _make_item(student_id_number="A1", student_name=""),  # missing name
        _make_item(student_id_number="A2", student_name="正常"),  # ok
    ]

    data, _ = service._prepare_excel_data(roster, items)

    assert len(data) == 1
    assert data[0]["身分證字號"] == "A2"
    assert data[0]["姓名"] == "正常"


# ---------------------------------------------------------------------------
# Amount columns — 單價 + 免稅給付 both pull from scholarship_amount
# ---------------------------------------------------------------------------


def test_row_data_amount_matches_scholarship_amount(service: ExcelExportService) -> None:
    """Both 單價 (unit price, col 10) AND 免稅給付 (tax-exempt payment,
    col 24) equal `item.scholarship_amount` as a float. They MUST agree
    — the accounting system reconciles col 10 × col 9 (quantity, always
    1) against col 24 for tax-exempt verification. Drift here triggers
    an audit kickback."""
    roster = _make_roster()
    item = _make_item(scholarship_amount=50000)

    data, _ = service._prepare_excel_data(roster, [item])

    assert data[0]["單價"] == 50000.0
    assert data[0]["免稅給付"] == 50000.0
    # Both are floats (not Decimal/int) — finance Excel parsing expects float
    assert isinstance(data[0]["單價"], float)
    assert isinstance(data[0]["免稅給付"], float)


def test_row_data_amount_zero_when_scholarship_amount_falsy(service: ExcelExportService) -> None:
    """Defensive: a None/0 scholarship_amount renders as 0 (not None,
    not blank, not crash). The downstream validator catches this as a
    warning rather than letting a blank propagate to the bank file."""
    roster = _make_roster()
    item = _make_item(scholarship_amount=None)

    data, _ = service._prepare_excel_data(roster, [item])

    assert data[0]["單價"] == 0
    assert data[0]["免稅給付"] == 0


# ---------------------------------------------------------------------------
# Fixed-value columns — bank code, identity, unit, etc.
# ---------------------------------------------------------------------------


def test_row_data_bank_code_always_700(service: ExcelExportService) -> None:
    """銀行代碼 (col 4) is HARDCODED to "700" — the Chunghwa Post code.
    This system only pays to postal accounts. If a future feature adds
    real-bank payments, this hardcode MUST be unwound deliberately
    (and this test updated). Pin so it can't drift silently."""
    roster = _make_roster()
    item = _make_item()

    data, _ = service._prepare_excel_data(roster, [item])

    assert data[0]["銀行代碼"] == "700"


def test_row_data_fixed_columns_contract(service: ExcelExportService) -> None:
    """Pin the small set of fixed-value columns the accounting system
    expects EXACTLY: 職別="學生", 身份別代碼="1", 單位="次", 數量="1",
    個人身分別="1" (本國人), 居留天數="是". Any drift triggers
    an audit kickback from the bank batch import."""
    roster = _make_roster()
    item = _make_item()

    data, _ = service._prepare_excel_data(roster, [item])

    row = data[0]
    assert row["職別(稱)"] == "學生"
    assert row["身份別代碼"] == "1"
    assert row["單位(ex:時,月,次...)"] == "次"
    assert row["數量"] == "1"
    assert row["個人身分別(1:本國人,2:外國人,3:大陸人)"] == "1"
    assert row["居留天數是否滿183天(是/否)"] == "是"


def test_row_data_bank_account_passed_through_from_item(service: ExcelExportService) -> None:
    """帳號 column (col 3) comes directly from item.bank_account (which
    upstream _create_roster_item populated from submitted_form_data, not
    student_data). Falsy bank_account → empty string, not None."""
    roster = _make_roster()
    with_account = _make_item(bank_account="00012345678")
    no_account = _make_item(student_id_number="A2", bank_account=None)

    data, _ = service._prepare_excel_data(roster, [with_account, no_account])

    assert data[0]["帳號"] == "00012345678"
    assert data[1]["帳號"] == ""


# ---------------------------------------------------------------------------
# Remarks (說明 col 25) — composed from period + scholarship + identity
# ---------------------------------------------------------------------------


def test_row_data_remarks_includes_period_and_scholarship(service: ExcelExportService) -> None:
    """說明 column is a space-joined breadcrumb of period / scholarship
    name / identity / allocation. Pin the core pieces so finance can
    grep for a known scholarship name in the payment report."""
    roster = _make_roster(period_label="2025-H1")
    item = _make_item(
        scholarship_name="博士班獎學金",
        application_identity="114新申請",
        allocated_sub_type="nstc",
        allocation_year=114,
    )

    data, _ = service._prepare_excel_data(roster, [item])

    remarks = data[0]["說明"]
    assert "期間:2025-H1" in remarks
    assert "獎學金:博士班獎學金" in remarks
    assert "身分:114新申請" in remarks
    # nstc → 國科會 sub_type display
    assert "分發:114年 國科會" in remarks


def test_row_data_remarks_flags_excluded_status(service: ExcelExportService) -> None:
    """An excluded item (is_included=False) gets 狀態:不合格 appended
    so the operator scanning the Excel can see at a glance which rows
    were dropped. Also flags missing bank info separately."""
    roster = _make_roster()
    item = _make_item(is_included=False, bank_account=None)

    data, _ = service._prepare_excel_data(roster, [item])

    remarks = data[0]["說明"]
    assert "狀態:不合格" in remarks
    assert "缺銀行資訊" in remarks


# ---------------------------------------------------------------------------
# Write-back side effect — item.excel_row_data and item.excel_remarks
# ---------------------------------------------------------------------------


def test_row_data_written_back_to_item(service: ExcelExportService) -> None:
    """`_prepare_excel_data` mutates `item.excel_row_data` and
    `item.excel_remarks` for downstream audit / preview. Pin so a
    refactor that "returns the rows but doesn't write back" doesn't
    break the audit trail silently (the preview API reads these
    columns to show what was emitted)."""
    roster = _make_roster()
    item = _make_item()

    data, _ = service._prepare_excel_data(roster, [item])

    assert item.excel_row_data == data[0]
    assert item.excel_remarks == data[0]["說明"]


# ---------------------------------------------------------------------------
# Empty roster → _validate_export_data returns is_valid=True with warning
# ---------------------------------------------------------------------------


def test_empty_roster_validates_as_header_only(service: ExcelExportService) -> None:
    """An empty roster (no qualified items) is NOT a hard error — the
    Excel still ships with just headers so finance can acknowledge the
    cycle ran to completion. Pin the contract: is_valid=True, no errors,
    one specific warning naming the header-only case.

    Without this branch a 0-qualified roster would fail validation and
    the whole roster generation would fall back to RosterGenerationError
    even though the upstream logic correctly identified zero qualified
    applicants — a false-positive failure."""
    result = service._validate_export_data([])

    assert result["is_valid"] is True
    assert result["errors"] == []
    assert any("No roster items to export" in w for w in result["warnings"])


def test_empty_roster_prepare_excel_data_returns_empty_list(service: ExcelExportService) -> None:
    """Companion: `_prepare_excel_data(roster, [])` returns an empty list,
    NOT None, so `_create_excel_file` writes a header-only sheet without
    a TypeError on `len(excel_data)`."""
    roster = _make_roster()

    data, _ = service._prepare_excel_data(roster, [])

    assert data == []


# ---------------------------------------------------------------------------
# Roster item scope filter — auto-excluded shown, manual removals hidden
# ---------------------------------------------------------------------------


def _mk_scope_item(name, *, is_included, exclusion_reason=None):
    it = _make_item(student_name=name, is_included=is_included)
    it.exclusion_reason = exclusion_reason
    return it


def test_is_manual_removal_detects_lock_and_reconcile_prefixes(service):
    assert (
        service._is_manual_removal(_mk_scope_item("a", is_included=False, exclusion_reason="鎖定後移除[停發]：x"))
        is True
    )
    assert (
        service._is_manual_removal(
            _mk_scope_item("b", is_included=False, exclusion_reason="比對分發移除：不在分發名單")
        )
        is True
    )
    assert (
        service._is_manual_removal(_mk_scope_item("c", is_included=False, exclusion_reason="學籍驗證未通過: graduated"))
        is False
    )
    assert service._is_manual_removal(_mk_scope_item("d", is_included=True, exclusion_reason=None)) is False


def test_get_roster_items_default_keeps_auto_excluded_hides_manual(service):
    included = _mk_scope_item("納入", is_included=True)
    auto_excluded = _mk_scope_item("自動排除", is_included=False, exclusion_reason="學籍驗證未通過: graduated")
    manual = _mk_scope_item("手動移除", is_included=False, exclusion_reason="鎖定後移除[停發]：x")
    roster = SimpleNamespace(items=[included, auto_excluded, manual])

    names = [i.student_name for i in service._get_roster_items(roster, include_excluded=False)]

    assert "納入" in names
    assert "自動排除" in names
    assert "手動移除" not in names


def test_get_roster_items_include_excluded_shows_everything(service):
    included = _mk_scope_item("納入", is_included=True)
    manual = _mk_scope_item("手動移除", is_included=False, exclusion_reason="鎖定後移除[停發]：x")
    roster = SimpleNamespace(items=[included, manual])

    names = [i.student_name for i in service._get_roster_items(roster, include_excluded=True)]

    assert names == ["手動移除", "納入"]  # sorted by student_name


# ---------------------------------------------------------------------------
# Rule column collection from frozen rule_validation_result snapshots
# ---------------------------------------------------------------------------


def _item_with_rules(name, details):
    """details: dict like {"rule_5": {"passed": True, "rule_name": "GPA", ...}}"""
    it = _make_item(student_name=name)
    it.rule_validation_result = {"is_eligible": True, "details": details}
    return it


def test_collect_rule_columns_orders_by_rule_id_and_dedupes(service):
    items = [
        _item_with_rules(
            "a",
            {
                "rule_5": {"passed": True, "rule_name": "GPA門檻"},
                "rule_2": {"passed": False, "rule_name": "國籍"},
            },
        ),
        _item_with_rules("b", {"rule_2": {"passed": True, "rule_name": "國籍"}, "no_rules_found": True}),
    ]

    cols = service._collect_rule_columns(items)

    assert cols == [(2, "國籍"), (5, "GPA門檻")]


def test_collect_rule_columns_disambiguates_duplicate_names(service):
    items = [
        _item_with_rules(
            "a",
            {
                "rule_2": {"passed": True, "rule_name": "門檻"},
                "rule_7": {"passed": True, "rule_name": "門檻"},
            },
        )
    ]

    cols = service._collect_rule_columns(items)

    assert cols == [(2, "門檻"), (7, "門檻（#7）")]


def test_collect_rule_columns_ignores_items_without_snapshot(service):
    plain = _make_item(student_name="x")
    plain.rule_validation_result = None
    assert service._collect_rule_columns([plain]) == []


def test_build_export_columns_layout(service):
    cols = service._build_export_columns([(2, "國籍"), (5, "GPA門檻")])

    base = list(service.template_columns)
    assert cols == base + ["學籍驗證", "規則資格", "國籍", "GPA門檻", "納入造冊", "排除原因"]


# ---------------------------------------------------------------------------
# Verification columns + per-rule columns + cell fill metadata
# ---------------------------------------------------------------------------


def test_prepare_returns_rows_and_parallel_fills(service):
    roster = _make_roster()
    item = _make_item()
    rows, fills = service._prepare_excel_data(roster, [item], [])
    assert len(rows) == len(fills) == 1


def test_verification_and_rule_columns_values(service):
    roster = _make_roster()
    item = _make_item(student_name="甲")
    item.verification_status = "verified"
    item.is_eligible = True
    item.rule_validation_result = {
        "is_eligible": True,
        "details": {
            "rule_2": {"passed": True, "rule_name": "國籍", "is_hard_rule": True},
            "rule_5": {"passed": False, "rule_name": "GPA門檻", "is_hard_rule": True},
            "rule_8": {"passed": False, "rule_name": "在學", "is_warning": True},
        },
    }
    rule_columns = service._collect_rule_columns([item])  # [(2,國籍),(5,GPA門檻),(8,在學)]

    rows, fills = service._prepare_excel_data(roster, [item], rule_columns)

    r, f = rows[0], fills[0]
    assert r["學籍驗證"] == "已驗證"
    assert r["規則資格"] == "符合"
    assert r["國籍"] == "通過"
    assert r["GPA門檻"] == "未通過"
    assert r["在學"] == "未通過"
    assert r["納入造冊"] == "是"
    # fills: hard fail → red, warning fail → amber, passed → no entry
    assert "國籍" not in f
    assert f["GPA門檻"] == "red"
    assert f["在學"] == "amber"


def test_unverified_and_excluded_and_missing_bank_fills_red(service):
    roster = _make_roster()
    item = _make_item(student_name="乙", bank_account=None, is_included=False)
    item.verification_status = "graduated"
    item.is_eligible = False
    item.exclusion_reason = "學籍驗證未通過: graduated"

    rows, fills = service._prepare_excel_data(roster, [item], [])

    r, f = rows[0], fills[0]
    assert r["學籍驗證"] == "已畢業"
    assert r["規則資格"] == "不符合"
    assert r["納入造冊"] == "否"
    assert r["排除原因"] == "學籍驗證未通過: graduated"
    assert f["學籍驗證"] == "red"
    assert f["規則資格"] == "red"
    assert f["帳號"] == "red"
    assert f["納入造冊"] == "red"
    assert f["排除原因"] == "red"


def test_untagged_failed_rule_defaults_to_red(service):
    """An errored/untagged rule result (passed=False but no is_hard_rule/
    is_warning, e.g. the exception path in _evaluate_scholarship_rule) must
    still color the cell — defaults to red, never left uncolored."""
    roster = _make_roster()
    item = _make_item(student_name="戊")
    item.is_eligible = False
    item.rule_validation_result = {
        "is_eligible": False,
        "details": {"rule_4": {"passed": False, "rule_name": "錯誤規則"}},
    }
    rule_columns = service._collect_rule_columns([item])

    rows, fills = service._prepare_excel_data(roster, [item], rule_columns)

    assert rows[0]["錯誤規則"] == "未通過"
    assert fills[0]["錯誤規則"] == "red"


def test_no_snapshot_item_renders_dashes_without_fill(service):
    roster = _make_roster()
    item = _make_item(student_name="丙")
    item.is_eligible = None
    item.rule_validation_result = None

    rows, fills = service._prepare_excel_data(roster, [item], [(2, "國籍")])

    assert rows[0]["規則資格"] == "—"
    assert rows[0]["國籍"] == "—"
    assert "規則資格" not in fills[0]
    assert "國籍" not in fills[0]


def test_excel_row_data_written_back_is_clean(service):
    """write-back 不含上色 metadata（fills 走平行回傳，不污染 DB JSON）。"""
    roster = _make_roster()
    item = _make_item(student_name="丁")
    rows, _ = service._prepare_excel_data(roster, [item], [])
    assert item.excel_row_data == rows[0]
    assert "__cell_fills__" not in item.excel_row_data


# ---------------------------------------------------------------------------
# End-to-end Excel write — headers, per-rule columns, red/amber cell fills
# ---------------------------------------------------------------------------


def test_create_excel_file_writes_headers_and_fills(service, tmp_path):
    import openpyxl

    roster = _make_roster()
    item = _make_item(student_name="甲", bank_account=None, is_included=False)
    item.verification_status = "withdrawn"
    item.is_eligible = False
    item.exclusion_reason = "學籍驗證未通過: withdrawn"
    item.rule_validation_result = {
        "is_eligible": False,
        "details": {"rule_3": {"passed": False, "rule_name": "GPA", "is_hard_rule": True}},
    }

    rule_columns = service._collect_rule_columns([item])
    columns = service._build_export_columns(rule_columns)
    rows, fills = service._prepare_excel_data(roster, [item], rule_columns)

    out = tmp_path / "roster.xlsx"
    service._create_excel_file(
        rows,
        fills,
        str(out),
        roster,
        template_path="/nonexistent.xlsx",
        columns=columns,
        include_header=True,
        include_statistics=False,
    )

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [ws.cell(row=1, column=c).value for c in range(1, len(columns) + 1)]
    assert header[-5:] == ["學籍驗證", "規則資格", "GPA", "納入造冊", "排除原因"]

    def fill_of(col_name):
        c = columns.index(col_name) + 1
        return ws.cell(row=2, column=c).fill.start_color.rgb

    assert "FFC7CE" in str(fill_of("學籍驗證"))  # red
    assert "FFC7CE" in str(fill_of("GPA"))  # red (hard fail)
    assert "FFC7CE" in str(fill_of("帳號"))  # red (missing bank)
    assert "FFC7CE" in str(fill_of("納入造冊"))  # red
    # 姓名欄不應上色
    assert "FFC7CE" not in str(fill_of("姓名"))


def test_create_excel_file_amber_for_warning_rule(service, tmp_path):
    import openpyxl

    roster = _make_roster()
    item = _make_item(student_name="乙")
    item.rule_validation_result = {
        "is_eligible": True,
        "details": {"rule_9": {"passed": False, "rule_name": "在學狀態", "is_warning": True}},
    }
    rule_columns = service._collect_rule_columns([item])
    columns = service._build_export_columns(rule_columns)
    rows, fills = service._prepare_excel_data(roster, [item], rule_columns)

    out = tmp_path / "roster_amber.xlsx"
    service._create_excel_file(
        rows,
        fills,
        str(out),
        roster,
        template_path="/nonexistent.xlsx",
        columns=columns,
        include_header=True,
        include_statistics=False,
    )

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    c = columns.index("在學狀態") + 1
    assert "FFEB9C" in str(ws.cell(row=2, column=c).fill.start_color.rgb)  # amber
