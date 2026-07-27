"""
Tests for the 匯入已領月份數 example workbook.

The point of these is anti-drift: the template is fed straight back through the
real parser, so a change to either side that breaks the other fails here rather
than in an admin's face.
"""

from io import BytesIO

import openpyxl
import pytest

from app.services.received_months_parser import (
    EXAMPLE_ROW_MARKER,
    HEADER_CURRENT_MONTH,
    HEADER_START_MONTH,
    HEADER_STATED_TOTAL,
    HEADER_STUDENT_NUMBER,
    parse_received_months_workbook,
)
from app.services.received_months_template import (
    BLANK_ROW_COUNT,
    EXAMPLE_ROW,
    SHEET_TITLE,
    TEMPLATE_COLUMNS,
    WORKBOOK_TITLE,
    build_received_months_template,
)


def _rows(content: bytes):
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    return rows


@pytest.fixture(scope="module")
def template_bytes() -> bytes:
    return build_received_months_template()


class TestShape:
    def test_sheet_is_titled_in_zh_tw(self, template_bytes):
        workbook = openpyxl.load_workbook(BytesIO(template_bytes))
        assert workbook.active.title == SHEET_TITLE
        workbook.close()

    def test_row_1_is_the_merged_title(self, template_bytes):
        assert _rows(template_bytes)[0][0] == WORKBOOK_TITLE

    def test_row_2_is_the_header(self, template_bytes):
        assert list(_rows(template_bytes)[1]) == TEMPLATE_COLUMNS

    def test_row_3_is_the_example(self, template_bytes):
        # openpyxl reads a written "" back as None, so compare on text.
        written = ["" if c is None else c for c in _rows(template_bytes)[2]]
        assert written == EXAMPLE_ROW

    def test_blank_rows_are_pre_numbered(self, template_bytes):
        rows = _rows(template_bytes)[3:]
        assert len(rows) == BLANK_ROW_COUNT
        assert [r[0] for r in rows] == list(range(1, BLANK_ROW_COUNT + 1))
        # Every other cell empty — nothing for the admin to clear out.
        assert all(all(c in ("", None) for c in r[1:]) for r in rows)

    def test_header_and_title_stay_visible_when_scrolling(self, template_bytes):
        workbook = openpyxl.load_workbook(BytesIO(template_bytes))
        assert workbook.active.freeze_panes == "A3"
        workbook.close()


class TestHeaderContract:
    """The columns the parser derives from must be present, by exact name."""

    @pytest.mark.parametrize(
        "header",
        [HEADER_STUDENT_NUMBER, HEADER_START_MONTH, HEADER_CURRENT_MONTH, HEADER_STATED_TOTAL],
    )
    def test_parser_relevant_header_present(self, header):
        assert header in TEMPLATE_COLUMNS

    def test_no_duplicate_headers(self):
        assert len(TEMPLATE_COLUMNS) == len(set(TEMPLATE_COLUMNS))


class TestRoundTripThroughTheParser:
    """The template must be readable by the very parser it feeds."""

    def test_blank_template_reports_no_importable_rows(self, template_bytes):
        # The 範例 row and the pre-numbered blanks are all skipped, so an
        # untouched template has nothing to import — and says so clearly.
        with pytest.raises(ValueError, match="沒有可匯入的資料列"):
            parse_received_months_workbook(_rows(template_bytes))

    def test_filled_template_parses_with_the_derived_month_count(self, template_bytes):
        workbook = openpyxl.load_workbook(BytesIO(template_bytes))
        sheet = workbook.active
        # Fill the first blank row (row 4) the way an admin would.
        for column, value in enumerate(
            ["1", "電機學院", "電機工程學系", "310460031", "陳小明", "113年9月", "115年8月", "116年8月", 24, "", ""],
            start=1,
        ):
            sheet.cell(row=4, column=column, value=value)
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        result = parse_received_months_workbook(_rows(buffer.getvalue()))

        assert len(result.rows) == 1, "the 範例 row must not be imported alongside it"
        parsed = result.rows[0]
        assert parsed.student_number == "310460031"
        assert parsed.months == 24
        assert parsed.error is None
        assert parsed.warning is None

    def test_the_example_row_is_never_imported(self, template_bytes):
        workbook = openpyxl.load_workbook(BytesIO(template_bytes))
        sheet = workbook.active
        sheet.cell(row=4, column=4, value="310460031")
        sheet.cell(row=4, column=6, value="114年1月")
        sheet.cell(row=4, column=7, value="114年3月")
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        result = parse_received_months_workbook(_rows(buffer.getvalue()))

        assert [r.student_number for r in result.rows] == ["310460031"]
        assert EXAMPLE_ROW_MARKER not in [r.student_number for r in result.rows]

    def test_the_examples_own_numbers_are_self_consistent(self, template_bytes):
        """國科會's demo row states 24 months for 113年9月→115年8月.

        If the stated total and the derived span ever disagreed, the template
        would be teaching admins a number the parser would warn about.
        """
        headers = TEMPLATE_COLUMNS
        example = dict(zip(headers, EXAMPLE_ROW))

        # Promote the demo row to a real one. Both the 學號 and the NO cell have
        # to change: the parser skips a row carrying 範例 in ANY cell, which is
        # exactly why the demo row is safe to ship inside the template.
        def promote(header):
            if header == HEADER_STUDENT_NUMBER:
                return "310460031"
            value = example[header]
            return 1 if value == EXAMPLE_ROW_MARKER else value

        rows = [
            (WORKBOOK_TITLE,) + (None,) * (len(headers) - 1),
            tuple(headers),
            tuple(promote(h) for h in headers),
        ]
        result = parse_received_months_workbook(rows)

        assert result.rows[0].months == example[HEADER_STATED_TOTAL]
        assert result.rows[0].warning is None
