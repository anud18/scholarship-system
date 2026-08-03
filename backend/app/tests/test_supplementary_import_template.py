"""補充匯入 範本 tests.

補充匯入 replaced 批次匯入 for colleges, so the file a college downloads must be the
file the admin panel downloads — same columns, same sub-type checkmarks, same
advisor fields. These tests pin that the two really are one artefact, and that the
downloaded file round-trips through the importer that consumes it.
"""

import io

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.application_field import ApplicationField
from app.models.scholarship import ScholarshipConfiguration, ScholarshipType, SubTypeSelectionMode
from app.services.batch_import_service import SUB_TYPE_CODE_BY_LABEL
from app.services.batch_import_template_service import build_batch_import_template
from app.services.supplementary_import_service import SupplementaryImportService


@pytest_asyncio.fixture
async def scholarship(db: AsyncSession) -> ScholarshipType:
    s = ScholarshipType(
        code="phd_tpl_test",
        name="範本測試獎學金",
        sub_type_list=["nstc", "moe_1w"],
        sub_type_selection_mode=SubTypeSelectionMode.single,
        status="active",
    )
    db.add(s)
    await db.flush()

    # The advisor columns only appear when some active configuration requires a
    # professor recommendation — seed one so the template under test has them.
    db.add(
        ScholarshipConfiguration(
            scholarship_type_id=s.id,
            academic_year=114,
            config_name="範本測試 114",
            config_code="phd_tpl_test_114",
            amount=30000,
            is_active=True,
            requires_professor_recommendation=True,
        )
    )
    await db.flush()
    return s


def _headers(payload: bytes) -> list:
    ws = load_workbook(io.BytesIO(payload)).active
    return [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]


@pytest.mark.asyncio
class TestTemplateIsSharedWithBatchImport:
    async def test_sub_type_columns_use_the_parser_labels(self, db: AsyncSession, scholarship: ScholarshipType):
        """A label the parser doesn't recognise would make the downloaded template
        un-importable — the generator and SUB_TYPE_CODE_BY_LABEL must agree."""
        headers = _headers(await build_batch_import_template(db, scholarship))
        for code in scholarship.sub_type_list:
            label = {c: lbl for lbl, c in SUB_TYPE_CODE_BY_LABEL.items()}[code]
            assert label in headers

    async def test_required_columns_present(self, db: AsyncSession, scholarship: ScholarshipType):
        headers = _headers(await build_batch_import_template(db, scholarship))
        assert headers[:3] == ["學號", "學生姓名", "郵局帳號"]

    async def test_custom_fields_are_appended(self, db: AsyncSession, scholarship: ScholarshipType):
        db.add(
            ApplicationField(
                scholarship_type=scholarship.code,
                field_name="contact_phone",
                field_label="聯絡電話",
                field_type="text",
                is_active=True,
                display_order=1,
            )
        )
        await db.flush()
        headers = _headers(await build_batch_import_template(db, scholarship))
        assert "聯絡電話" in headers


@pytest.mark.asyncio
class TestTemplateRoundTrip:
    async def test_downloaded_template_parses_through_supplementary_import(
        self, db: AsyncSession, scholarship: ScholarshipType
    ):
        """The template ships two sample rows; feeding it straight back into the
        importer must yield those two students with their sub-types resolved."""
        payload = await build_batch_import_template(db, scholarship)

        service = SupplementaryImportService(db)
        rows, errors = await service.parse_file(
            payload,
            scholarship_type_id=scholarship.id,
            academic_year=114,
            semester=None,
        )

        assert errors == []
        assert [r["student_id"] for r in rows] == ["111111111", "222222222"]
        # Row 1 marks every sub-type, row 2 only the first — the sample data
        # deliberately contrasts 1 and 0 so the checkmark semantics are visible.
        assert rows[0]["sub_types"], "first sample row must carry its marked sub-types"
        assert all(st in scholarship.sub_type_list for st in rows[0]["sub_types"])

    async def test_advisor_employee_id_survives_the_round_trip(self, db: AsyncSession, scholarship: ScholarshipType):
        """指導教授本校人事編號 is what lets 補充匯入 route a brand-new student to a
        professor at all, so it must be in the template AND read back out."""
        from app.services.application_field_service import ApplicationFieldService

        requires_advisor = await ApplicationFieldService(db).check_requires_professor_recommendation(scholarship.code)
        assert requires_advisor, "fixture must require a professor recommendation for this column to exist"

        payload = await build_batch_import_template(db, scholarship)
        assert "指導教授本校人事編號" in _headers(payload)
        rows, errors = await SupplementaryImportService(db).parse_file(
            payload, scholarship_type_id=scholarship.id, academic_year=114, semester=None
        )
        assert errors == []
        assert rows[0]["advisor_nycu_id"]

    async def test_renewal_rows_are_rejected(self, db: AsyncSession, scholarship: ScholarshipType):
        """續領 has its own admin-only import; a 續領年份 column here would otherwise
        mint a brand-new application carrying a renewal year."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        for col, header in enumerate(["學號", "學生姓名", "續領年份", "國科會"], start=1):
            ws.cell(row=1, column=col, value=header)
        for col, value in enumerate(["310460111", "續領生", 113, 1], start=1):
            ws.cell(row=2, column=col, value=value)
        buf = io.BytesIO()
        wb.save(buf)

        _rows, errors = await SupplementaryImportService(db).parse_file(
            buf.getvalue(), scholarship_type_id=scholarship.id, academic_year=114, semester=None
        )
        assert any("續領" in e for e in errors)
        assert any("310460111" in e for e in errors)
