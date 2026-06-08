"""Unit tests for export_summary_tables: pure helpers + build_embedded_summary_tables.

These tests avoid DB / MinIO / reportlab font by monkeypatching load_export_aux_data
and constructing transient Application objects via Application(**defaults) (never
Application.__new__, which leaves _sa_instance_state unset).
"""

import io
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.models.application import Application
from app.services.export_summary_tables import (
    _dept_name_from_apps,
    _sort_key,
    build_embedded_summary_tables,
)


def _mk_app(app_id, user_id, dep_no, dep_name, std_code, cname="某生", academy="某學院"):
    a = Application(
        user_id=user_id,
        scholarship_type_id=1,
        academic_year=114,
        student_data={
            "trm_depno": dep_no,
            "trm_depname": dep_name,
            "trm_academyname": academy,
            "std_stdcode": std_code,
            "std_cname": cname,
        },
    )
    a.id = app_id
    return a


class TestSortKey:
    def test_sorts_by_student_code_blanks_last(self):
        a1 = _mk_app(1, 11, "1000", "A系", "002")
        a2 = _mk_app(2, 12, "1000", "A系", "001")
        a3 = _mk_app(3, 13, "1000", "A系", "")  # blank code → last
        ordered = sorted([a1, a2, a3], key=_sort_key)
        assert [a.id for a in ordered] == [2, 1, 3]


class TestDeptNameFromApps:
    def test_takes_first_non_empty_depname(self):
        apps = [_mk_app(1, 11, "1000", "", "001"), _mk_app(2, 12, "1000", "教育博", "002")]
        assert _dept_name_from_apps(apps) == "教育博"

    def test_falls_back_when_all_blank(self):
        apps = [_mk_app(1, 11, "1000", "", "001")]
        assert _dept_name_from_apps(apps) == "未知系所"


@pytest.fixture
def _patch_aux(monkeypatch):
    async def _fake_aux(db, *, scholarship_type, applications):
        # (dynamic_fields, sub_type_labels, account_by_user, advisor_by_user)
        return ([], {}, {}, {})

    monkeypatch.setattr("app.services.export_summary_tables.load_export_aux_data", _fake_aux)


def _read_col(xlsx_bytes, col_idx):
    """Return values of `col_idx` (1-based) for all data rows (row >= 3).

    openpyxl round-trips empty-string cells as None; normalise to "" so
    assertions can use string literals uniformly.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    return [ws.cell(row=r, column=col_idx).value or "" for r in range(3, ws.max_row + 1)]


@pytest.mark.asyncio
async def test_builds_college_and_per_dept_tables(_patch_aux):
    stype = SimpleNamespace(name="某獎學金", code="phd", sub_type_configs=[])
    dept_groups = {
        "1000_A系": [
            _mk_app(2, 12, "1000", "A系", "002"),
            _mk_app(1, 11, "1000", "A系", "001"),  # deliberately out of order
        ],
        "2000_B系": [_mk_app(3, 13, "2000", "B系", "003")],
    }

    tables = await build_embedded_summary_tables(
        db=None,
        scholarship_type=stype,
        dept_groups=dept_groups,
        college_name="某學院",
        academic_year=114,
    )

    college_key = "114學年度某獎學金學生資料彙整表_某學院.xlsx"
    a_key = "1000_A系/114學年度某獎學金學生資料彙整表_A系.xlsx"
    b_key = "2000_B系/114學年度某獎學金學生資料彙整表_B系.xlsx"
    assert set(tables.keys()) == {college_key, a_key, b_key}

    # Per-dept row counts == dept sizes; codes sorted within dept
    assert _read_col(tables[a_key], 13) == ["001", "002"]
    assert _read_col(tables[b_key], 13) == ["003"]

    # College table: grouped by dept folder order (A系 before B系), sorted within
    assert _read_col(tables[college_key], 13) == ["001", "002", "003"]
    assert _read_col(tables[college_key], 5) == ["A系", "A系", "B系"]  # 系所 column

    # 排序 column (學院初審會議之學院排序, col 2) is empty for every row
    assert _read_col(tables[college_key], 2) == ["", "", ""]


@pytest.mark.asyncio
async def test_college_name_falls_back_to_全校(_patch_aux):
    stype = SimpleNamespace(name="某獎學金", code="phd", sub_type_configs=[])
    dept_groups = {"1000_A系": [_mk_app(1, 11, "1000", "A系", "001")]}
    tables = await build_embedded_summary_tables(
        db=None,
        scholarship_type=stype,
        dept_groups=dept_groups,
        college_name=None,
        academic_year=114,
    )
    assert "114學年度某獎學金學生資料彙整表_全校.xlsx" in tables


@pytest.mark.asyncio
async def test_table_build_failure_writes_error_txt(_patch_aux, monkeypatch):
    def _boom(self, *, rows, dynamic_fields, sub_type_labels, title, sheet_name):
        raise RuntimeError("render exploded")

    monkeypatch.setattr(
        "app.services.export_summary_tables.CollegeRankingExportService.build_workbook",
        _boom,
    )
    stype = SimpleNamespace(name="某獎學金", code="phd", sub_type_configs=[])
    dept_groups = {"1000_A系": [_mk_app(1, 11, "1000", "A系", "001")]}

    tables = await build_embedded_summary_tables(
        db=None,
        scholarship_type=stype,
        dept_groups=dept_groups,
        college_name="某學院",
        academic_year=114,
    )

    # Each failed table becomes an error .txt instead of aborting the whole build
    assert "1000_A系/_錯誤_總表生成失敗.txt" in tables
    assert "_錯誤_學院總表生成失敗.txt" in tables
    assert b"render exploded" in tables["1000_A系/_錯誤_總表生成失敗.txt"]
    # No xlsx produced when every build fails
    assert not any(k.endswith(".xlsx") for k in tables)
