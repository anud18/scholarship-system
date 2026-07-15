"""Integration tests for the college distribution-results Excel/PDF export.

Cross-college isolation is asserted by PARSING the workbook rows — never by
scanning response bytes, which would false-negative on xlsx compression.
"""

import io

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import require_college
from app.main import app
from app.models.application import Application
from app.models.college_review import CollegeRanking, CollegeRankingItem
from app.models.scholarship import (
    ScholarshipConfiguration,
    ScholarshipStatus,
    ScholarshipType,
    SubTypeSelectionMode,
)
from app.models.user import AdminScholarship, User, UserRole, UserType

EXPORT_URL = "/api/v1/college-review/distribution-results/export"


@pytest_asyncio.fixture
async def sch_type(db: AsyncSession) -> ScholarshipType:
    st = ScholarshipType(
        code="cde_phd",
        name="CDE PhD Scholarship",
        description="college-distribution-export test",
        status=ScholarshipStatus.active.value,
    )
    db.add(st)
    await db.commit()
    await db.refresh(st)
    return st


@pytest_asyncio.fixture
async def config(db: AsyncSession, sch_type) -> ScholarshipConfiguration:
    cfg = ScholarshipConfiguration(
        scholarship_type_id=sch_type.id,
        config_name="CDE 114-1",
        config_code="CDE-114-1",
        academic_year=114,
        semester="first",
        amount=40000,
        is_active=True,
        allow_college_view_distribution=True,
    )
    db.add(cfg)
    await db.commit()
    await db.refresh(cfg)
    return cfg


@pytest_asyncio.fixture
async def college_client(db: AsyncSession, client: AsyncClient, sch_type) -> AsyncClient:
    user = User(
        nycu_id="cde_college_A",
        email="cde_college_A@university.edu",
        name="College A",
        user_type=UserType.employee,
        role=UserRole.college,
        college_code="A",
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)
    db.add(AdminScholarship(admin_id=user.id, scholarship_id=sch_type.id))
    await db.commit()

    async def override_college():
        return user

    app.dependency_overrides[require_college] = override_college
    try:
        yield client
    finally:
        app.dependency_overrides.pop(require_college, None)


async def _seed(db, sch_type):
    """Two colleges' rankings; college A has 正取/未錄取, college B has one 正取."""
    rankings = {}
    for code in ("A", "B"):
        r = CollegeRanking(
            scholarship_type_id=sch_type.id,
            sub_type_code="nstc",
            academic_year=114,
            semester="first",
            college_code=code,
            ranking_name=f"nstc 114-1 {code}",
            total_applications=2,
            is_finalized=True,
            distribution_executed=True,
            allocated_count=1,
        )
        db.add(r)
        rankings[code] = r
    await db.commit()
    for r in rankings.values():
        await db.refresh(r)

    def app_row(sid, name, academy, dept):
        student = User(
            nycu_id=f"cde_student_{sid}",
            email=f"cde_student_{sid}@university.edu",
            name=name,
            user_type=UserType.student,
            role=UserRole.student,
        )
        db.add(student)
        return Application(
            app_id=f"APP-CDE-{sid}",
            student=student,
            scholarship_type_id=sch_type.id,
            academic_year=114,
            semester="first",
            status="approved",
            sub_type_selection_mode=SubTypeSelectionMode.single,
            student_data={
                "std_stdcode": sid,
                "std_cname": name,
                "std_academyno": academy,
                "trm_depname": dept,
            },
        )

    a_admit = app_row("A001", "王小明", "A", "電子研")
    a_reject = app_row("A003", "張三", "A", "資工研")
    b_admit = app_row("B001", "他院生", "B", "機械研")
    for a in (a_admit, a_reject, b_admit):
        db.add(a)
    await db.commit()
    for a in (a_admit, a_reject, b_admit):
        await db.refresh(a)

    db.add_all(
        [
            CollegeRankingItem(
                ranking_id=rankings["A"].id,
                application_id=a_admit.id,
                rank_position=1,
                is_allocated=True,
                allocated_sub_type="nstc",
                status="allocated",
            ),
            CollegeRankingItem(
                ranking_id=rankings["A"].id,
                application_id=a_reject.id,
                rank_position=2,
                is_allocated=False,
                status="rejected",
            ),
            CollegeRankingItem(
                ranking_id=rankings["B"].id,
                application_id=b_admit.id,
                rank_position=1,
                is_allocated=True,
                allocated_sub_type="nstc",
                status="allocated",
            ),
        ]
    )
    await db.commit()


def _rows(payload: bytes):
    ws = load_workbook(io.BytesIO(payload)).active
    return [[c.value for c in row] for row in ws.iter_rows(min_row=3)]


@pytest.mark.asyncio
async def test_export_xlsx_contains_only_this_college(college_client, config, sch_type, db):
    """THE isolation test: parse the workbook and assert college B's student is
    absent from college A's export."""
    await _seed(db, sch_type)
    resp = await college_client.get(
        EXPORT_URL, params={"scholarship_type_id": sch_type.id, "academic_year": 114, "semester": "first"}
    )
    assert resp.status_code == 200
    rows = _rows(resp.content)
    numbers = {r[3] for r in rows}
    assert numbers == {"A001", "A003"}
    assert "B001" not in numbers
    names = {r[4] for r in rows}
    assert "他院生" not in names


@pytest.mark.asyncio
async def test_export_xlsx_row_contents(college_client, config, sch_type, db):
    await _seed(db, sch_type)
    resp = await college_client.get(
        EXPORT_URL, params={"scholarship_type_id": sch_type.id, "academic_year": 114, "semester": "first"}
    )
    rows = _rows(resp.content)
    admitted = next(r for r in rows if r[3] == "A001")
    assert admitted[1] == "正取"
    assert admitted[2] == 1
    assert admitted[4] == "王小明"
    assert admitted[5] == "電子研"

    rejected = next(r for r in rows if r[3] == "A003")
    assert rejected[1] == "未錄取"
    assert rejected[2] == 2  # 名次 populated on 未錄取 too
    assert rejected[5] == "資工研"


@pytest.mark.asyncio
async def test_export_pdf_returns_pdf_bytes(college_client, config, sch_type, db):
    await _seed(db, sch_type)
    resp = await college_client.get(
        EXPORT_URL,
        params={"scholarship_type_id": sch_type.id, "academic_year": 114, "semester": "first", "format": "pdf"},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF")


@pytest.mark.asyncio
async def test_export_sets_utf8_filename_header(college_client, config, sch_type, db):
    await _seed(db, sch_type)
    resp = await college_client.get(
        EXPORT_URL, params={"scholarship_type_id": sch_type.id, "academic_year": 114, "semester": "first"}
    )
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith("attachment; filename*=UTF-8''")
    assert resp.headers["content-length"] == str(len(resp.content))


@pytest.mark.asyncio
async def test_export_403_when_flag_off(college_client, config, sch_type, db):
    """The export inherits the loader's gate — no separate check to forget."""
    config.allow_college_view_distribution = False
    await db.commit()
    await _seed(db, sch_type)
    resp = await college_client.get(
        EXPORT_URL, params={"scholarship_type_id": sch_type.id, "academic_year": 114, "semester": "first"}
    )
    assert resp.status_code == 403
    body = resp.json()
    assert "分發結果尚未開放查看" in (body.get("detail") or body.get("message") or "")


@pytest.mark.asyncio
async def test_export_403_without_grant(client, config, sch_type, db):
    await _seed(db, sch_type)
    ungranted = User(
        nycu_id="cde_college_nogrant",
        email="cde_college_nogrant@university.edu",
        name="No Grant",
        user_type=UserType.employee,
        role=UserRole.college,
        college_code="A",
    )
    db.add(ungranted)
    await db.commit()

    async def override_college():
        return ungranted

    app.dependency_overrides[require_college] = override_college
    try:
        resp = await client.get(
            EXPORT_URL, params={"scholarship_type_id": sch_type.id, "academic_year": 114, "semester": "first"}
        )
    finally:
        app.dependency_overrides.pop(require_college, None)
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_export_rejects_unknown_format(college_client, config, sch_type, db):
    await _seed(db, sch_type)
    resp = await college_client.get(
        EXPORT_URL,
        params={"scholarship_type_id": sch_type.id, "academic_year": 114, "semester": "first", "format": "csv"},
    )
    assert resp.status_code == 422  # FastAPI Literal validation
