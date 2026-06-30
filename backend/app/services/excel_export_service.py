"""
Excel匯出服務
Excel export service for payment roster generation
"""

import hashlib
import logging
import os
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.exceptions import FileStorageError
from app.models.payment_roster import (
    MANUAL_REMOVAL_PREFIXES,
    PaymentRoster,
    PaymentRosterItem,
    StudentVerificationStatus,
)
from app.services.minio_service import MinIOService

logger = logging.getLogger(__name__)


class ExcelExportService:
    """Excel匯出服務"""

    # SECURITY: Allowlist of valid Excel template files to prevent path traversal
    ALLOWED_TEMPLATES = {
        "STD_UP_MIXLISTA.xlsx",
        "payment_roster_template.xlsx",
        "scholarship_roster.xlsx",
    }

    # 資格驗證欄名（附加於 base 30 欄之後）。逐條規則欄為動態，介於
    # COL_RULE_SUMMARY 與 COL_INCLUDED 之間（見 _collect_rule_columns）。
    COL_VERIFICATION = "學籍驗證"  # verification_status 標籤
    COL_RULE_SUMMARY = "規則資格"  # is_eligible → 符合/不符合/—
    COL_INCLUDED = "納入造冊"  # is_included → 是/否
    COL_EXCLUSION = "排除原因"  # exclusion_reason
    COL_BANK_ACCOUNT = "帳號"  # base 第 3 欄（缺帳號時標紅）

    # 不符合 → 紅底（FFC7CE）；警告 → 琥珀底（FFEB9C），Excel 標準條件格式色。
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # 需千分位數字格式的欄位（皆為金額欄）
    NUMERIC_FORMAT_COLUMNS = frozenset({"單價", "免稅給付"})

    def __init__(self):
        self.export_base_path = getattr(settings, "roster_export_dir", "./exports")
        self.template_dir = getattr(settings, "roster_template_dir", "./app/templates")
        default_template = getattr(settings, "roster_excel_template", "STD_UP_MIXLISTA.xlsx")
        self.template_path = os.path.join(self.template_dir, default_template)
        self.default_template_name = default_template
        self.ensure_export_directory()
        self._load_template_structure()

    def _get_template_paths(self) -> Dict[str, str]:
        """Get hardcoded template path mapping (CodeQL requirement)"""
        return {
            "STD_UP_MIXLISTA.xlsx": os.path.join(self.template_dir, "STD_UP_MIXLISTA.xlsx"),
            "payment_roster_template.xlsx": os.path.join(self.template_dir, "payment_roster_template.xlsx"),
            "scholarship_roster.xlsx": os.path.join(self.template_dir, "scholarship_roster.xlsx"),
        }

    def _load_template_structure(self):
        """Load template structure from STD_UP_MIXLISTA.xlsx"""
        try:
            if os.path.exists(self.template_path):
                wb = load_workbook(self.template_path)
                ws = wb.active

                # Extract headers from template
                self.template_columns = []
                for col in range(1, 33):  # 32 columns max
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        self.template_columns.append(str(cell_value))
                    else:
                        break

                logger.info(f"Loaded {len(self.template_columns)} columns from template")
                wb.close()
            else:
                logger.warning(f"Template not found at {self.template_path}, using default columns")
                self._set_default_columns()
        except Exception:
            logger.exception("Failed to load template; using default columns")
            self._set_default_columns()

    def _set_default_columns(self):
        """Set default 28-column structure for payment roster export"""
        self.template_columns = [
            "身分證字號",  # 1. 必填
            "姓名",  # 2. 必填
            "帳號",  # 3. 郵局帳號
            "銀行代碼",  # 4. 統一填700(郵局)
            "職別(稱)",  # 5. 固定值"學生"
            "戶籍地址",  # 6. 選填
            "身份別代碼",  # 7. 學生固定為1
            "單位(ex:時,月,次...)",  # 8. 固定值"次"
            "數量",  # 9. 固定值"1"
            "單價",  # 10. 獎學金金額
            "機關負擔勞保費",  # 11. 學生不適用，留空
            "機關負擔健保費",  # 12. 學生不適用，留空
            "機關負擔補充保費",  # 13. 學生不適用，留空
            "機關負擔勞退金",  # 14. 學生不適用，留空
            "機關負擔離職金",  # 15. 學生不適用，留空
            "機關負擔職災",  # 16. 學生不適用，留空
            "個人自付勞保費",  # 17. 學生不適用，留空
            "個人自付健保費",  # 18. 學生不適用，留空
            "個人自付補充保費",  # 19. 學生不適用，留空
            "個人自付勞退金",  # 20. 學生不適用，留空
            "個人自付離職金",  # 21. 學生不適用，留空
            "代扣所得",  # 22. 學生不適用，留空
            "其他代扣",  # 23. 學生不適用，留空
            "免稅給付",  # 24. 獎學金金額(全額免稅)
            "說明",  # 25. 組合字串：期間+獎學金+狀態
            "E-MAIL",  # 26. 選填
            "個人身分別(1:本國人,2:外國人,3:大陸人)",  # 27. 本國人=1
            "居留天數是否滿183天(是/否)",  # 28. 本國人預設"是"
            "申請身分",  # 29. 申請身分別 (e.g. 114新申請, 114續領)
            "分發獎學金",  # 30. 分發到的獎學金 (e.g. 114年 國科會)
        ]

        # Field mapping for easy access
        self.field_mapping = {
            "id_number": 0,  # 身分證字號
            "name": 1,  # 姓名
            "bank_account": 2,  # 帳號
            "bank_code": 3,  # 銀行代碼
            "job_title": 4,  # 職別(稱)
            "address": 5,  # 戶籍地址
            "identity_code": 6,  # 身份別代碼
            "unit": 7,  # 單位(ex:時,月,次...)
            "quantity": 8,  # 數量
            "amount": 9,  # 單價
            "org_labor_insurance": 10,  # 機關負擔勞保費
            "org_health_insurance": 11,  # 機關負擔健保費
            "org_supplementary_premium": 12,  # 機關負擔補充保費
            "org_labor_pension": 13,  # 機關負擔勞退金
            "org_severance": 14,  # 機關負擔離職金
            "org_occupational_injury": 15,  # 機關負擔職災
            "personal_labor_insurance": 16,  # 個人自付勞保費
            "personal_health_insurance": 17,  # 個人自付健保費
            "personal_supplementary_premium": 18,  # 個人自付補充保費
            "personal_labor_pension": 19,  # 個人自付勞退金
            "personal_severance": 20,  # 個人自付離職金
            "withheld_income": 21,  # 代扣所得
            "other_deductions": 22,  # 其他代扣
            "tax_exempt_payment": 23,  # 免稅給付
            "remarks": 24,  # 說明
            "email": 25,  # E-MAIL
            "personal_identity": 26,  # 個人身分別
            "residence_days": 27,  # 居留天數是否滿183天
        }

    def ensure_export_directory(self):
        """確保匯出目錄存在；若無權限則退回系統暫存目錄（容器中 cwd 可能不可寫）"""
        try:
            Path(self.export_base_path).mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            import tempfile

            fallback = os.path.join(tempfile.gettempdir(), "scholarship-exports")
            logger.warning(
                "Cannot create export dir %s; falling back to %s",
                self.export_base_path,
                fallback,
                exc_info=True,
            )
            Path(fallback).mkdir(parents=True, exist_ok=True)
            self.export_base_path = fallback

    def export_roster_to_excel(
        self,
        roster: PaymentRoster,
        include_excluded: bool = False,
        *,
        template_name: Optional[str] = None,
        include_header: bool = True,
        include_statistics: bool = True,
        async_mode: bool = False,
        skip_minio_upload: bool = False,
    ) -> Dict[str, Any]:
        """
        匯出造冊至Excel檔案

        Args:
            roster: 造冊對象 (MUST be already persisted with a valid ID)
            include_excluded: 是否包含被排除的項目
            template_name: 指定模板名稱 (預設為設定檔定義)
            include_header: 是否在檔案中包含表頭
            include_statistics: 是否加入統計資訊工作表
            async_mode: 是否以非同步模式回傳背景任務資訊
            skip_minio_upload: 是否跳過 MinIO 上傳並保留本地檔案 (用於 use_minio=False 的下載路徑)

        Returns:
            Dict[str, Any]: 匯出結果
            {
                "file_path": str,
                "file_name": str,
                "file_size": int,
                "file_hash": str,
                "minio_object_name": str,  # MinIO object path (if upload succeeded)
                "total_rows": int,
                "qualified_count": int,
                "disqualified_count": int,
                "template_name": str,
                "include_header": bool,
                "include_statistics": bool,
            }

        Raises:
            FileStorageError: 檔案儲存失敗

        Important:
            - This method modifies roster.minio_object_name in memory
            - CALLER MUST call db.commit() to persist the MinIO object name
            - If MinIO upload fails, export still succeeds with local file only
            - roster.id must be valid (not None) for MinIO upload to work
        """
        resolved_template_path = self._resolve_template_path(template_name)
        resolved_template_name = os.path.basename(resolved_template_path)

        if async_mode:
            task_id = f"roster-export-{uuid4().hex}"
            estimated_completion = (datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat()
            return {
                "task_id": task_id,
                "status": "queued",
                "template_name": resolved_template_name,
                "include_header": include_header,
                "include_statistics": include_statistics,
                "estimated_completion": estimated_completion,
            }

        file_name: Optional[str] = None
        try:
            # 取得造冊明細
            roster_items = self._get_roster_items(roster, include_excluded)

            # 產生檔案名稱
            file_name = roster.generate_excel_filename()
            file_path = os.path.join(self.export_base_path, file_name)

            # 計算動態欄位（base 30 欄 + 驗證欄 + 逐條規則欄）
            rule_columns = self._collect_rule_columns(roster_items)
            export_columns = self._build_export_columns(rule_columns)

            # 準備Excel資料 + 上色 metadata
            excel_data, cell_fills = self._prepare_excel_data(roster, roster_items, rule_columns)

            # 驗證資料品質
            validation_result = self._validate_export_data(excel_data)
            if validation_result["warnings"]:
                logger.warning(f"Data validation warnings: {validation_result['warnings']}")
            if not validation_result["is_valid"]:
                logger.error(f"Data validation failed: {validation_result['errors']}")
                raise FileStorageError(
                    f"Data validation failed: {'; '.join(validation_result['errors'])}",
                    file_name=file_name,
                )

            # 建立Excel檔案
            self._create_excel_file(
                excel_data,
                cell_fills,
                file_path,
                roster,
                template_path=resolved_template_path,
                columns=export_columns,
                include_header=include_header,
                include_statistics=include_statistics,
            )

            # 計算檔案資訊
            file_size = os.path.getsize(file_path)
            file_hash = self._calculate_file_hash(file_path)

            # 更新造冊檔案資訊
            roster.excel_filename = file_name
            roster.excel_file_path = file_path
            roster.excel_file_size = file_size
            roster.excel_file_hash = file_hash

            # Upload to MinIO automatically (if roster has valid ID and not explicitly skipped)
            if skip_minio_upload:
                logger.info(
                    f"Skipping MinIO upload for roster {roster.id} (skip_minio_upload=True); "
                    f"local file retained at: {file_path}"
                )
            elif roster.id is None:
                logger.warning(
                    "Roster ID is None, cannot upload to MinIO. " "Roster must be persisted to database before export."
                )
            else:
                try:
                    # File size check (default limit: 50MB)
                    max_file_size = 50 * 1024 * 1024  # 50MB
                    if file_size > max_file_size:
                        logger.warning(
                            f"Excel file size ({file_size} bytes) exceeds recommended limit "
                            f"({max_file_size} bytes). Upload may be slow."
                        )

                    minio_service = MinIOService()

                    # Read the Excel file
                    with open(file_path, "rb") as f:
                        file_content = f.read()

                    # Upload to MinIO
                    minio_result = minio_service.upload_roster_file(
                        file_content=file_content,
                        filename=file_name,
                        roster_id=roster.id,
                        metadata={"export_type": "excel", "template": resolved_template_name},
                    )

                    # Store MinIO object name in roster (caller must commit to persist)
                    roster.minio_object_name = minio_result["object_name"]

                    logger.info(
                        f"Excel file uploaded to MinIO: {minio_result['object_name']} "
                        f"(size: {file_size} bytes, hash: {file_hash[:8]}...)"
                    )

                    # Clean up local file after successful MinIO upload
                    try:
                        os.remove(file_path)
                        logger.info(f"Cleaned up local file after MinIO upload: {file_path}")
                    except Exception as cleanup_error:
                        logger.warning(f"Failed to cleanup local file {file_path}: {cleanup_error}")

                except Exception as minio_error:
                    # Log error but don't fail the export - local file still exists
                    logger.error(f"Failed to upload Excel file to MinIO: {minio_error}", exc_info=True)
                    logger.warning(f"Excel file remains available locally at: {file_path}")

            qualified_count = sum(1 for item in roster_items if item.is_qualified and item.is_included)
            disqualified_count = len(roster_items) - qualified_count

            logger.info(f"Excel export completed: {file_name} ({file_size} bytes)")

            return {
                "file_path": file_path,
                "file_name": file_name,
                "file_size": file_size,
                "file_hash": file_hash,
                "minio_object_name": roster.minio_object_name,  # Include MinIO object name
                "total_rows": len(roster_items),
                "qualified_count": qualified_count,
                "disqualified_count": disqualified_count,
                "validation_result": validation_result,
                "template_columns": export_columns,
                "template_name": resolved_template_name,
                "include_header": include_header,
                "include_statistics": include_statistics,
            }

        except Exception as e:
            logger.exception("Excel export failed")
            raise FileStorageError(f"Failed to export Excel file: {e}", file_name=file_name) from e

    @staticmethod
    def _is_manual_removal(item) -> bool:
        """True 表示此 item 是「鎖定後手動移除」或「比對分發移除」，
        應排除於資格驗證視圖之外（自動因資格不符排除者則保留並標紅）。
        前綴常數與 roster_service 產生端共用（MANUAL_REMOVAL_PREFIXES）。"""
        reason = getattr(item, "exclusion_reason", None)
        return bool(reason and reason.startswith(MANUAL_REMOVAL_PREFIXES))

    @staticmethod
    def _rule_details(item) -> Dict[str, Any]:
        """取出凍結快照的逐條規則 details；缺漏／格式不符時回傳空 dict。"""
        rvr = getattr(item, "rule_validation_result", None)
        details = rvr.get("details") if isinstance(rvr, dict) else None
        return details if isinstance(details, dict) else {}

    def _get_roster_items(self, roster: PaymentRoster, include_excluded: bool) -> List[PaymentRosterItem]:
        """取得造冊明細。

        預設（include_excluded=False）：保留「納入」與「因資格不符自動排除」者，
        隱藏「手動移除」者。include_excluded=True 時回傳全部（含手動移除）。
        """
        items = list(roster.items)

        if not include_excluded:
            items = [item for item in items if item.is_included or not self._is_manual_removal(item)]

        return sorted(items, key=lambda x: x.student_name)

    def _collect_rule_columns(self, roster_items: List[PaymentRosterItem]) -> List[Tuple[int, str]]:
        """從所有 item 的凍結快照 rule_validation_result["details"] 收集規則欄。

        回傳依 rule_id 升冪排序的 (rule_id, header)；header 以 rule_name 為主，
        若與 base/驗證欄名或其他規則 header 衝突，加上「（#id）」後綴確保唯一。
        非規則鍵（no_rules_found / error）忽略。Excel 層只讀不重跑規則。
        """
        seen: Dict[int, str] = {}
        for item in roster_items:
            for key, res in self._rule_details(item).items():
                if not isinstance(key, str) or not key.startswith("rule_"):
                    continue
                suffix = key[len("rule_") :]
                if not suffix.isdigit():
                    continue
                rid = int(suffix)
                if rid in seen:
                    continue
                name = (res.get("rule_name") if isinstance(res, dict) else None) or f"規則{rid}"
                seen[rid] = name

        columns: List[Tuple[int, str]] = []
        used = set(self.template_columns) | {
            self.COL_VERIFICATION,
            self.COL_RULE_SUMMARY,
            self.COL_INCLUDED,
            self.COL_EXCLUSION,
        }
        for rid in sorted(seen):
            header = seen[rid]
            if header in used:
                header = f"{header}（#{rid}）"
            used.add(header)
            columns.append((rid, header))
        return columns

    def _build_export_columns(self, rule_columns: List[Tuple[int, str]]) -> List[str]:
        """組出本次匯出的完整欄位順序：base 30 欄 + 驗證欄 + 逐條規則欄 + 納入/排除。"""
        return (
            list(self.template_columns)
            + [self.COL_VERIFICATION, self.COL_RULE_SUMMARY]
            + [header for _, header in rule_columns]
            + [self.COL_INCLUDED, self.COL_EXCLUSION]
        )

    def _resolve_template_path(self, template_name: Optional[str]) -> str:
        """
        Resolve Excel template path based on optional template name.

        SECURITY: Validates template against allowlist to prevent path traversal attacks.
        """
        if not template_name:
            return self.template_path

        candidate = template_name if template_name.lower().endswith(".xlsx") else f"{template_name}.xlsx"

        # SECURITY: Check allowlist FIRST before any path operations
        if candidate not in self.ALLOWED_TEMPLATES:
            logger.warning(f"Template '{candidate}' not in allowlist. Using default template.")
            return self.template_path

        # SECURITY: Use hardcoded path mapping instead of string interpolation (CodeQL requirement)
        template_paths = self._get_template_paths()
        candidate_path = template_paths.get(candidate)

        if candidate_path and os.path.exists(candidate_path):
            return candidate_path

        logger.warning("Template %s not found. Falling back to default template.", candidate)
        return self.template_path

    @staticmethod
    def _format_allocation_display(item: PaymentRosterItem) -> str:
        """Format allocated sub-type + year for Excel display"""
        if not item.allocated_sub_type:
            return ""
        sub_type_map = {
            "nstc": "國科會",
            "moe_1w": "教育部(1萬)",
            "moe_2w": "教育部(2萬)",
        }
        display = sub_type_map.get(item.allocated_sub_type, item.allocated_sub_type)
        if item.allocation_year:
            return f"{item.allocation_year}年 {display}"
        return display

    def _get_filtered_roster_items(self, roster: PaymentRoster, include_excluded: bool) -> List[PaymentRosterItem]:
        """Compatibility helper for legacy calls that expect filtered roster items"""
        return self._get_roster_items(roster, include_excluded)

    def _prepare_excel_data(
        self,
        roster: PaymentRoster,
        roster_items: List[PaymentRosterItem],
        rule_columns: Optional[List[Tuple[int, str]]] = None,
    ) -> Tuple[List[Dict], List[Dict[str, str]]]:
        """準備 Excel 資料 — base 30 欄 + 資格驗證欄。

        回傳 (excel_data, cell_fills)：cell_fills 與 excel_data 等長平行，
        每筆為 {欄名: "red"|"amber"}，供 _create_excel_file 套色。
        item.excel_row_data 仍寫回乾淨 row dict（不含 fills）。
        """
        rule_columns = rule_columns or []
        # 預先算好每欄的快照鍵，避免逐列重建 f"rule_{rid}"
        rule_lookup = [(header, f"rule_{rid}") for rid, header in rule_columns]
        excel_data: List[Dict] = []
        cell_fills: List[Dict[str, str]] = []

        for idx, item in enumerate(roster_items, start=1):
            # 取得學生相關資訊

            # 驗證必填欄位
            if not item.student_id_number or not item.student_name:
                logger.warning(
                    f"Skipping item {idx} due to missing required fields: "
                    f"ID={item.student_id_number}, Name={item.student_name}"
                )
                continue

            # 生成備註 (第25欄) - 包含期間標籤、獎學金名稱、分發資訊和狀態
            remarks_parts = []
            if hasattr(roster, "period_label") and roster.period_label:
                remarks_parts.append(f"期間:{roster.period_label}")
            remarks_parts.append(f"獎學金:{item.scholarship_name}")
            # 分發資訊（從快照欄位取得）
            if item.application_identity:
                remarks_parts.append(f"身分:{item.application_identity}")
            if item.allocated_sub_type:
                remarks_parts.append(f"分發:{self._format_allocation_display(item)}")
            if not item.is_included:
                remarks_parts.append("狀態:不合格")
            if not item.bank_account:
                remarks_parts.append("缺銀行資訊")
            remarks = " ".join(remarks_parts)

            # 建立28欄位標準格式資料映射
            row_data = {
                # 1. 身分證字號 (必填)
                "身分證字號": item.student_id_number,
                # 2. 姓名 (必填)
                "姓名": item.student_name,
                # 3. 帳號 (郵局帳號)
                "帳號": item.bank_account or "",
                # 4. 銀行代碼 (郵局代碼固定為700)
                "銀行代碼": "700",
                # 5. 職別(稱) (固定值"學生")
                "職別(稱)": "學生",
                # 6. 戶籍地址 (選填)
                "戶籍地址": item.permanent_address or "",
                # 7. 身份別代碼 (學生固定為1)
                "身份別代碼": "1",
                # 8. 單位(ex:時,月,次...) (固定值"次")
                "單位(ex:時,月,次...)": "次",
                # 9. 數量 (固定值"1")
                "數量": "1",
                # 10. 單價 (獎學金金額)
                "單價": float(item.scholarship_amount) if item.scholarship_amount else 0,
                # 11-16. 機關負擔項目 (學生不適用，留空)
                "機關負擔勞保費": "",
                "機關負擔健保費": "",
                "機關負擔補充保費": "",
                "機關負擔勞退金": "",
                "機關負擔離職金": "",
                "機關負擔職災": "",
                # 17-21. 個人自付項目 (學生不適用，留空)
                "個人自付勞保費": "",
                "個人自付健保費": "",
                "個人自付補充保費": "",
                "個人自付勞退金": "",
                "個人自付離職金": "",
                # 22-23. 代扣項目 (學生不適用，留空)
                "代扣所得": "",
                "其他代扣": "",
                # 24. 免稅給付 (獎學金金額，全額免稅)
                "免稅給付": float(item.scholarship_amount) if item.scholarship_amount else 0,
                # 25. 說明 (期間+獎學金+狀態)
                "說明": remarks,
                # 26. E-MAIL (選填)
                "E-MAIL": item.student_email or "",
                # 27. 個人身分別(1:本國人,2:外國人,3:大陸人) (本國人=1)
                "個人身分別(1:本國人,2:外國人,3:大陸人)": "1",
                # 28. 居留天數是否滿183天(是/否) (本國人預設"是")
                "居留天數是否滿183天(是/否)": "是",
                # 29. 申請身分 (快照欄位)
                "申請身分": item.application_identity or "",
                # 30. 分發獎學金 (快照欄位)
                "分發獎學金": self._format_allocation_display(item),
            }

            fills: Dict[str, str] = {}

            # 銀行帳號缺漏 → 帳號欄紅
            if not item.bank_account:
                fills[self.COL_BANK_ACCOUNT] = "red"

            # 學籍驗證
            status = item.verification_status
            row_data[self.COL_VERIFICATION] = self._get_verification_status_label(status)
            status_value = status.value if hasattr(status, "value") else str(status)
            if status_value != StudentVerificationStatus.VERIFIED.value:
                fills[self.COL_VERIFICATION] = "red"

            # 整體規則資格（重用 is_eligible property；stub 以屬性提供）
            elig = getattr(item, "is_eligible", None)
            row_data[self.COL_RULE_SUMMARY] = "符合" if elig is True else ("不符合" if elig is False else "—")
            if elig is False:
                fills[self.COL_RULE_SUMMARY] = "red"

            # 逐條規則欄（讀凍結快照，不重跑）
            details = self._rule_details(item)
            for header, rule_key in rule_lookup:
                res = details.get(rule_key)
                if isinstance(res, dict) and "passed" in res:
                    if res.get("passed"):
                        row_data[header] = "通過"
                    else:
                        row_data[header] = "未通過"
                        # warning 規則 → 琥珀；硬性或無嚴重度標記（含評估錯誤）→ 紅
                        fills[header] = "amber" if (res.get("is_warning") and not res.get("is_hard_rule")) else "red"
                else:
                    row_data[header] = "—"

            # 納入造冊 / 排除原因
            row_data[self.COL_INCLUDED] = "是" if item.is_included else "否"
            row_data[self.COL_EXCLUSION] = item.exclusion_reason or ""
            if not item.is_included:
                fills[self.COL_INCLUDED] = "red"
                fills[self.COL_EXCLUSION] = "red"

            # 儲存Excel行資料到資料庫（乾淨 row dict，不含 fills metadata）
            item.excel_row_data = row_data
            item.excel_remarks = remarks

            excel_data.append(row_data)
            cell_fills.append(fills)

        logger.info(f"Prepared {len(excel_data)} rows for export ({len(rule_columns)} rule columns)")
        return excel_data, cell_fills

    def _validate_export_data(self, excel_data: List[Dict]) -> Dict[str, Any]:
        """驗證STD_UP_MIXLISTA格式匯出資料品質"""
        validation_result = {
            "is_valid": True,
            "warnings": [],
            "errors": [],
            "statistics": {},
        }

        if not excel_data:
            validation_result["warnings"].append("No roster items to export — generating header-only file")
            return validation_result

        # 統計資料
        total_rows = len(excel_data)
        missing_bank_info = 0
        missing_required_info = 0
        invalid_amounts = 0
        invalid_format = 0

        # STD_UP_MIXLISTA必填欄位檢查
        required_fields = ["身分證字號", "姓名"]  # 第1、2欄必填

        for _idx, row in enumerate(excel_data, start=1):
            # 檢查必填欄位
            missing_required = False
            for field in required_fields:
                if not row.get(field) or str(row.get(field)).strip() == "":
                    missing_required = True
                    break

            if missing_required:
                missing_required_info += 1

            # 檢查銀行資訊 (第3、4欄)
            if not row.get("銀行代碼") or not row.get("帳號"):
                missing_bank_info += 1

            # 檢查金額格式 (第8欄給付總額)
            amount = row.get("給付總額", 0)
            if not isinstance(amount, (int, float)) or amount <= 0:
                invalid_amounts += 1

            # 檢查職別固定值 (第5欄)
            if row.get("職別(稱)") != "學生":
                invalid_format += 1

            # 檢查學生標記 (第7欄)
            if row.get("是否為學生") != "1":
                invalid_format += 1

            # 檢查扣繳憑單類別 (第28欄)
            if row.get("扣繳憑單類別") != "50":
                invalid_format += 1

        # 產生錯誤和警告
        if missing_required_info > 0:
            validation_result["errors"].append(
                f"{missing_required_info} records missing required fields (身分證字號/姓名)"
            )
            validation_result["is_valid"] = False

        if missing_bank_info > 0:
            validation_result["warnings"].append(
                f"{missing_bank_info} records missing bank information (將標註於說明欄)"
            )

        if invalid_amounts > 0:
            validation_result["warnings"].append(f"{invalid_amounts} records have invalid amounts")

        if invalid_format > 0:
            validation_result["warnings"].append(f"{invalid_format} records have format issues")

        # 計算完整性統計
        complete_records = total_rows - missing_bank_info
        completion_rate = (complete_records / total_rows * 100) if total_rows > 0 else 0

        validation_result["statistics"] = {
            "total_rows": total_rows,
            "missing_bank_info": missing_bank_info,
            "missing_required_info": missing_required_info,
            "invalid_amounts": invalid_amounts,
            "invalid_format": invalid_format,
            "completion_rate": completion_rate,
            "std_up_mixlista_compliant": missing_required_info == 0 and invalid_format == 0,
        }

        logger.info(
            f"STD_UP_MIXLISTA validation: {total_rows} total rows, "
            f"{missing_required_info} missing required, {missing_bank_info} missing bank info, "
            f"completion rate: {completion_rate:.1f}%"
        )

        return validation_result

    def _create_excel_file(
        self,
        excel_data: List[Dict],
        cell_fills: List[Dict[str, str]],
        file_path: str,
        roster: PaymentRoster,
        *,
        template_path: str,
        columns: List[str],
        include_header: bool,
        include_statistics: bool,
    ):
        """建立 Excel 檔案 — 依 columns 寫表頭/資料並套用紅/琥珀底。"""
        try:
            use_template = include_header and os.path.exists(template_path)

            if use_template:
                wb = load_workbook(template_path)
                ws = wb.active
                logger.info("Using template file for Excel generation: %s", template_path)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "印領清冊"
                logger.info("Created new Excel file using default structure (include_header=%s)", include_header)

            # 清掉既有列，統一依 columns 重寫表頭與資料
            if ws.max_row >= 1:
                ws.delete_rows(1, ws.max_row)

            if include_header:
                for col_idx, column_name in enumerate(columns, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=column_name)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                start_row = 2
            else:
                start_row = 1

            for row_idx, (row_data, fills) in enumerate(zip(excel_data, cell_fills), start=start_row):
                for col_idx, column_name in enumerate(columns, start=1):
                    value = row_data.get(column_name, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    if column_name in self.NUMERIC_FORMAT_COLUMNS and isinstance(value, (int, float)):
                        cell.number_format = "#,##0"

                    kind = fills.get(column_name)
                    if kind == "red":
                        cell.fill = self.RED_FILL
                    elif kind == "amber":
                        cell.fill = self.AMBER_FILL

            total_rows = max(len(excel_data) + (1 if include_header else 0), 1)
            self._apply_excel_styling(ws, total_rows, include_header, columns)

            if include_statistics:
                self._add_worksheet_info(wb, roster)

            wb.save(file_path)
            logger.info("Excel file created successfully: %s", file_path)

        except Exception as e:
            logger.exception("Failed to create Excel file")
            raise FileStorageError(
                f"Failed to create Excel file: {e}",
                file_name=os.path.basename(file_path),
            ) from e

    def _apply_excel_styling(self, ws, max_row: int, include_header: bool, columns: List[str]):
        """應用Excel樣式"""
        self._set_column_widths(ws, columns)
        self._set_borders(ws, max_row, columns)
        ws.freeze_panes = "A2" if include_header else None

    def _set_column_widths(self, ws, columns: List[str]):
        """設定欄寬 - 智慧調整基於欄位內容"""
        # 預設欄寬設定
        default_widths = {
            "序號": 6,
            "流水號": 6,
            "身分證字號": 12,
            "姓名": 10,
            "學校代碼": 8,
            "學校名稱": 15,
            "郵遞區號": 8,
            "地址": 25,
            "電話": 12,
            "電子信箱": 20,
            "銀行代號": 8,
            "銀行名稱": 12,
            "帳號": 15,
            "戶名": 10,
            "金額": 10,
            "所得類別": 8,
            "扣繳稅額": 8,
            "身分別": 8,
            "居留天數是否滿183天": 12,
            "統一編號": 10,
            "機關代號": 8,
            "獎學金名稱": 20,
            "獎學金子類型": 15,
            "學年度": 8,
            "學期": 8,
            "系所": 15,
            "年級": 6,
            "學號": 12,
            "指導教授": 10,
            "申請日期": 10,
            "核准日期": 10,
            "造冊日期": 10,
            "審核狀態": 8,
            "備註": 20,
            "驗證狀態": 10,
            "申請身分": 14,
            "分發獎學金": 18,
            "學籍驗證": 12,
            "規則資格": 10,
            "納入造冊": 10,
            "排除原因": 30,
        }

        for col_idx, column_name in enumerate(columns, start=1):
            width = default_widths.get(column_name, 12)
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = width

    def _set_borders(self, ws, max_row: int, columns: List[str]):
        """設定邊框"""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in range(1, max_row + 1):
            for col in range(1, len(columns) + 1):
                ws.cell(row=row, column=col).border = thin_border

    def _add_worksheet_info(self, wb: Workbook, roster: PaymentRoster):
        """加入工作表資訊"""
        info_ws = wb.create_sheet("造冊資訊")

        info_data = [
            ["造冊代碼", roster.roster_code],
            ["期間標記", roster.period_label],
            ["學年度", roster.academic_year],
            ["造冊週期", roster.roster_cycle.value],
            ["觸發方式", roster.trigger_type.value],
            [
                "產生時間",
                roster.started_at.strftime("%Y-%m-%d %H:%M:%S") if roster.started_at else "",
            ],
            [
                "完成時間",
                roster.completed_at.strftime("%Y-%m-%d %H:%M:%S") if roster.completed_at else "",
            ],
            ["總申請數", roster.total_applications or 0],
            ["合格人數", roster.qualified_count or 0],
            ["不合格人數", roster.disqualified_count or 0],
            ["總金額", float(roster.total_amount) if roster.total_amount else 0],
            ["學籍驗證啟用", "是" if roster.student_verification_enabled else "否"],
            ["API失敗次數", roster.verification_api_failures or 0],
        ]

        for row_idx, (label, value) in enumerate(info_data, start=1):
            info_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            info_ws.cell(row=row_idx, column=2, value=value)

        info_ws.column_dimensions["A"].width = 15
        info_ws.column_dimensions["B"].width = 20

    def _extract_postal_code(self, address: Optional[str]) -> str:
        """從地址中提取郵遞區號"""
        if not address:
            return ""

        # 簡單的郵遞區號提取邏輯
        import re

        match = re.match(r"^(\d{3,5})", address.strip())
        return match.group(1) if match else ""

    def _get_advisor_name(self, student) -> str:
        """取得指導教授姓名"""
        if not student:
            return ""

        try:
            # 查詢學生的指導教授關係
            from sqlalchemy.orm import sessionmaker

            from app.db.session import engine
            from app.models.professor_student_relationship import ProfessorStudentRelationship

            SessionLocal = sessionmaker(bind=engine)
            db = SessionLocal()

            try:
                relationship = (
                    db.query(ProfessorStudentRelationship)
                    .filter(
                        ProfessorStudentRelationship.student_id == student.id,
                        ProfessorStudentRelationship.is_active.is_(True),
                    )
                    .first()
                )

                if relationship and relationship.professor:
                    return relationship.professor.name

            finally:
                db.close()

        except Exception as e:
            logger.warning(f"Failed to get advisor name for student {student.id}", exc_info=True)

        return ""

    def _generate_remarks(self, item: PaymentRosterItem, roster: PaymentRoster) -> str:
        """產生備註欄內容"""
        remarks = []

        # 基本造冊資訊
        remarks.append(f"造冊期間: {roster.period_label}")
        if roster.scholarship_configuration:
            remarks.append(f"獎學金: {roster.scholarship_configuration.config_code}")

        # 判斷狀態
        if not item.is_included:
            if item.exclusion_reason:
                remarks.append(f"排除原因: {item.exclusion_reason}")
        elif item.verification_status.value != "verified":
            remarks.append(f"學籍狀態: {self._get_verification_status_label(item.verification_status)}")
        elif not item.bank_account:
            remarks.append("缺少銀行資訊")
        else:
            remarks.append("合格")

        # 添加規則驗證警告（如果有）
        if item.warning_rules:
            warning_msg = "警告: " + "; ".join(item.warning_rules)
            remarks.append(warning_msg)

        return "; ".join(remarks)

    def _get_verification_status_label(self, status) -> str:
        """取得驗證狀態標籤"""
        labels = {
            "verified": "已驗證",
            "graduated": "已畢業",
            "suspended": "休學中",
            "withdrawn": "已退學",
            "api_error": "驗證錯誤",
            "not_found": "查無此人",
        }
        return labels.get(status.value if hasattr(status, "value") else str(status), str(status))

    def _calculate_file_hash(self, file_path: str) -> str:
        """計算檔案SHA256雜湊值"""
        hash_sha256 = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()

    def get_excel_file_info(self, roster: PaymentRoster) -> Optional[Dict[str, Any]]:
        """取得Excel檔案資訊"""
        if not roster.excel_file_path or not os.path.exists(roster.excel_file_path):
            return None

        return {
            "file_name": roster.excel_filename,
            "file_path": roster.excel_file_path,
            "file_size": roster.excel_file_size,
            "file_hash": roster.excel_file_hash,
            "created_at": roster.completed_at,
        }

    def delete_excel_file(self, roster: PaymentRoster) -> bool:
        """刪除Excel檔案"""
        if not roster.excel_file_path:
            return False

        try:
            if os.path.exists(roster.excel_file_path):
                os.remove(roster.excel_file_path)

            # 清除資料庫中的檔案資訊
            roster.excel_filename = None
            roster.excel_file_path = None
            roster.excel_file_size = None
            roster.excel_file_hash = None

            logger.info(f"Deleted Excel file: {roster.excel_file_path}")
            return True

        except Exception:
            logger.exception("Failed to delete Excel file")
            return False

    def preview_roster_export(
        self,
        roster: PaymentRoster,
        template_name: str = "STD_UP_MIXLISTA",
        include_header: bool = True,
        max_preview_rows: int = 10,
        *,
        include_excluded: bool = False,
    ) -> Dict[str, Any]:
        """
        預覽造冊匯出內容
        Preview roster export content without creating actual file
        """
        try:
            # 取得造冊項目
            roster_items = self._get_filtered_roster_items(roster, include_excluded=include_excluded)

            # 計算動態欄位 + 準備Excel資料（預覽不需 fills）
            rule_columns = self._collect_rule_columns(roster_items)
            export_columns = self._build_export_columns(rule_columns)
            excel_data, _ = self._prepare_excel_data(roster, roster_items, rule_columns)

            # 驗證資料
            validation_result = self._validate_export_data(excel_data)

            # 限制預覽行數
            preview_data = excel_data[:max_preview_rows] if max_preview_rows > 0 else excel_data

            # 產生metadata
            metadata = {
                "template_name": template_name,
                "total_rows": len(excel_data),
                "preview_rows": len(preview_data),
                "roster_code": roster.roster_code,
                "period_label": getattr(roster, "period_label", ""),
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "std_up_mixlista_format": True,
            }

            return {
                "preview_data": preview_data,
                "total_rows": len(excel_data),
                "column_headers": export_columns,
                "validation_result": validation_result,
                "metadata": metadata,
            }

        except Exception as e:
            logger.exception("Failed to preview roster export")
            raise FileStorageError(f"預覽產生失敗: {str(e)}") from e

    def process_async_export(
        self,
        roster_id: int,
        task_id: str,
        user_id: int,
        *,
        template_name: Optional[str] = None,
        include_header: bool = True,
        include_statistics: bool = True,
        include_excluded: bool = False,
    ):
        """
        處理非同步匯出任務
        Process asynchronous export task
        """
        try:
            from app.core.database import get_db_sync
            from app.services.audit_service import audit_service
            from app.services.minio_service import minio_service

            # 取得資料庫session
            db = next(get_db_sync())

            try:
                # 取得造冊
                roster = db.query(PaymentRoster).filter(PaymentRoster.id == roster_id).first()
                if not roster:
                    raise ValueError(f"Roster {roster_id} not found")

                # 執行匯出
                export_result = self.export_roster_to_excel(
                    roster=roster,
                    template_name=template_name or self.default_template_name,
                    include_header=include_header,
                    include_statistics=include_statistics,
                    include_excluded=include_excluded,
                    async_mode=False,
                )

                # 上傳到MinIO
                with open(export_result["file_path"], "rb") as f:
                    file_content = f.read()

                minio_result = minio_service.upload_roster_file(
                    file_content=file_content,
                    filename=os.path.basename(export_result["file_path"]),
                    roster_id=roster_id,
                    metadata={"task_id": task_id, "user_id": str(user_id)},
                )

                # 記錄稽核日誌
                audit_service.log_excel_export(
                    roster_id=roster_id,
                    filename=minio_result["object_name"],
                    file_size=minio_result["file_size"],
                    record_count=len(self._get_filtered_roster_items(roster, include_excluded=include_excluded)),
                    user_id=user_id,
                    user_name=f"user_{user_id}",
                    export_format="xlsx",
                    db=db,
                )

                logger.info(f"Async export completed: roster_id={roster_id}, task_id={task_id}")

            finally:
                db.close()

        except Exception:
            logger.exception(f"Async export failed: roster_id={roster_id}, task_id={task_id}")
            # 這裡可以實作錯誤通知機制
