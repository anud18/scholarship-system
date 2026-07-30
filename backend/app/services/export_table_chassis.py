"""Shared table-rendering chassis for tabular xlsx/PDF exports.

The chassis was born in ``college_ranking_export_service``, forked into
``college_distribution_export_service`` (see the fork warning that used to live
there), and extracted here when the manual-distribution 分發名單 export became
the third consumer. It owns the pieces every tabular export shares:

- reportlab paragraph styles (CJK font + ``wordWrap="CJK"``) and the grid
  ``TableStyle`` recipe;
- the ``KeepInFrame`` cell wrapper that stops one overlong cell from raising
  ``LayoutError`` and 500-ing the whole export (a reportlab ``Table`` cannot
  split a single row across pages);
- weight-driven column sizing, shared between the PDF (points) and xlsx
  (character-width) renderers so the two formats keep one source of truth;
- the openpyxl title/header/border helpers.

Renderers stay responsible for their own column model, row values and
``sanitize_excel_cell`` calls — the chassis never decides content.
"""

from __future__ import annotations

from typing import Any, List, Sequence, Tuple

# `escape` is a pure string-escaping helper (`<` → `&lt;` …) used to sanitise
# cell values before they go into reportlab Paragraph markup. It does not parse
# untrusted XML, so the B406 warning is a false positive here.
from xml.sax.saxutils import escape as xml_escape  # nosec B406

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import KeepInFrame, Paragraph

from app.services.pdf_fonts import CJK_FONT_NAME
from app.utils.excel_safety import sanitize_excel_cell

PDF_MARGIN_PT = 10 * mm
# Vertical space reserved (per page) for the title + spacer + ONE repeated header
# row. Renderers with a taller repeated header (e.g. two header rows) must NOT
# reuse this — measure instead, via `pdf_cell_max_height(..., header_height=...)`.
PDF_HEADER_RESERVE_PT = 60

# reportlab's SimpleDocTemplate gives its Frame 6pt of padding top and bottom,
# which is not part of the page margins and so has to come off the usable height.
PDF_FRAME_PADDING_PT = 12

# Excel width per unit of PDF column weight, plus a fixed padding allowance.
XLSX_WIDTH_PER_WEIGHT = 8
XLSX_WIDTH_PADDING = 6

_HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")


def safe_str(value: Any) -> str:
    return "" if value is None else str(value)


# -------- PDF helpers --------


def make_pdf_styles(prefix: str) -> Tuple[ParagraphStyle, ParagraphStyle, ParagraphStyle]:
    """Return (title, header, cell) paragraph styles for one export's PDF.

    ``wordWrap="CJK"`` on header/cell styles breaks long CJK runs AND unspaced
    ASCII (emails, student IDs) that would otherwise overflow the column.
    """
    title_style = ParagraphStyle(
        f"{prefix}PdfTitle",
        fontName=CJK_FONT_NAME,
        fontSize=12,
        leading=15,
        alignment=1,  # center
    )
    header_style = ParagraphStyle(
        f"{prefix}PdfHeader",
        fontName=CJK_FONT_NAME,
        fontSize=8,
        leading=10,
        alignment=1,
        wordWrap="CJK",
    )
    cell_style = ParagraphStyle(
        f"{prefix}PdfCell",
        fontName=CJK_FONT_NAME,
        fontSize=7.5,
        leading=9,
        wordWrap="CJK",
    )
    return title_style, header_style, cell_style


def pdf_col_widths(weights: Sequence[float], usable_width: float) -> List[float]:
    """Normalise relative column weights to the usable page width."""
    total = sum(weights) or 1.0
    return [usable_width * w / total for w in weights]


# Slack left over the measured header so rounding and a slightly taller wrap
# (fonts metrics differ per platform) still can't push a capped row over.
PDF_CELL_SAFETY_PT = 10


def pdf_cell_max_height(page_height: float, *, header_height: float) -> float:
    """Tallest a single body cell may be and still fit a CONTINUATION page.

    A reportlab ``Table`` cannot split one row across pages, so an overlong cell
    raises ``LayoutError`` and fails the WHOLE export. The binding constraint is a
    continuation page, which carries the repeated header rows but not the title —
    so the cap is the frame height minus the REAL header height.

    Pass the measured header height (``table.wrap(...)`` on a header-only table)
    rather than a guessed constant: a renderer with two header rows needs roughly
    twice the reserve of one, and guessing 60pt for a ~61pt header silently
    produced a cap ~13pt too generous, which raised LayoutError on free text as
    short as 700 characters.
    """
    usable = page_height - (PDF_MARGIN_PT * 2) - PDF_FRAME_PADDING_PT
    return max(usable - header_height - PDF_CELL_SAFETY_PT, 1.0)


def pdf_paragraph(text: Any, style: ParagraphStyle) -> Paragraph:
    """A Paragraph with reportlab-markup characters escaped."""
    return Paragraph(xml_escape(safe_str(text)), style)


def pdf_cell(value: Any, *, width: float, max_height: float, style: ParagraphStyle) -> KeepInFrame:
    """One body cell, capped to the usable page height.

    A reportlab ``Table`` cannot split ONE row across pages, so a single very
    long cell would raise ``LayoutError`` and fail the WHOLE export. Capping
    each cell and letting ``KeepInFrame`` shrink anything taller keeps the
    export alive no matter what free text arrives from SIS/students.
    """
    return KeepInFrame(width, max_height, [pdf_paragraph(value, style)], mode="shrink")


def base_pdf_table_style(header_rows: int = 1) -> List[tuple]:
    """The shared grid recipe; consumers may append SPAN/extra commands."""
    return [
        ("GRID", (0, 0), (-1, -1), 0.4, colors.Color(0.6, 0.6, 0.6)),
        ("BACKGROUND", (0, 0), (-1, header_rows - 1), colors.Color(0.87, 0.87, 0.87)),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, -1), CJK_FONT_NAME),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]


# -------- xlsx helpers --------


def write_xlsx_title_row(ws, *, title: str, total_cols: int) -> None:
    """Row 1: title merged across all columns.

    Sanitized like the data rows: callers' f-strings happen to lead with an int
    today, but that is incidental — this keeps a formula trigger impossible
    regardless of how the title is built.
    """
    ws.cell(row=1, column=1, value=sanitize_excel_cell(title))
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")


def style_xlsx_header_cell(cell) -> None:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = _HEADER_FILL


def apply_xlsx_borders(ws, *, max_row: int, max_col: int, first_row: int = 2) -> None:
    """Thin borders from ``first_row`` down — the title row stays borderless."""
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(first_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = border


def apply_xlsx_column_widths(ws, weights: Sequence[float]) -> None:
    """Size xlsx columns from the SAME weights the PDF uses.

    Deriving width from header text length instead would clamp every 2-char CJK
    header to an identical width, silently ignoring the weights.
    """
    for idx, weight in enumerate(weights, start=1):
        width = round(XLSX_WIDTH_PER_WEIGHT * weight) + XLSX_WIDTH_PADDING
        ws.column_dimensions[get_column_letter(idx)].width = width
