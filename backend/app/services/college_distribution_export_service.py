"""College distribution-results export service.

Pure rendering logic — receives the ``sub_types`` structure produced by
``load_college_distribution_results`` and returns xlsx or PDF bytes. The endpoint
layer is responsible for loading and authorizing the data.

``build_workbook`` and ``build_pdf`` share one source of truth for the column set
(``_headers``), per-row cell values (``_row_cells``) and column sizing
(``_COLUMNS`` weights), so the two formats never drift apart.

The table-rendering chassis (PDF styles, KeepInFrame cells, border/width
helpers) lives in ``app.services.export_table_chassis``, shared with
``manual_distribution_export_service``. ``college_ranking_export_service`` still
carries its own pre-extraction copy — a chassis fix likely belongs there too.

Unlike the 學生資料彙整表 export this carries NO PII (no 身分證字號, no 匯款帳號):
colleges get exactly what the 分發結果 panel already shows them.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.export_table_chassis import (
    PDF_HEADER_RESERVE_PT,
    PDF_MARGIN_PT,
    apply_xlsx_borders,
    apply_xlsx_column_widths,
    base_pdf_table_style,
    make_pdf_styles,
    pdf_cell,
    pdf_col_widths,
    pdf_paragraph,
    safe_str,
    style_xlsx_header_cell,
    write_xlsx_title_row,
)
from app.services.pdf_fonts import ensure_cjk_font
from app.utils.excel_safety import sanitize_excel_cell

# (header label, PDF column weight). HEADERS and the PDF weight vector are both
# derived from this one list so they can never drift — renaming a label can't
# silently mis-size a column. The normalize-to-page-width math lives in
# the chassis' pdf_col_widths.
_COLUMNS: List[Tuple[str, float]] = [
    ("類別", 1.4),
    ("結果", 0.7),
    ("名次", 0.6),
    ("學號", 1.2),
    ("姓名", 1.2),
    ("系所", 1.4),
    ("申請獎學金", 1.8),
]

HEADERS: List[str] = [label for label, _ in _COLUMNS]
_COL_WEIGHTS: List[float] = [weight for _, weight in _COLUMNS]


@dataclass(frozen=True)
class DistributionExportRow:
    """One student's outcome within one sub-type."""

    sub_type_label: str
    outcome: str  # 正取 / 未錄取
    position: Optional[int]
    student_number: str
    student_name: str
    department: str
    applied_scholarships: str


def _export_row(sub_type_label: str, outcome: str, student: Dict[str, Any]) -> DistributionExportRow:
    return DistributionExportRow(
        sub_type_label=sub_type_label,
        outcome=outcome,
        position=student.get("rank_position"),
        student_number=student.get("student_number") or "",
        student_name=student.get("student_name") or "",
        department=student.get("department") or "",
        applied_scholarships="、".join(student.get("applied_sub_types") or []),
    )


def _rank_key(student: Dict[str, Any]) -> tuple:
    rank = student.get("rank_position")
    return (rank is None, rank or 0)


def flatten_sub_types(sub_types: List[Dict[str, Any]]) -> List[DistributionExportRow]:
    """Flatten the grouped loader payload into export rows, mirroring the panel:
    per-sub-type 正取 blocks first, then every remaining student pooled into one
    未錄取 block sorted by college rank.

    Rejected students arrive grouped under their RANKING's sub-type code (e.g. a
    literal "default"), which is meaningless to the college — so 未錄取 rows carry
    "—" as 類別 instead of that label. Backend "backup" rows fold into 未錄取 too:
    nothing in the current distribution flow writes backup_allocations, so 備取 is
    not a real outcome.
    """
    admitted_rows: List[DistributionExportRow] = []
    leftover_students: List[Dict[str, Any]] = []
    for group in sub_types:
        label = group.get("label") or group.get("code") or ""
        for student in group.get("admitted") or []:
            admitted_rows.append(_export_row(label, "正取", student))
        leftover_students.extend(group.get("backup") or [])
        leftover_students.extend(group.get("rejected") or [])
    rejected_rows = [_export_row("—", "未錄取", s) for s in sorted(leftover_students, key=_rank_key)]
    return admitted_rows + rejected_rows


class CollegeDistributionExportService:
    """Builds 分發結果 workbooks and PDFs."""

    def build_workbook(self, *, rows: List[DistributionExportRow], title: str, sheet_name: str) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = self._headers()
        total_cols = len(headers)

        write_xlsx_title_row(ws, title=title, total_cols=total_cols)

        # Row 2: header
        for col_idx, header in enumerate(headers, start=1):
            style_xlsx_header_cell(ws.cell(row=2, column=col_idx, value=header))

        # Data rows — written from the same _row_cells used by the PDF export
        for idx, row in enumerate(rows, start=1):
            excel_row = idx + 2  # +2 because rows 1-2 are title/header
            for col_idx, value in enumerate(self._row_cells(row), start=1):
                # SECURITY: neutralize spreadsheet formula injection — openpyxl writes
                # a leading "=" as a LIVE formula and 姓名/系所 come from SIS.
                ws.cell(row=excel_row, column=col_idx, value=sanitize_excel_cell(value))

        max_row = len(rows) + 2
        apply_xlsx_borders(ws, max_row=max_row, max_col=total_cols)
        apply_xlsx_column_widths(ws, _COL_WEIGHTS)
        ws.freeze_panes = "A3"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def build_pdf(self, *, rows: List[DistributionExportRow], title: str) -> bytes:
        """Render the same 分發結果 table as an A4-landscape PDF.

        Mirrors ``build_workbook`` exactly (same columns, rows and ordering via
        ``_headers`` / ``_row_cells``). Column widths are normalised to the usable
        page width; rows paginate vertically with the header repeated per page.

        ``sanitize_excel_cell`` is deliberately NOT applied here: reportlab has no
        formula semantics, so the apostrophe prefix would be a visible artifact.
        This is the one place the two formats legitimately diverge.
        """
        ensure_cjk_font()

        headers = self._headers()

        page_width, page_height = landscape(A4)
        usable_width = page_width - (PDF_MARGIN_PT * 2)
        col_widths = self._pdf_col_widths(usable_width)
        cell_max_height = page_height - (PDF_MARGIN_PT * 2) - PDF_HEADER_RESERVE_PT

        title_style, header_style, cell_style = make_pdf_styles("Distribution")

        data: List[list] = [[pdf_paragraph(h, header_style) for h in headers]]
        for row in rows:
            values = self._row_cells(row)
            data.append(
                [
                    pdf_cell(safe_str(v), width=col_widths[col], max_height=cell_max_height, style=cell_style)
                    for col, v in enumerate(values)
                ]
            )

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(base_pdf_table_style()))

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=PDF_MARGIN_PT,
            rightMargin=PDF_MARGIN_PT,
            topMargin=PDF_MARGIN_PT,
            bottomMargin=PDF_MARGIN_PT,
        )
        doc.build([pdf_paragraph(title, title_style), Spacer(1, 4 * mm), table])
        return buf.getvalue()

    # -------- PDF layout helpers --------

    def _pdf_col_widths(self, usable_width: float) -> List[float]:
        # Unlike the ranking service the column set here is static, so the weights come
        # straight from _COL_WEIGHTS — no headers argument needed.
        return pdf_col_widths(_COL_WEIGHTS, usable_width)

    # -------- Shared column/value model (single source of truth) --------

    def _headers(self) -> List[str]:
        return list(HEADERS)

    def _row_cells(self, row: DistributionExportRow) -> List[Any]:
        """Ordered cell values for one row.

        The xlsx writer keeps the native int for 名次 (proper Excel typing); the PDF
        renderer stringifies it. Both share this list so the formats render identical
        content.
        """
        return [
            row.sub_type_label,
            row.outcome,
            row.position if row.position is not None else "",
            row.student_number,
            row.student_name,
            row.department,
            row.applied_scholarships,
        ]
