"""Admin 分發名單 (受獎名冊) export service.

Pure rendering logic for the manual-distribution 分發結果名單 export: receives
the allocated groups the ``/manual-distribution/distribution-summary`` endpoint
loads (one group per sub_type × 消耗年度配額) and returns xlsx or PDF bytes.
The endpoint layer is responsible for loading and authorizing the data.

The column layout follows the 國科會「補助大學校院推動研究生獎學金試辦方案」
受獎博士生名冊: 序號 / 基本資料 (學院・系所・姓名・國籍・性別・碩士畢業院校系所)
/ 學籍資料 (首次註冊入學日期・學號) / 請領資格檢核 (left blank for the 承辦人
to fill in). Three extra columns automate the parts of the 第三點第一款至四款
ineligibility check that the SIS snapshot can answer (see the ``derive_*``
functions); everything the SIS cannot see — e.g. outside full-time employment —
stays with the manual blank column.

The auto-check columns are derived from the ``Application.student_data``
snapshot taken at submission time, NOT a fresh SIS pull — a student who 休學
after being awarded still reads 在學 here. Roster generation re-verifies with
fresh SIS data; these columns are a screening aid, hence the group header
註明「依申請時學籍資料自動判定」.

``build_workbook`` and ``build_pdf`` share one source of truth for the column
set, group-header spans, per-row cell values and column sizing (``_COLUMNS``),
so the two formats never drift apart. The table chassis (PDF styles,
KeepInFrame cells, border/width helpers) comes from
``app.services.export_table_chassis``.

SCOPE — this layout is deliberately fixed, not configuration-driven. Unlike the
scholarship RULES (which are DB-driven per project CLAUDE.md §3), these columns
reproduce one external government form verbatim, so they are owned by 國科會 and
not by an administrator: making them configurable would let a config edit
produce a roster 國科會 rejects. The consequence is that this renderer is
PhD-flavoured — 受獎博士生首次註冊入學日期, 碩士畢業院/校/系所 and
``MASTER_SCHOOL_FIELD`` (a ``phd``-seeded dynamic field) — while the endpoint
applies it to whatever ``scholarship_type_id`` it is given. That is fine today
because the 分發 sub-type flow (nstc / moe_*) is the PhD scholarship, and every
field degrades to an empty cell rather than wrong data for other types. If a
second scholarship ever needs a DIFFERENT official roster form, select the
renderer per scholarship type here — do not parameterise these columns.

Carries NO high-sensitivity PII (no 身分證字號, no 匯款帳號), but 姓名/學號/國籍/
性別 plus the derived 學籍 flags are still personal data about identified
students, so the endpoint writes a ``pii_access`` AuditLog.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Any, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.college_mappings import COLLEGE_MAPPINGS
from app.services.export_table_chassis import (
    PDF_MARGIN_PT,
    apply_xlsx_borders,
    apply_xlsx_column_widths,
    base_pdf_table_style,
    make_pdf_styles,
    pdf_cell,
    pdf_cell_max_height,
    pdf_col_widths,
    pdf_paragraph,
    safe_str,
    style_xlsx_header_cell,
    write_xlsx_title_row,
)
from app.services.pdf_fonts import CJK_FONT_NAME, ensure_cjk_font
from app.utils.excel_safety import sanitize_excel_cell

# Snapshot coercion/rendering lives in a LEAF util so ManualDistributionService can
# share it without dragging reportlab/openpyxl into the core service import graph.
from app.utils.student_snapshot_fields import as_int as _as_int
from app.utils.student_snapshot_fields import as_text as _as_text
from app.utils.student_snapshot_fields import format_enrollment_date_roc

# -------- Column model (single source of truth for BOTH formats) --------

_BASIC_GROUP = "基本資料"
_REGISTRY_GROUP = "學籍資料"
_MANUAL_CHECK_GROUP = "請領資格檢核"
AUTO_CHECK_GROUP = "學籍系統檢核(依申請時學籍資料自動判定)"

# (group header or None, column header, PDF column weight). A None group means
# the column header spans both header rows (the form's 序號 column). HEADERS and
# the weight vector derive from this one list so they can never drift.
_COLUMNS: List[Tuple[Optional[str], str, float]] = [
    (None, "序號", 0.5),
    (_BASIC_GROUP, "學院", 0.9),
    (_BASIC_GROUP, "系所", 1.2),
    (_BASIC_GROUP, "姓名", 0.9),
    (_BASIC_GROUP, "國籍", 0.8),
    (_BASIC_GROUP, "性別", 0.5),
    (_BASIC_GROUP, "碩士畢業院/校/系所", 1.7),
    (_REGISTRY_GROUP, "受獎博士生首次註冊入學日期(民國年.月.日)", 1.1),
    (_REGISTRY_GROUP, "學號", 0.9),
    (_MANUAL_CHECK_GROUP, "有無本試辦方案第三點第一款至四款之情形(填寫有或無)", 1.2),
    (AUTO_CHECK_GROUP, "1.於公私立機構從事專職全時之有給職工作或以在職生身分報考者", 1.3),
    (AUTO_CHECK_GROUP, "2.以香港、澳門及大陸地區學生身分入學者", 1.2),
    (AUTO_CHECK_GROUP, "3.錄取後辦理休學、保留入學資格或未完成註冊者", 1.3),
]

HEADERS: List[str] = [label for _, label, _ in _COLUMNS]
_COL_WEIGHTS: List[float] = [weight for _, _, weight in _COLUMNS]

# -------- Auto-check verdicts --------

CHECK_FLAGGED = "有"
CHECK_CLEAR = "無"
CHECK_UNVERIFIABLE = "無法檢核"

# SIS code tables — seeded by alembic 6d5b1940bf8a (reference tables) and
# app/core/enroll_types.py. Only the codes this export interprets are pinned
# here; anything else is treated as "not flagged".
_SCHOOLID_IN_SERVICE = 2  # std_schoolid (school_identities): 2 = 在職生
_ENROLL_TYPES_IN_SERVICE = frozenset({2, 5})  # std_enrolltype: 招生考試在職生 / 推甄在職生
_IDENTITY_MAINLAND = 17  # std_identity (identities): 17 = 陸生
_ENROLL_TYPE_MAINLAND = 17  # std_enrolltype: 17 = 陸生
_STUDY_STATUS_FLAGGED = frozenset({4, 9, 10})  # studying_statuses: 休學 / 保留學籍 / 放棄入學
# 港澳 has NO SIS identity code (the identities table stops at 僑生/外籍生/陸生),
# so the only available signal is the free-text 國籍 / 僑居地 fields. Matched as
# SUBSTRINGS (and identically in both fields) because the SIS free text is not a
# controlled vocabulary — "中國", "中國大陸" and "中華人民共和國香港特別行政區" all
# occur. 中華民國 is safe against every keyword here: substrings are contiguous,
# and 中華民國 contains neither "中國" (中華民國 → 中華/華民/民國) nor "大陸".
_CROSS_STRAIT_KEYWORDS = ("中國", "中華人民共和國", "大陸", "香港", "澳門")


def derive_employment_check(student_data: dict) -> str:
    """試辦方案第三點第一款: 從事專職全時有給職工作「或以在職生身分報考」.

    Only the 在職生 half is visible to the SIS (身分別 or 入學管道); outside
    employment is not — that part stays with the manual 請領資格檢核 column.
    """
    schoolid = _as_int(student_data.get("std_schoolid"))
    enroll_type = _as_int(student_data.get("std_enrolltype"))
    if schoolid == _SCHOOLID_IN_SERVICE or enroll_type in _ENROLL_TYPES_IN_SERVICE:
        return CHECK_FLAGGED
    if schoolid is None and enroll_type is None:
        return CHECK_UNVERIFIABLE
    return CHECK_CLEAR


def derive_cross_strait_check(student_data: dict) -> str:
    """試辦方案第三點第二款: 以香港、澳門及大陸地區學生身分入學."""
    identity = _as_int(student_data.get("std_identity"))
    enroll_type = _as_int(student_data.get("std_enrolltype"))
    nation = _as_text(student_data.get("std_nation"))
    oversea = _as_text(student_data.get("std_overseaplace"))
    if identity == _IDENTITY_MAINLAND or enroll_type == _ENROLL_TYPE_MAINLAND:
        return CHECK_FLAGGED
    if any(k in nation or k in oversea for k in _CROSS_STRAIT_KEYWORDS):
        return CHECK_FLAGGED
    if identity is None and enroll_type is None and not nation and not oversea:
        return CHECK_UNVERIFIABLE
    return CHECK_CLEAR


def derive_study_status_check(student_data: dict) -> str:
    """試辦方案第三點第三款: 錄取後辦理休學、保留入學資格或未完成註冊.

    Reads BOTH the term-level and student-level status because either may lag
    the other in the snapshot; a flag on either counts.
    """
    term_status = _as_int(student_data.get("trm_studystatus"))
    student_status = _as_int(student_data.get("std_studingstatus"))
    if term_status in _STUDY_STATUS_FLAGGED or student_status in _STUDY_STATUS_FLAGGED:
        return CHECK_FLAGGED
    if term_status is None and student_status is None:
        return CHECK_UNVERIFIABLE
    return CHECK_CLEAR


# -------- Snapshot field rendering --------


def _gender_label(student_data: dict) -> str:
    sex = _as_int(student_data.get("std_sex"))
    if sex == 1:
        return "男"
    if sex == 2:
        return "女"
    return ""


def _college_label(student_data: dict) -> str:
    name = _as_text(student_data.get("trm_academyname"))
    if name:
        return name
    code = _as_text(student_data.get("std_academyno")) or _as_text(student_data.get("trm_academyno"))
    return COLLEGE_MAPPINGS.get(code, code)


MASTER_SCHOOL_FIELD = "master_school_info"


def _master_school_info(application: Any) -> str:
    """碩士畢業院/校/系所 — the student-typed dynamic form field carries the full
    校/院/系所 string; the SIS ``std_highestschname`` fallback knows the school
    name only."""
    form_data = getattr(application, "submitted_form_data", None) or {}
    fields_map = form_data.get("fields") if isinstance(form_data, dict) else None
    entry = fields_map.get(MASTER_SCHOOL_FIELD) if isinstance(fields_map, dict) else None
    if isinstance(entry, dict):
        value = entry.get("value")
        if value not in (None, ""):
            return str(value)
    student_data = getattr(application, "student_data", None) or {}
    return _as_text(student_data.get("std_highestschname"))


# -------- Row / group model --------


@dataclass(frozen=True)
class RecipientExportRow:
    """One awarded student, in form column order."""

    # 序號 is assigned by the CALLER and runs continuously across every group in
    # one export (it does NOT restart per sheet/page): the 名冊 is filed as a
    # single document, so 序號 N must identify one student unambiguously.
    seq: int
    college: str
    department: str
    student_name: str
    nationality: str
    gender: str
    master_school: str
    enrollment_date: str
    student_number: str
    employment_check: str
    cross_strait_check: str
    study_status_check: str


@dataclass(frozen=True)
class RecipientExportGroup:
    """One 分發 group: sub_type × 消耗配額年度."""

    label: str  # ScholarshipSubTypeConfig.name, falling back to the raw code
    sub_type_code: str
    allocation_year: Optional[int]
    rows: List[RecipientExportRow]


def build_recipient_row(seq: int, application: Any) -> RecipientExportRow:
    """Derive one export row from an Application's SIS snapshot + form data."""
    student_data = getattr(application, "student_data", None) or {}
    return RecipientExportRow(
        seq=seq,
        college=_college_label(student_data),
        department=_as_text(student_data.get("trm_depname")),
        student_name=_as_text(student_data.get("std_cname")),
        nationality=_as_text(student_data.get("std_nation")),
        gender=_gender_label(student_data),
        master_school=_master_school_info(application),
        enrollment_date=format_enrollment_date_roc(student_data),
        student_number=_as_text(student_data.get("std_stdcode")),
        employment_check=derive_employment_check(student_data),
        cross_strait_check=derive_cross_strait_check(student_data),
        study_status_check=derive_study_status_check(student_data),
    )


# -------- Header layout (shared by both formats) --------


def _group_runs() -> List[Tuple[Optional[str], int, int]]:
    """Consecutive (group, first_col, last_col) runs over _COLUMNS, 0-based."""
    runs: List[Tuple[Optional[str], int, int]] = []
    for idx, (group, _, _) in enumerate(_COLUMNS):
        if runs and runs[-1][0] is not None and runs[-1][0] == group:
            runs[-1] = (group, runs[-1][1], idx)
        else:
            runs.append((group, idx, idx))
    return runs


_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_SHEET_NAME_MAX = 31


def _sheet_name(group: RecipientExportGroup, used: set) -> str:
    base = group.label or group.sub_type_code or "分發名單"
    if group.allocation_year is not None:
        base = f"{base}_{group.allocation_year}"
    base = _INVALID_SHEET_CHARS.sub("_", base)[:_SHEET_NAME_MAX] or "分發名單"
    name = base
    suffix = 2
    while name in used:
        tail = f"_{suffix}"
        name = base[: _SHEET_NAME_MAX - len(tail)] + tail
        suffix += 1
    used.add(name)
    return name


def _group_heading(group: RecipientExportGroup) -> str:
    year_part = f"　{group.allocation_year}年度配額" if group.allocation_year is not None else ""
    return f"{group.label}({group.sub_type_code}){year_part}　共{len(group.rows)}人"


class ManualDistributionExportService:
    """Builds the admin 分發名單 recipient roster as xlsx or PDF."""

    def build_workbook(self, *, groups: Sequence[RecipientExportGroup], title: str) -> bytes:
        """One sheet per 分發 group, each with the two-row grouped form header."""
        wb = Workbook()
        wb.remove(wb.active)
        used_names: set = set()

        for group in groups:
            ws = wb.create_sheet(_sheet_name(group, used_names))
            self._write_sheet(ws, group.rows, title=f"{title}－{_group_heading(group)}")

        if not wb.sheetnames:
            # Zero groups: an openpyxl workbook with no sheets saves as a corrupt
            # file. The endpoint 404s before this, so this is renderer-level
            # robustness only: emit one empty roster sheet instead.
            self._write_sheet(wb.create_sheet("分發名單"), [], title=title)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _write_sheet(self, ws, rows: Sequence[RecipientExportRow], *, title: str) -> None:
        total_cols = len(_COLUMNS)

        write_xlsx_title_row(ws, title=title, total_cols=total_cols)

        # Rows 2-3: the form's grouped header. Group cells merge across their
        # columns; the un-grouped 序號 column merges down both rows instead.
        for grp, first, last in _group_runs():
            if grp is None:
                ws.cell(row=2, column=first + 1, value=_COLUMNS[first][1])
                ws.merge_cells(start_row=2, start_column=first + 1, end_row=3, end_column=first + 1)
            else:
                ws.cell(row=2, column=first + 1, value=grp)
                if last > first:
                    ws.merge_cells(start_row=2, start_column=first + 1, end_row=2, end_column=last + 1)
                for col in range(first, last + 1):
                    ws.cell(row=3, column=col + 1, value=_COLUMNS[col][1])
        for row_idx in (2, 3):
            for col_idx in range(1, total_cols + 1):
                style_xlsx_header_cell(ws.cell(row=row_idx, column=col_idx))

        # Data rows — written from the same _row_cells used by the PDF export.
        for offset, row in enumerate(rows):
            excel_row = offset + 4  # rows 1-3 are title + two header rows
            for col_idx, value in enumerate(self._row_cells(row), start=1):
                # SECURITY: neutralize spreadsheet formula injection — 姓名/系所/
                # 碩士畢業校系 are SIS- or student-supplied free text.
                ws.cell(row=excel_row, column=col_idx, value=sanitize_excel_cell(value))

        apply_xlsx_borders(ws, max_row=len(rows) + 3, max_col=total_cols)
        apply_xlsx_column_widths(ws, _COL_WEIGHTS)
        ws.freeze_panes = "A4"

    def build_pdf(self, *, groups: Sequence[RecipientExportGroup], title: str) -> bytes:
        """Render the same roster as an A4-landscape PDF, one page-run per group.

        Mirrors ``build_workbook`` exactly (same columns, spans, rows via
        ``_group_runs`` / ``_row_cells``). ``sanitize_excel_cell`` is deliberately
        NOT applied here: reportlab has no formula semantics, so the apostrophe
        prefix would be a visible artifact. This is the one place the two formats
        legitimately diverge.
        """
        ensure_cjk_font()

        page_width, page_height = landscape(A4)
        usable_width = page_width - (PDF_MARGIN_PT * 2)
        col_widths = pdf_col_widths(_COL_WEIGHTS, usable_width)

        title_style, header_style, cell_style = make_pdf_styles("RecipientRoster")

        # Cap body cells against the MEASURED 2-row header, not a fixed reserve:
        # this table's repeated header is ~twice the single-row one the chassis
        # constant was sized for, and a cap even slightly too generous makes
        # reportlab raise LayoutError on free text as short as 700 characters.
        cell_max_height = pdf_cell_max_height(
            page_height,
            header_height=self._measure_header_height(
                col_widths=col_widths, header_style=header_style, page_height=page_height
            ),
        )
        section_style = ParagraphStyle(
            "RecipientRosterPdfSection",
            fontName=CJK_FONT_NAME,
            fontSize=10,
            leading=13,
            wordWrap="CJK",
        )

        elements: List[Any] = [pdf_paragraph(title, title_style), Spacer(1, 4 * mm)]
        for group_idx, group in enumerate(groups):
            if group_idx > 0:
                elements.append(PageBreak())
            elements.append(pdf_paragraph(_group_heading(group), section_style))
            elements.append(Spacer(1, 2 * mm))
            elements.append(
                self._group_table(
                    group,
                    col_widths=col_widths,
                    cell_max_height=cell_max_height,
                    header_style=header_style,
                    cell_style=cell_style,
                )
            )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=PDF_MARGIN_PT,
            rightMargin=PDF_MARGIN_PT,
            topMargin=PDF_MARGIN_PT,
            bottomMargin=PDF_MARGIN_PT,
        )
        doc.build(elements)
        return buf.getvalue()

    # -------- PDF table assembly --------

    def _header_rows(self, header_style: ParagraphStyle) -> Tuple[List[Any], List[Any], List[tuple]]:
        """The two grouped header rows + the SPAN commands that shape them.

        Shared by the real table and the measurement probe so the height the cap
        is derived from is the height actually rendered.
        """
        group_row: List[Any] = [""] * len(_COLUMNS)
        label_row: List[Any] = [""] * len(_COLUMNS)
        span_commands: List[tuple] = []
        for grp, first, last in _group_runs():
            if grp is None:
                group_row[first] = pdf_paragraph(_COLUMNS[first][1], header_style)
                span_commands.append(("SPAN", (first, 0), (first, 1)))
            else:
                group_row[first] = pdf_paragraph(grp, header_style)
                if last > first:
                    span_commands.append(("SPAN", (first, 0), (last, 0)))
                for col in range(first, last + 1):
                    label_row[col] = pdf_paragraph(_COLUMNS[col][1], header_style)
        return group_row, label_row, span_commands

    def _measure_header_height(
        self, *, col_widths: List[float], header_style: ParagraphStyle, page_height: float
    ) -> float:
        """Rendered height of the repeated 2-row header, in points.

        Measured rather than assumed: the headers are long CJK strings whose wrap
        depends on the column widths, and a guessed reserve that is even slightly
        too small makes ``build_pdf`` raise ``LayoutError`` on ordinary free text.
        """
        group_row, label_row, span_commands = self._header_rows(header_style)
        probe = Table([group_row, label_row], colWidths=col_widths)
        probe.setStyle(TableStyle(base_pdf_table_style(header_rows=2) + span_commands))
        _, height = probe.wrap(sum(col_widths), page_height)
        return height

    def _group_table(
        self,
        group: RecipientExportGroup,
        *,
        col_widths: List[float],
        cell_max_height: float,
        header_style: ParagraphStyle,
        cell_style: ParagraphStyle,
    ) -> Table:
        group_row, label_row, span_commands = self._header_rows(header_style)

        data: List[list] = [group_row, label_row]
        for row in group.rows:
            data.append(
                [
                    pdf_cell(safe_str(v), width=col_widths[col], max_height=cell_max_height, style=cell_style)
                    for col, v in enumerate(self._row_cells(row))
                ]
            )

        table = Table(data, colWidths=col_widths, repeatRows=2)
        table.setStyle(TableStyle(base_pdf_table_style(header_rows=2) + span_commands))
        return table

    # -------- Shared column/value model (single source of truth) --------

    def _row_cells(self, row: RecipientExportRow) -> List[Any]:
        """Ordered cell values for one row.

        The xlsx writer keeps the native int for 序號 (proper Excel typing); the
        PDF renderer stringifies it. The 請領資格檢核 column is deliberately
        empty — the 承辦人 fills it in by hand (the SIS cannot answer 第四款,
        already-receiving-another-government-scholarship, at all).
        """
        return [
            row.seq,
            row.college,
            row.department,
            row.student_name,
            row.nationality,
            row.gender,
            row.master_school,
            row.enrollment_date,
            row.student_number,
            "",  # 請領資格檢核 — manual
            row.employment_check,
            row.cross_strait_check,
            row.study_status_check,
        ]
