"""College ranking export service.

Pure rendering logic — receives prepared data structures, returns xlsx or PDF
bytes. The endpoint layer is responsible for loading rows, dynamic field configs,
and sub-type labels from the database.

``build_workbook`` and ``build_pdf`` share one source of truth for the column set
(``_headers``) and per-row cell values (``_row_cells``), so the two formats never
drift apart.
"""

from __future__ import annotations

import io
import math
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

# `escape` is a pure string-escaping helper (`<` → `&lt;` …) used to sanitise
# cell values before they go into reportlab Paragraph markup. It does not parse
# untrusted XML, so the B406 warning is a false positive here.
from xml.sax.saxutils import escape as xml_escape  # nosec B406

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import KeepInFrame, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.pdf_fonts import CJK_FONT_NAME, ensure_cjk_font
from app.utils.excel_security import excel_safe_cell_value

# Static columns shared by the xlsx and PDF exports: (header label, PDF column
# weight). STATIC_HEADERS and the PDF weight vector are both derived from this one
# list so they can never drift — renaming a label can't silently mis-size a column.
# Weights are relative (narrow id/flag columns get less, free-text columns more);
# the normalize-to-page-width math lives in _pdf_col_widths. Dynamic columns use
# _PDF_COL_WEIGHT_DEFAULT.
_STATIC_COLUMNS: List[Tuple[str, float]] = [
    ("NO.", 0.5),
    ("學院初審會議之學院排序", 0.8),
    ("申請獎學金類別", 1.4),
    ("學院", 1.0),
    ("系所", 1.2),
    ("年級", 0.5),
    ("是否為逕博學生", 0.7),
    ("學生中文姓名", 1.0),
    ("學生英文姓名", 1.4),
    ("國籍", 0.7),
    ("性別", 0.5),
    ("註冊入學日期", 0.9),
    ("學號", 1.0),
    ("學生身分證字號", 1.1),
    ("學生匯款帳號", 1.2),
    ("學生E-mail", 1.7),
    ("學生通訊地址", 2.0),
    ("指導教授姓名", 1.3),
]

STATIC_HEADERS: List[str] = [label for label, _ in _STATIC_COLUMNS]
_STATIC_COL_WEIGHTS: List[float] = [weight for _, weight in _STATIC_COLUMNS]


# std_enrolltype codes that indicate 逕讀博士 (direct-track PhD)
DIRECT_PHD_ENROLLTYPE_CODES = {8, 9, 10, 11}


@dataclass(frozen=True)
class DynamicFieldSpec:
    field_name: str
    field_label: str
    export_column_label: Optional[str]
    display_order: int


@dataclass
class ExportRow:
    """One ranked application's data."""

    rank_position: Optional[int]
    application: Any  # Duck-typed: needs sub_type_preferences, sub_scholarship_type, student_data, submitted_form_data
    bank_account: Optional[str] = None  # 郵局帳號 (from user_profiles.account_number)
    advisor_names: Optional[str] = None  # 指導教授姓名 (comma-joined if multiple advisors)


class CollegeRankingExportService:
    """Builds 學生資料彙整表 workbooks."""

    def build_workbook(
        self,
        *,
        rows: List[ExportRow],
        dynamic_fields: List[DynamicFieldSpec],
        sub_type_labels: Dict[str, str],
        title: str,
        sheet_name: str,
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        sorted_dynamic = self._sort_dynamic(dynamic_fields)
        headers = self._headers(sorted_dynamic)
        total_cols = len(headers)

        # Row 1: title (merged across all columns)
        ws.cell(row=1, column=1, value=title)
        if total_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row 2: header
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")

        # Data rows — written from the same _row_cells used by the PDF export
        for idx, row in enumerate(rows, start=1):
            excel_row = idx + 2  # +2 because rows 1-2 are title/header
            for col_idx, value in enumerate(self._row_cells(row, idx, sub_type_labels, sorted_dynamic), start=1):
                # SECURITY (#1081-G): free-text form fields flow into these cells;
                # neutralize spreadsheet formula injection before writing.
                ws.cell(row=excel_row, column=col_idx, value=excel_safe_cell_value(value))

        max_row = len(rows) + 2
        self._apply_borders(ws, max_row=max_row, max_col=total_cols)
        self._apply_column_widths(ws, headers)
        ws.freeze_panes = "A3"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def build_pdf(
        self,
        *,
        rows: List[ExportRow],
        dynamic_fields: List[DynamicFieldSpec],
        sub_type_labels: Dict[str, str],
        title: str,
    ) -> bytes:
        """Render the same 學生資料彙整表 as an A4-landscape PDF.

        Mirrors ``build_workbook`` exactly (same columns, rows and ordering via
        ``_headers`` / ``_row_cells``). Column widths are normalised to the usable
        page width so the wide table never overflows horizontally — long cell text
        wraps, and rows paginate vertically with the header repeated on each page.
        """
        ensure_cjk_font()

        sorted_dynamic = self._sort_dynamic(dynamic_fields)
        headers = self._headers(sorted_dynamic)

        page_width, page_height = landscape(A4)
        usable_width = page_width - (self._PDF_MARGIN_PT * 2)
        col_widths = self._pdf_col_widths(headers, usable_width)
        # A reportlab Table cannot split ONE row across pages, so a single very
        # long free-text cell (e.g. a verbose dynamic field) would raise
        # LayoutError and fail the whole export. Cap each cell to the usable
        # content height and let KeepInFrame shrink anything taller to fit.
        cell_max_height = page_height - (self._PDF_MARGIN_PT * 2) - self._PDF_HEADER_RESERVE_PT

        title_style = ParagraphStyle(
            "RankingPdfTitle",
            fontName=CJK_FONT_NAME,
            fontSize=12,
            leading=15,
            alignment=1,  # center
        )
        header_style = ParagraphStyle(
            "RankingPdfHeader",
            fontName=CJK_FONT_NAME,
            fontSize=6.5,
            leading=8,
            alignment=1,
            wordWrap="CJK",
        )
        cell_style = ParagraphStyle(
            "RankingPdfCell",
            fontName=CJK_FONT_NAME,
            fontSize=6,
            leading=7.5,
            wordWrap="CJK",  # break long CJK and unspaced ASCII (emails, IDs)
        )

        data: List[list] = [[Paragraph(xml_escape(h), header_style) for h in headers]]
        for idx, row in enumerate(rows, start=1):
            values = self._row_cells(row, idx, sub_type_labels, sorted_dynamic)
            data.append(
                [
                    KeepInFrame(
                        col_widths[col],
                        cell_max_height,
                        [Paragraph(xml_escape(self._safe_str(v)), cell_style)],
                        mode="shrink",
                    )
                    for col, v in enumerate(values)
                ]
            )

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.Color(0.6, 0.6, 0.6)),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.87, 0.87, 0.87)),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, -1), CJK_FONT_NAME),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 1),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ]
            )
        )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=self._PDF_MARGIN_PT,
            rightMargin=self._PDF_MARGIN_PT,
            topMargin=self._PDF_MARGIN_PT,
            bottomMargin=self._PDF_MARGIN_PT,
        )
        doc.build([Paragraph(xml_escape(title), title_style), Spacer(1, 4 * mm), table])
        return buf.getvalue()

    # -------- PDF layout helpers --------

    _PDF_MARGIN_PT = 10 * mm
    # Vertical space reserved (per page) for the title + spacer + repeated header
    # row, subtracted from page height to bound how tall a single data cell may be.
    _PDF_HEADER_RESERVE_PT = 60
    # Width weight for dynamic (form-field) columns; static columns carry their own
    # weights in _STATIC_COL_WEIGHTS (derived from _STATIC_COLUMNS).
    _PDF_COL_WEIGHT_DEFAULT = 1.2

    def _pdf_col_widths(self, headers: List[str], usable_width: float) -> List[float]:
        # headers is always STATIC_HEADERS + dynamic, so the first N columns map
        # index-for-index onto _STATIC_COL_WEIGHTS; the rest are dynamic.
        n_static = len(_STATIC_COL_WEIGHTS)
        weights = [
            _STATIC_COL_WEIGHTS[i] if i < n_static else self._PDF_COL_WEIGHT_DEFAULT for i in range(len(headers))
        ]
        total = sum(weights) or 1.0
        return [usable_width * w / total for w in weights]

    # -------- Shared column/value model (single source of truth) --------

    @staticmethod
    def _sort_dynamic(dynamic_fields: List[DynamicFieldSpec]) -> List[DynamicFieldSpec]:
        return sorted(dynamic_fields, key=lambda f: (f.display_order, f.field_name))

    def _headers(self, sorted_dynamic: List[DynamicFieldSpec]) -> List[str]:
        return STATIC_HEADERS + [(f.export_column_label or f.field_label) for f in sorted_dynamic]

    def _row_cells(
        self,
        row: ExportRow,
        row_index: int,
        sub_type_labels: Dict[str, str],
        sorted_dynamic: List[DynamicFieldSpec],
    ) -> List[Any]:
        """Ordered cell values for one row: static columns then dynamic columns.

        The xlsx writer keeps the native int values (NO., rank, grade) for proper
        Excel typing; the PDF renderer stringifies them. Both share this list so
        the two formats render identical content.
        """
        return self._static_values(row, row_index, sub_type_labels) + self._dynamic_values(row, sorted_dynamic)

    def _static_values(
        self,
        row: ExportRow,
        row_index: int,
        sub_type_labels: Dict[str, str],
    ) -> List[Any]:
        sd = getattr(row.application, "student_data", None) or {}
        return [
            row_index,
            row.rank_position if row.rank_position is not None else "",
            self._render_scholarship_type(row.application, sub_type_labels),
            self._safe_str(sd.get("trm_academyname")),
            self._safe_str(sd.get("trm_depname")),
            self._compute_grade(sd.get("trm_termcount")),
            self._render_direct_phd(sd.get("std_enrolltype")),
            self._safe_str(sd.get("std_cname")),
            self._safe_str(sd.get("std_ename")),
            self._safe_str(sd.get("std_nation")),
            self._render_gender(sd.get("std_sex")),
            self._render_enrollment_date(sd.get("std_enrollyear"), sd.get("std_enrollterm")),
            self._safe_str(sd.get("std_stdcode")),
            self._safe_str(sd.get("std_pid")),
            self._safe_str(row.bank_account),
            self._safe_str(sd.get("com_email")),
            self._safe_str(sd.get("com_commadd")),
            self._safe_str(row.advisor_names),
        ]

    def _safe_str(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value)

    def _render_gender(self, std_sex: Any) -> str:
        if std_sex == 1:
            return "男"
        if std_sex == 2:
            return "女"
        return ""

    def _render_direct_phd(self, std_enrolltype: Any) -> str:
        try:
            code = int(std_enrolltype)
        except (TypeError, ValueError):
            return ""
        return "是" if code in DIRECT_PHD_ENROLLTYPE_CODES else "否"

    def _compute_grade(self, trm_termcount: Any):
        try:
            tc = int(trm_termcount)
        except (TypeError, ValueError):
            return ""
        if tc <= 0:
            return ""
        return math.ceil(tc / 2)

    def _render_enrollment_date(self, year: Any, term: Any) -> str:
        if year in (None, ""):
            return ""
        try:
            year_int = int(year)
        except (TypeError, ValueError):
            return ""
        month = 9 if term in (None, "", 1, "1") else 2
        return f"{year_int}.{month}.1"

    def _render_scholarship_type(self, application: Any, sub_type_labels: Dict[str, str]) -> str:
        prefs: Optional[List[str]] = getattr(application, "sub_type_preferences", None)
        if prefs:
            labels = [sub_type_labels.get(code, code) for code in prefs]
            if len(labels) == 1:
                return labels[0]
            return f"{labels[0]}(第一志願)暨{labels[1]}(第二志願)"

        fallback_code = getattr(application, "sub_scholarship_type", None)
        if not fallback_code:
            return ""
        return sub_type_labels.get(fallback_code, fallback_code)

    # -------- Dynamic column rendering --------

    def _dynamic_values(self, row: ExportRow, sorted_dynamic: List[DynamicFieldSpec]) -> List[str]:
        form_data = getattr(row.application, "submitted_form_data", None) or {}
        fields_map = form_data.get("fields") if isinstance(form_data, dict) else None
        fields_map = fields_map if isinstance(fields_map, dict) else {}
        return [self._extract_dynamic_value(fields_map, spec.field_name) for spec in sorted_dynamic]

    def _extract_dynamic_value(self, fields_map: Dict[str, Any], field_name: str) -> str:
        entry = fields_map.get(field_name)
        if not isinstance(entry, dict):
            return ""
        raw = entry.get("value")
        if raw is None or raw == "":
            return ""
        return str(raw)

    # -------- Formatting --------

    def _apply_borders(self, ws, *, max_row: int, max_col: int) -> None:
        thin = Side(style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).border = border

    def _apply_column_widths(self, ws, headers: List[str]) -> None:
        for idx, header in enumerate(headers, start=1):
            text_len = max(len(str(header)), 6)
            width = min(max(text_len + 4, 10), 30)
            ws.column_dimensions[get_column_letter(idx)].width = width
