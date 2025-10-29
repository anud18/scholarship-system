"""
Whitelist Excel Import/Export Service
"""

import io
import logging
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class WhitelistExcelService:
    """白名單 Excel 匯入/匯出服務"""

    def __init__(self):
        # Excel 欄位定義
        self.column_headers = ["學號", "姓名", "子獎學金類型", "備註"]
        self.column_mapping = {"nycu_id": 0, "name": 1, "sub_type": 2, "note": 3}

        # 樣式定義
        self.header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        self.cell_font = Font(name="微軟正黑體", size=10)
        self.cell_alignment = Alignment(horizontal="left", vertical="center")

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def export_whitelist(
        self, whitelist_data: Dict[str, List[Dict[str, str]]], scholarship_name: str = "獎學金"
    ) -> io.BytesIO:
        """
        匯出白名單為 Excel 檔案

        Args:
            whitelist_data: 白名單資料，格式: {"sub_type": [{"nycu_id": "xxx", "name": "xxx", "note": "xxx"}]}
            scholarship_name: 獎學金名稱

        Returns:
            BytesIO: Excel 文件內容
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "白名單"

        # 寫入標題行
        for col_idx, header in enumerate(self.column_headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 寫入資料
        row = 2
        for sub_type, students in whitelist_data.items():
            for student in students:
                ws.cell(row=row, column=1).value = student.get("nycu_id", "")
                ws.cell(row=row, column=2).value = student.get("name", "")
                ws.cell(row=row, column=3).value = sub_type
                ws.cell(row=row, column=4).value = student.get("note", "")

                # 應用樣式
                for col_idx in range(1, len(self.column_headers) + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.font = self.cell_font
                    cell.alignment = self.cell_alignment
                    cell.border = self.border

                row += 1

        # 調整欄寬
        column_widths = [15, 20, 25, 30]  # 學號、姓名、子獎學金類型、備註
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # 儲存到 BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    def parse_import_excel(
        self, file_content: bytes, valid_sub_types: List[str]
    ) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
        """
        解析匯入的 Excel 檔案

        Args:
            file_content: Excel 文件內容
            valid_sub_types: 有效的子獎學金類型列表

        Returns:
            Tuple[成功的資料列表, 錯誤列表]
        """
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active

        success_data = []
        errors = []

        # 驗證標題行
        header_row = [cell.value for cell in ws[1]]
        # Expected headers: ["學號", "姓名", "子獎學金類型", "備註"]

        # 靈活匹配標題（允許部分標題缺失）
        if not any(h in header_row for h in ["學號", "子獎學金類型"]):
            errors.append({"row": "1", "nycu_id": "", "error": "Excel 格式錯誤：缺少必要欄位（學號、子獎學金類型）"})
            return success_data, errors

        # 找出欄位索引
        col_indices = {}
        for idx, header in enumerate(header_row):
            if header == "學號":
                col_indices["nycu_id"] = idx
            elif header == "姓名":
                col_indices["name"] = idx
            elif header == "子獎學金類型":
                col_indices["sub_type"] = idx
            elif header == "備註":
                col_indices["note"] = idx

        # 解析資料行（從第2行開始）
        seen_students = set()  # 記錄已處理的學號，避免重複

        for row_num in range(2, ws.max_row + 1):
            row_data = [cell.value for cell in ws[row_num]]

            # 跳過空行
            if all(cell is None or str(cell).strip() == "" for cell in row_data):
                continue

            # 提取資料
            nycu_id = str(row_data[col_indices["nycu_id"]]).strip() if "nycu_id" in col_indices else ""
            sub_type = str(row_data[col_indices["sub_type"]]).strip() if "sub_type" in col_indices else ""
            name = (
                str(row_data[col_indices["name"]]).strip()
                if "name" in col_indices and row_data[col_indices["name"]]
                else ""
            )
            note = (
                str(row_data[col_indices["note"]]).strip()
                if "note" in col_indices and row_data[col_indices["note"]]
                else ""
            )

            # 驗證必填欄位
            if not nycu_id:
                errors.append({"row": str(row_num), "nycu_id": "", "error": "學號不能為空"})
                continue

            if not sub_type:
                errors.append({"row": str(row_num), "nycu_id": nycu_id, "error": "子獎學金類型不能為空"})
                continue

            # 驗證子獎學金類型
            if sub_type not in valid_sub_types:
                errors.append(
                    {
                        "row": str(row_num),
                        "nycu_id": nycu_id,
                        "error": f"無效的子獎學金類型: {sub_type}，有效值: {', '.join(valid_sub_types)}",
                    }
                )
                continue

            # 檢查重複
            student_key = f"{nycu_id}_{sub_type}"
            if student_key in seen_students:
                errors.append(
                    {"row": str(row_num), "nycu_id": nycu_id, "error": f"重複的學號-子類型組合: {nycu_id} - {sub_type}"}
                )
                continue

            seen_students.add(student_key)

            # 加入成功資料
            success_data.append({"nycu_id": nycu_id, "sub_type": sub_type, "name": name, "note": note})

        return success_data, errors

    def generate_template(self, sub_types: List[str]) -> io.BytesIO:
        """
        生成白名單匯入模板

        Args:
            sub_types: 子獎學金類型列表

        Returns:
            BytesIO: 模板 Excel 文件
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "白名單匯入模板"

        # 寫入標題行
        for col_idx, header in enumerate(self.column_headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 寫入範例資料
        example_data = [
            ["0856001", "王小明", sub_types[0] if sub_types else "general", "範例備註"],
            ["0856002", "李小華", sub_types[1] if len(sub_types) > 1 else sub_types[0] if sub_types else "general", ""],
        ]

        for row_idx, data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = self.cell_font
                cell.alignment = self.cell_alignment
                cell.border = self.border

        # 調整欄寬
        column_widths = [15, 20, 25, 30]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # 新增說明工作表
        ws_info = wb.create_sheet("使用說明")
        instructions = [
            ["白名單匯入格式說明"],
            [""],
            ["1. 必填欄位：學號、子獎學金類型"],
            ["2. 選填欄位：姓名、備註"],
            [""],
            ["3. 有效的子獎學金類型："],
        ]

        for sub_type in sub_types:
            instructions.append([f"   - {sub_type}"])

        instructions.extend(
            [
                [""],
                ["4. 注意事項："],
                ["   - 請勿更改標題列"],
                ["   - 學號請使用正確的學校學號"],
                ["   - 同一學號可以在不同子獎學金類型中出現"],
                ["   - 重複的學號-子類型組合將被忽略"],
            ]
        )

        for row_idx, data in enumerate(instructions, start=1):
            for col_idx, value in enumerate(data, start=1):
                ws_info.cell(row=row_idx, column=col_idx).value = value

        # 儲存到 BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file


# Service instance
whitelist_excel_service = WhitelistExcelService()
