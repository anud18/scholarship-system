"""Batch-import template builder.

ONE generator for the 批次匯入 example workbook, shared by the admin
批次匯入 endpoint and the college 補充匯入 endpoint. 補充匯入 replaced 批次匯入
for colleges and must offer an identical file — if the two ever built their
own columns, a template downloaded from one panel would stop importing
through the other.
"""

import logging
from io import BytesIO
from typing import Any, Dict, List, Tuple

import pandas as pd
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.application_field import ApplicationField
from app.models.scholarship import ScholarshipType

logger = logging.getLogger(__name__)

TEMPLATE_SHEET_NAME = "批次匯入範例"


async def build_batch_import_template(db: AsyncSession, scholarship: ScholarshipType) -> bytes:
    """Render the batch-import example workbook for one scholarship type.

    Columns: 學號 / 學生姓名 / 郵局帳號, the advisor trio when the scholarship
    requires a professor recommendation, one checkmark column per sub-type,
    then the scholarship's active custom fields — plus two sample rows.
    """
    # Define base columns (Traditional Chinese)
    columns = [
        "學號",  # student_id - 必填
        "學生姓名",  # student_name - 必填
        "郵局帳號",  # postal_account - 可選
    ]

    # Mapping for internal use (Chinese to English)
    column_mapping = {
        "學號": "student_id",
        "學生姓名": "student_name",
        "郵局帳號": "postal_account",
    }

    # Check if scholarship requires professor recommendation for advisor fields
    from app.services.application_field_service import ApplicationFieldService

    field_service = ApplicationFieldService(db)
    requires_advisor = await field_service.check_requires_professor_recommendation(scholarship.code)

    # Add advisor fixed fields if required (after postal_account, before sub_types)
    if requires_advisor:
        columns.extend(
            [
                "指導教授姓名",  # advisor_name
                "指導教授Email",  # advisor_email
                "指導教授本校人事編號",  # advisor_nycu_id
            ]
        )
        column_mapping.update(
            {
                "指導教授姓名": "advisor_name",
                "指導教授Email": "advisor_email",
                "指導教授本校人事編號": "advisor_nycu_id",
            }
        )

    # Sub-type label mapping — inverted from the parser's shared constant so
    # a downloaded template is always importable (labels can never drift).
    from app.services.batch_import_service import SUB_TYPE_CODE_BY_LABEL

    sub_type_labels = {code: label for label, code in SUB_TYPE_CODE_BY_LABEL.items()}

    # Add sub_type columns if scholarship has sub types (Traditional Chinese)
    if scholarship.sub_type_list:
        for sub_type_code in scholarship.sub_type_list:
            label = sub_type_labels.get(sub_type_code, sub_type_code)
            columns.append(label)
            column_mapping[label] = f"sub_type_{sub_type_code}"

    # Query custom fields for this scholarship type
    custom_fields_stmt = (
        select(ApplicationField)
        .where(ApplicationField.scholarship_type == scholarship.code)
        .where(ApplicationField.is_active)
        .order_by(ApplicationField.display_order)
    )
    custom_fields_result = await db.execute(custom_fields_stmt)
    custom_fields = custom_fields_result.scalars().all()

    # Add custom field columns (Traditional Chinese), skipping any that would
    # duplicate a base/fixed column already present. postal_account and the
    # advisor fields are also seeded as ApplicationFields, so without this the
    # column list gets duplicate labels — which makes pandas return a DataFrame
    # for df[label] and breaks the column-width pass with
    # "'DataFrame' object has no attribute 'tolist'".
    reserved_field_names = {"student_id", "student_name", "postal_account"}
    if requires_advisor:
        reserved_field_names.update({"advisor_name", "advisor_email", "advisor_nycu_id"})

    template_custom_fields = []
    for field in custom_fields:
        if field.field_name in reserved_field_names or field.field_label in columns:
            continue
        template_custom_fields.append(field)
        columns.append(field.field_label)  # Use Chinese label
        column_mapping[field.field_label] = f"custom_{field.field_name}"

    # Create sample data (2 example rows)
    sample_data = [
        {
            "學號": "111111111",
            "學生姓名": "王小明",
            "郵局帳號": "1234567890123",
        },
        {
            "學號": "222222222",
            "學生姓名": "陳小華",
            "郵局帳號": "9876543210987",
        },
    ]

    # Add advisor field sample values if required
    if requires_advisor:
        sample_data[0].update(
            {
                "指導教授姓名": "張教授",
                "指導教授Email": "professor.chang@nycu.edu.tw",
                "指導教授本校人事編號": "P001234",
            }
        )
        sample_data[1].update(
            {
                "指導教授姓名": "李教授",
                "指導教授Email": "professor.lee@nycu.edu.tw",
                "指導教授本校人事編號": "P005678",
            }
        )

    # Add sub_type sample values if applicable.
    # Sub-type cells are checkmarks: 1 (or V/✓) = applying for that category,
    # 0 or blank = not applying. Preference order is NOT read from these
    # cells — the system forces MOE (moe_1w) as first preference, mirroring
    # the student wizard. The two sample rows deliberately contrast 1 and 0
    # so the semantics are visible in the file itself; the header comments
    # added below spell them out.
    if scholarship.sub_type_list:
        for row_index, row in enumerate(sample_data):
            for st_index, sub_type_code in enumerate(scholarship.sub_type_list):
                label = sub_type_labels.get(sub_type_code, sub_type_code)
                row[label] = 1 if row_index == 0 or st_index == 0 else 0

    # Add custom field sample values
    for field in template_custom_fields:
        for i, row in enumerate(sample_data):
            # Provide sample values based on field type
            if field.field_type == "text":
                row[field.field_label] = f"範例{field.field_label}{i + 1}"
            elif field.field_type == "number":
                row[field.field_label] = 100 + i
            elif field.field_type == "select":
                # Use first option if available
                if field.field_options and len(field.field_options) > 0:
                    row[field.field_label] = field.field_options[0].get("label", "")
                else:
                    row[field.field_label] = ""
            elif field.field_type == "checkbox":
                row[field.field_label] = "Y" if i == 0 else ""
            else:
                row[field.field_label] = ""

    # Create DataFrame
    df = pd.DataFrame(sample_data, columns=columns)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="批次匯入範例")

        # Auto-adjust column widths
        from openpyxl.comments import Comment
        from openpyxl.utils import get_column_letter

        # Sub-type headers (國科會/教育部…) get a hover comment explaining the
        # checkmark semantics — the 1/0 cell values alone don't tell staff
        # what they mean.
        SUB_TYPE_COMMENT_TEXT = "1 = 有申請此類別；0 或空白 = 未申請（亦可填 V 或 ✓）"
        SUB_TYPE_COMMENT_BOX_HEIGHT = 80
        SUB_TYPE_COMMENT_BOX_WIDTH = 280
        sub_type_column_labels = {sub_type_labels.get(code, code) for code in (scholarship.sub_type_list or [])}

        worksheet = writer.sheets["批次匯入範例"]

        # SECURITY (#1081 G / #1223 A): df.to_excel assigns through openpyxl, so a
        # value beginning with =/+/-/@ would be written as a LIVE formula. Sweep the
        # sheet pandas just produced.
        #
        # min_row=2 skips the HEADER row deliberately: this template is downloaded,
        # filled in and re-uploaded, and build_submitted_form_data matches columns
        # by exact header string (custom_field_mapping). An apostrophe-prefixed
        # header would match nothing on re-upload and every value in that column
        # would be silently dropped. The headers are admin-authored field labels
        # (dynamic form configuration), not applicant free-text.
        from app.utils.excel_safety import neutralise_worksheet

        neutralise_worksheet(worksheet, min_row=2)

        for idx, col in enumerate(df.columns, 1):
            # Calculate max length for this column. Use positional access so a
            # duplicate column label can never turn df[col] into a DataFrame
            # (which has no .tolist()).
            column_values = df.iloc[:, idx - 1].astype(str).tolist()

            # Collect all content in this column (header + all data)
            all_content = [str(col)] + column_values

            # Calculate max character length
            max_length = max(len(text) for text in all_content) if all_content else 0

            # Count Chinese characters in each cell and find the max
            # Chinese characters need approximately 2x the width of English characters
            max_chinese_in_cell = (
                max(sum(1 for c in text if "\u4e00" <= c <= "\u9fff") for text in all_content) if all_content else 0
            )

            # Adjusted width calculation:
            # - Base width from character count
            # - Add extra width for Chinese characters (they're wider)
            # - Add padding
            adjusted_width = max_length + max_chinese_in_cell * 1.2 + 2

            # Apply width to column
            column_letter = get_column_letter(idx)
            worksheet.column_dimensions[column_letter].width = adjusted_width

            if col in sub_type_column_labels:
                worksheet.cell(row=1, column=idx).comment = Comment(
                    SUB_TYPE_COMMENT_TEXT,
                    "獎學金系統",
                    height=SUB_TYPE_COMMENT_BOX_HEIGHT,
                    width=SUB_TYPE_COMMENT_BOX_WIDTH,
                )

    output.seek(0)
    return output.getvalue()
