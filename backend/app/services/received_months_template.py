"""
Builds the 匯入已領月份數 example workbook.

A faithful reproduction of 國科會's blank 「獲獎生已領月份統計表」: a merged
title row, the header row, 國科會's own 範例 demonstration row, and pre-numbered
empty rows for the admin to fill in.

The header row is built from the parser's own constants, so the template and the
parser cannot drift apart — see test_received_months_template.py, which feeds
this workbook straight back through the parser.
"""

from io import BytesIO
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.services.received_months_parser import (
    EXAMPLE_ROW_MARKER,
    HEADER_CURRENT_MONTH,
    HEADER_START_MONTH,
    HEADER_STATED_TOTAL,
    HEADER_STUDENT_NUMBER,
)

SHEET_TITLE = "獲獎生已領月份統計表"
WORKBOOK_TITLE = "國科會博士生獎學金-獲獎生已領月份統計表"

# 國科會's column order. The parser maps by header text, not position, so this
# order is presentational — but matching the real form keeps it familiar.
TEMPLATE_COLUMNS: List[str] = [
    "NO",
    "學院",
    "系所",
    HEADER_STUDENT_NUMBER,
    "學生姓名",
    HEADER_START_MONTH,
    HEADER_CURRENT_MONTH,
    "領獎結束月份",
    HEADER_STATED_TOTAL,
    "休學/退學/畢業",
    "備註",
]

# 國科會's own demonstration row. The parser skips any row carrying 範例, so this
# shows the expected formats without ever being imported.
EXAMPLE_ROW: List[object] = [
    EXAMPLE_ROW_MARKER,
    "人文藝術與社會學院",
    "應用藝術研究所",
    "412262001",
    "王美美",
    "113年9月",
    "115年8月",
    "116年8月",
    24,
    "115年9月休學",
    "",
]

BLANK_ROW_COUNT = 10

# Roughly the widths of the real form; Chinese glyphs need ~2 columns each.
_COLUMN_WIDTHS = [6, 22, 22, 14, 12, 16, 16, 16, 20, 18, 16]


def build_received_months_template() -> bytes:
    """Return the example workbook as .xlsx bytes."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    last_column = get_column_letter(len(TEMPLATE_COLUMNS))

    sheet.append([WORKBOOK_TITLE])
    sheet.merge_cells(f"A1:{last_column}1")
    title_cell = sheet["A1"]
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.append(TEMPLATE_COLUMNS)
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.append(EXAMPLE_ROW)

    # Pre-numbered empty rows, exactly as 國科會 ships the form.
    for number in range(1, BLANK_ROW_COUNT + 1):
        sheet.append([number] + [""] * (len(TEMPLATE_COLUMNS) - 1))

    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    # Keep the title and header visible while scrolling a long roster.
    sheet.freeze_panes = "A3"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
