"""Supplementary import service — adds new students to an existing ranking post-distribution."""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Column indices (1-based, matching 學生資料彙整表 export format)
_COL_RANK = 2
_COL_SCHOLARSHIP_TYPE = 3
_COL_STUDENT_ID = 13
_COL_BANK_ACCOUNT = 15
_COL_ADVISOR_NAME = 18
_STATIC_COL_COUNT = 18


@dataclass
class SupplementaryRow:
    """Parsed data for one student from the supplementary import Excel."""

    student_id: str
    excel_rank: int  # Value from col 2 (will be offset by max existing rank)
    sub_type_preferences: List[str]
    bank_account: Optional[str]
    advisor_name: Optional[str]
    submitted_form_fields: Dict[str, str]  # field_name -> raw cell value


def parse_scholarship_type_cell(cell_value: str, label_to_code: Dict[str, str]) -> List[str]:
    """Parse 申請獎學金類別 cell into ordered sub_type_preference codes.

    Formats:
        "XXX"                                      -> [code_of_XXX]
        "XXX(第一志願)暨YYY(第二志願)"              -> [code_of_XXX, code_of_YYY]
    """
    cell_value = (cell_value or "").strip()
    dual_match = re.fullmatch(r"(.+?)\(第一志願\)暨(.+?)\(第二志願\)", cell_value)
    if dual_match:
        first_label = dual_match.group(1).strip()
        second_label = dual_match.group(2).strip()
        for label in (first_label, second_label):
            if label not in label_to_code:
                raise ValueError(f"無法識別的獎學金類別：「{label}」")
        return [label_to_code[first_label], label_to_code[second_label]]

    if cell_value in label_to_code:
        return [label_to_code[cell_value]]

    raise ValueError(f"無法識別的獎學金類別：「{cell_value}」")


class SupplementaryImportService:
    """Handles all logic for supplementary student import after distribution."""

    def __init__(self, db, student_service=None):
        from app.services.student_service import StudentService

        self.db = db
        self.student_service = student_service or StudentService()

    # -------- Pure helpers (no DB / no HTTP) --------

    @staticmethod
    def parse_excel(
        file_bytes: bytes,
        label_to_code: Dict[str, str],
        dynamic_field_names: List[str],
    ) -> Tuple[List[SupplementaryRow], List[str]]:
        """Parse a 學生資料彙整表 Excel file.

        Returns (rows, errors). If errors is non-empty the caller should
        abort and return them to the client; rows may be partially populated.
        """
        errors: List[str] = []
        rows: List[SupplementaryRow] = []
        seen_student_ids: Dict[str, int] = {}  # student_id -> first excel row number
        seen_ranks: Dict[int, int] = {}  # rank -> first excel row number

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            ws = wb.active
        except Exception as exc:
            return [], [f"無法讀取 Excel 檔案：{exc}"]

        # Row 1 = title, Row 2 = headers, Row 3+ = data
        excel_row_num = 2  # last header row
        for excel_row in ws.iter_rows(min_row=3, values_only=True):
            excel_row_num += 1
            student_id_raw = excel_row[_COL_STUDENT_ID - 1] if len(excel_row) >= _COL_STUDENT_ID else None
            if not student_id_raw:
                continue  # skip empty rows

            student_id = str(student_id_raw).strip()

            # Duplicate student ID check
            if student_id in seen_student_ids:
                errors.append(f"學號重複：{student_id}（首次出現在第 {seen_student_ids[student_id]} 行）")
                continue
            seen_student_ids[student_id] = excel_row_num

            # Parse rank (col 2)
            rank_raw = excel_row[_COL_RANK - 1] if len(excel_row) >= _COL_RANK else None
            try:
                excel_rank = int(rank_raw)
                if excel_rank < 1:
                    raise ValueError()
            except (TypeError, ValueError):
                errors.append(f"排名無效（學號 {student_id}）：必須為正整數，收到 '{rank_raw}'")
                continue

            if excel_rank in seen_ranks:
                errors.append(f"排名重複：第 {excel_rank} 名出現超過一次（學號 {student_id}）")
                continue
            seen_ranks[excel_rank] = excel_row_num

            # Parse 申請獎學金類別 (col 3)
            scholarship_cell_raw = (
                excel_row[_COL_SCHOLARSHIP_TYPE - 1] if len(excel_row) >= _COL_SCHOLARSHIP_TYPE else None
            )
            scholarship_cell = str(scholarship_cell_raw or "").strip()
            try:
                sub_type_preferences = parse_scholarship_type_cell(scholarship_cell, label_to_code)
            except ValueError as exc:
                errors.append(f"學號 {student_id}：{exc}")
                continue

            # Other static columns
            bank_account_raw = excel_row[_COL_BANK_ACCOUNT - 1] if len(excel_row) >= _COL_BANK_ACCOUNT else None
            bank_account = str(bank_account_raw).strip() if bank_account_raw else None

            advisor_raw = excel_row[_COL_ADVISOR_NAME - 1] if len(excel_row) >= _COL_ADVISOR_NAME else None
            advisor_name = str(advisor_raw).strip() if advisor_raw else None

            # Dynamic columns (col 19+)
            submitted_form_fields: Dict[str, str] = {}
            for idx, field_name in enumerate(dynamic_field_names):
                col_idx = _STATIC_COL_COUNT + idx  # 0-based
                if col_idx < len(excel_row):
                    raw = excel_row[col_idx]
                    if raw is not None and str(raw).strip():
                        submitted_form_fields[field_name] = str(raw).strip()

            rows.append(
                SupplementaryRow(
                    student_id=student_id,
                    excel_rank=excel_rank,
                    sub_type_preferences=sub_type_preferences,
                    bank_account=bank_account,
                    advisor_name=advisor_name,
                    submitted_form_fields=submitted_form_fields,
                )
            )

        # Validate rank sequence is consecutive starting from 1
        if rows and not errors:
            expected = set(range(1, len(rows) + 1))
            actual = {r.excel_rank for r in rows}
            missing = expected - actual
            if missing:
                errors.append(f"排名不連續：缺少第 {', '.join(str(r) for r in sorted(missing))} 名")

        return rows, errors

    # -------- DB + SIS helpers --------

    async def validate_no_duplicate_applications(
        self,
        rows: List["SupplementaryRow"],
        scholarship_type_id: int,
        academic_year: int,
        semester: Optional[str],
    ) -> List[str]:
        """Return list of student_ids that already have an application for this scholarship/year/semester."""
        from sqlalchemy import select, and_, or_
        from app.models.application import Application
        from app.models.user import User

        student_ids = [r.student_id for r in rows]
        if not student_ids:
            return []

        user_stmt = select(User.id, User.nycu_id).where(User.nycu_id.in_(student_ids))
        user_result = await self.db.execute(user_stmt)
        nycu_to_user_id = {nycu_id: uid for uid, nycu_id in user_result.all()}

        if not nycu_to_user_id:
            return []

        user_ids = list(nycu_to_user_id.values())

        if semester == "yearly":
            sem_cond = or_(
                Application.semester.is_(None),
                Application.semester == "yearly",
            )
        else:
            sem_cond = Application.semester == semester

        app_stmt = select(Application.user_id).where(
            and_(
                Application.user_id.in_(user_ids),
                Application.scholarship_type_id == scholarship_type_id,
                Application.academic_year == academic_year,
                sem_cond,
                Application.deleted_at.is_(None),
            )
        )
        result = await self.db.execute(app_stmt)
        conflicting_user_ids = {row[0] for row in result.all()}

        user_id_to_nycu = {v: k for k, v in nycu_to_user_id.items()}
        return [user_id_to_nycu[uid] for uid in conflicting_user_ids if uid in user_id_to_nycu]

    async def fetch_student_data_bulk(
        self,
        student_ids: List[str],
        academic_year: int,
        semester: Optional[str],
    ) -> Tuple[Dict[str, dict], List[str]]:
        """Fetch student_data from SIS API for each student_id.

        Returns (data_map, missing_ids).
        Raises ValueError if SIS API is not enabled.
        Uses get_student_snapshot to capture API 1 + API 2 + metadata.
        """
        from app.core.exceptions import NotFoundError

        if not getattr(self.student_service, "api_enabled", False):
            raise ValueError("學生 API 未啟用，無法驗證學生資料")

        data_map: Dict[str, dict] = {}
        missing: List[str] = []

        for student_id in student_ids:
            try:
                data = await self.student_service.get_student_snapshot(
                    student_id,
                    academic_year=str(academic_year),
                    semester=semester,
                )
            except NotFoundError:
                missing.append(student_id)
                continue
            except Exception as exc:
                logger.warning("SIS API error for %s: %s", student_id, exc, exc_info=True)
                missing.append(student_id)
                continue

            data_map[student_id] = data

        return data_map, missing

    async def find_or_create_users(self, student_data_map: Dict[str, dict]) -> Dict[str, object]:
        """Return {student_id: User} — creates User if not found."""
        from sqlalchemy import select
        from app.models.user import User, UserRole, UserType

        student_ids = list(student_data_map.keys())
        if not student_ids:
            return {}

        stmt = select(User).where(User.nycu_id.in_(student_ids))
        result = await self.db.execute(stmt)
        user_map: Dict[str, "User"] = {u.nycu_id: u for u in result.scalars().all()}

        for student_id, sis_data in student_data_map.items():
            if student_id in user_map:
                continue
            new_user = User(
                nycu_id=student_id,
                name=sis_data.get("std_cname") or student_id,
                email=sis_data.get("com_email"),
                user_type=UserType.student,
                role=UserRole.student,
                dept_code=sis_data.get("std_depno"),
            )
            self.db.add(new_user)
            await self.db.flush()
            user_map[student_id] = new_user

        return user_map

    async def upsert_user_profiles(
        self,
        user_map: Dict[str, object],
        rows: List["SupplementaryRow"],
    ) -> None:
        """Create or update UserProfile with bank_account and advisor_name from Excel."""
        from sqlalchemy import select
        from app.models.user_profile import UserProfile

        user_ids = [u.id for u in user_map.values()]
        if not user_ids:
            return

        existing_stmt = select(UserProfile).where(UserProfile.user_id.in_(user_ids))
        existing_result = await self.db.execute(existing_stmt)
        profile_map: Dict[int, UserProfile] = {p.user_id: p for p in existing_result.scalars().all()}

        row_map = {r.student_id: r for r in rows}

        for student_id, user in user_map.items():
            row = row_map.get(student_id)
            if not row:
                continue
            if user.id in profile_map:
                profile = profile_map[user.id]
                if row.bank_account:
                    profile.account_number = row.bank_account
                if row.advisor_name:
                    profile.advisor_name = row.advisor_name
            else:
                profile = UserProfile(
                    user_id=user.id,
                    account_number=row.bank_account,
                    advisor_name=row.advisor_name,
                )
                self.db.add(profile)

        await self.db.flush()

    async def create_applications_and_items(
        self,
        rows: List["SupplementaryRow"],
        user_map: Dict[str, object],
        student_data_map: Dict[str, dict],
        ranking,  # CollegeRanking ORM object
        max_existing_rank: int,
    ) -> int:
        """Create Application + CollegeRankingItem for each supplementary row.

        Returns count of created items.
        """
        from sqlalchemy import select
        from app.models.application import Application, ApplicationStatus
        from app.models.application_sequence import ApplicationSequence
        from app.models.college_review import CollegeRankingItem

        if not rows:
            return 0

        semester_str = ranking.semester if ranking.semester else "yearly"

        # Lock sequence once, reserve len(rows) slots
        seq_stmt = (
            select(ApplicationSequence)
            .where(
                ApplicationSequence.academic_year == ranking.academic_year,
                ApplicationSequence.semester == semester_str,
            )
            .with_for_update()
        )
        seq_result = await self.db.execute(seq_stmt)
        seq_record = seq_result.scalar_one_or_none()
        if not seq_record:
            seq_record = ApplicationSequence(
                academic_year=ranking.academic_year,
                semester=semester_str,
                last_sequence=0,
            )
            self.db.add(seq_record)
            await self.db.flush()

        base_seq = seq_record.last_sequence
        seq_record.last_sequence = base_seq + len(rows)
        await self.db.flush()

        created = 0
        for idx, row in enumerate(rows):
            user = user_map.get(row.student_id)
            if not user:
                continue

            sis_data = student_data_map.get(row.student_id, {})

            app_id = ApplicationSequence.format_app_id(ranking.academic_year, semester_str, base_seq + idx + 1)

            submitted_form_data = {
                "fields": {
                    field_name: {
                        "field_id": field_name,
                        "field_type": "text",
                        "value": value,
                    }
                    for field_name, value in row.submitted_form_fields.items()
                }
            }

            # scholarship_subtype_list is what the manual-distribution panel reads
            # as `applied_sub_types`; sub_type_preferences is the ordered preference list
            # used by allocation logic. Set both from the Excel 申請獎學金類別 column so
            # admin can see + distribute supplementary students.
            app = Application(
                app_id=app_id,
                user_id=user.id,
                scholarship_type_id=ranking.scholarship_type_id,
                academic_year=ranking.academic_year,
                semester=ranking.semester,
                status=ApplicationStatus.submitted,
                sub_type_selection_mode=ranking.scholarship_type.sub_type_selection_mode,
                student_data=sis_data,
                scholarship_subtype_list=list(row.sub_type_preferences or []),
                sub_type_preferences=row.sub_type_preferences,
                submitted_form_data=submitted_form_data,
            )
            self.db.add(app)
            await self.db.flush()

            rank_item = CollegeRankingItem(
                ranking_id=ranking.id,
                application_id=app.id,
                rank_position=max_existing_rank + row.excel_rank,
                is_supplementary=True,
                status="ranked",
                college_rejected=False,
                is_allocated=False,
            )
            self.db.add(rank_item)
            created += 1

        await self.db.flush()
        return created
